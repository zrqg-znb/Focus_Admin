import os
from datetime import datetime

import openpyxl
from openpyxl.workbook import Workbook

from application.settings import STATIC_URL


def excel_to_dict(file_path, label_field_dict):
    # 加载工作簿
    workbook = openpyxl.load_workbook(filename=file_path)
    sheet = workbook.active  # 获取活动表单

    # 获取第一行作为键
    keys = [label_field_dict.get(cell.value) for cell in sheet[1]]

    # 遍历其他行作为值
    data_list = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        data_list.append(dict(zip(keys, row)))

    return data_list


def dict_to_excel(list_data: list[dict]):
    # 创建并初始化Excel工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    
    # 生成唯一的文件名并设置文件路径
    file_name = datetime.now().strftime('%Y%m%d%H%M%S%f') + '.xlsx'
    current_ymd = datetime.now().strftime('%Y%m%d')
    file_path = os.path.join(STATIC_URL, current_ymd)
    file_url = os.path.join(file_path, file_name)
    # 确保导出文件的目录存在
    if not os.path.exists(file_path):
        os.makedirs(file_path)
    
    # 如果没有数据，创建空工作簿
    if not list_data:
        wb.save(file_url)
        wb.close()
        return file_url
    
    # 写入表头（只写入一次）
    headers = list(list_data[0].keys())
    ws.append(headers)
    
    # 向Excel中写入数据行
    for data in list_data:
        row = []
        for header in headers:
            value = data.get(header, '')
            # 处理列表类型的值
            if isinstance(value, list):
                value = ','.join(str(v) for v in value) if value else ''
            elif value is None:
                value = ''
            row.append(value)
        ws.append(row)

    # 保存Excel文件到指定路径
    wb.save(file_url)
    wb.close()
    return file_url
