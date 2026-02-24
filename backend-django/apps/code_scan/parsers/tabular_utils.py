import csv
import os
import re
from typing import Any, Dict, Iterable, List, Optional, Tuple

from openpyxl import load_workbook


def normalize_header(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower()
    text = text.replace("（", "(").replace("）", ")")
    text = re.sub(r"[\s_\-]+", "", text)
    text = re.sub(r"[^0-9a-zA-Z\u4e00-\u9fff()]+", "", text)
    return text


def as_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def as_int(value: Any, default: int = 0) -> int:
    if value in (None, ""):
        return default
    try:
        if isinstance(value, str):
            value = value.strip()
            if not value:
                return default
        return int(float(value))
    except Exception:
        return default


def join_path(path_value: Any, file_value: Any) -> str:
    path_text = as_text(path_value)
    file_text = as_text(file_value)
    if path_text and file_text:
        normalized_path = path_text.replace("\\", "/")
        normalized_file = file_text.replace("\\", "/")
        if normalized_path.endswith(normalized_file):
            return normalized_path
        return f"{normalized_path.rstrip('/')}/{normalized_file.lstrip('/')}"
    return path_text or file_text


def load_rows_by_extension(file_path: str) -> Tuple[List[Any], List[List[Any]]]:
    lower = file_path.lower()
    if lower.endswith(".xlsx"):
        return _load_xlsx_rows(file_path)
    if lower.endswith(".csv"):
        return _load_csv_rows(file_path)
    raise ValueError("不支持的文件格式，请上传 .xlsx 或 .csv 文件")


def build_header_index(header_row: Iterable[Any]) -> Dict[str, int]:
    header_index: Dict[str, int] = {}
    for index, cell in enumerate(header_row):
        key = normalize_header(cell)
        if key and key not in header_index:
            header_index[key] = index
    return header_index


def find_col_index(header_index: Dict[str, int], aliases: Iterable[str]) -> Optional[int]:
    alias_keys = [normalize_header(alias) for alias in aliases if normalize_header(alias)]
    if not alias_keys:
        return None

    for alias in alias_keys:
        if alias in header_index:
            return header_index[alias]

    for alias in alias_keys:
        for header_key, idx in header_index.items():
            if alias and alias in header_key:
                return idx

    return None


def value_by_index(row: List[Any], index: Optional[int]) -> Any:
    if index is None:
        return None
    if index < 0 or index >= len(row):
        return None
    return row[index]


def _load_xlsx_rows(file_path: str) -> Tuple[List[Any], List[List[Any]]]:
    workbook = load_workbook(file_path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        iterator = sheet.iter_rows(values_only=True)
        header: Optional[Tuple[Any, ...]] = None
        for row in iterator:
            if _has_content(row):
                header = row
                break
        if header is None:
            return [], []

        rows: List[List[Any]] = []
        for row in iterator:
            if not _has_content(row):
                continue
            rows.append(list(row))
        return list(header), rows
    finally:
        workbook.close()


def _load_csv_rows(file_path: str) -> Tuple[List[Any], List[List[Any]]]:
    last_error: Optional[Exception] = None
    for encoding in ("utf-8-sig", "utf-8", "gbk", "latin-1"):
        try:
            with open(file_path, "r", encoding=encoding, newline="") as f:
                sample = f.read(4096)
                f.seek(0)
                try:
                    dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
                except Exception:
                    dialect = csv.excel
                reader = csv.reader(f, dialect)
                rows = [row for row in reader if _has_content(row)]
                if not rows:
                    return [], []
                header = rows[0]
                return header, rows[1:]
        except Exception as exc:
            last_error = exc
            continue
    raise ValueError(f"无法读取 CSV 文件: {os.path.basename(file_path)}") from last_error


def _has_content(row: Optional[Iterable[Any]]) -> bool:
    if row is None:
        return False
    for cell in row:
        if as_text(cell):
            return True
    return False
