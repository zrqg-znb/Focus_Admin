import datetime
import hashlib
import logging
import json
import math
import os
import random
import re
import time
from typing import Any

import openpyxl
import requests
from django.conf import settings
from django.core.cache import cache
from django.http import HttpResponse
from django.utils import timezone
from ninja.errors import HttpError

from apps.project_manager.project.project_model import Project

from .requirement_board_model import (
    CATEGORY_ORDER,
    STATUS_LABELS,
    STATUS_ORDER,
    TIME_FIELD_OPTIONS,
    UNKNOWN_TEAM_NAME,
    VERIFICATION_POLICY_LABELS,
    VERIFICATION_POLICY_ORDER,
)
from .requirement_board_schemas import (
    RequirementBoardDataQuerySchema,
    RequirementBoardExportQuerySchema,
    RequirementBoardSummaryQuerySchema,
)

logger = logging.getLogger(__name__)

_DATA_CACHE_TTL_SECONDS = 10 * 60
_SUMMARY_CACHE_TTL_SECONDS = 10 * 60
_FILTERED_SCAN_CACHE_TTL_SECONDS = 5 * 60
_LOCK_TTL_SECONDS = 30
_UPSTREAM_PAGE_SIZE = 500
_MAX_SCAN_PAGES = 200
_DELAY_PREVIEW_LIMIT = 8
_EXPORT_SHEET_TITLE = "需求数据"
_EXPORT_HEADERS = (
    "项目名",
    "团队",
    "需求类型",
    "验证策略",
    "需求 ID",
    "标题",
    "状态代码",
    "状态名称",
    "计划转测时间",
    "计划完成时间",
    "开发完成时间",
    "测试完成时间",
    "开发延期",
    "测试延期",
    "工作量(人天)",
    "代码量(KLOC)",
    "开发责任人",
    "测试责任人",
)
_STATUS_FIELD_MAP = {
    "I": "i_count",
    "D": "d_count",
    "P": "p_count",
    "C": "c_count",
    "A": "a_count",
}
_STATUS_ALIASES = {
    "I": "I",
    "INIT": "I",
    "INITIAL": "I",
    "INITIALIZED": "I",
    "INITIALIZATION": "I",
    "初始化": "I",
    "D": "D",
    "DEFINE": "D",
    "DEFINED": "D",
    "DEFINITION": "D",
    "已完成定义": "D",
    "已定义完成": "D",
    "P": "P",
    "PROGRESS": "P",
    "INPROGRESS": "P",
    "IN_PROGRESS": "P",
    "IN-PROGRESS": "P",
    "IN PROGRESS": "P",
    "WORKING": "P",
    "开发中": "P",
    "正在工作": "P",
    "C": "C",
    "COMPLETE": "C",
    "COMPLETED": "C",
    "CODECOMPLETE": "C",
    "CODE_COMPLETE": "C",
    "已开发完成": "C",
    "开发已完成": "C",
    "已开发完成（转测）": "C",
    "A": "A",
    "ACCEPT": "A",
    "ACCEPTED": "A",
    "ACCEPTANCE": "A",
    "测试完成": "A",
    "测试完成（已置A）": "A",
    "测试验收完成": "A",
}
_OWNER_SPLIT_RE = re.compile(r"[,，]")
_OWNER_PAREN_RE = re.compile(r"[（(](.*?)[)）]")


def _get_setting(name: str, default: Any = None):
    return getattr(settings, name, os.environ.get(name, default))


def _serialize_log_payload(payload: Any) -> str:
    try:
        return json.dumps(payload, ensure_ascii=False, sort_keys=True, default=str)
    except TypeError:
        return str(payload)


def _is_debug_log_enabled() -> bool:
    return _to_bool(_get_setting("REQUIREMENT_BOARD_DEBUG_LOG", False), False)


def _debug_log(event: str, **payload: Any) -> None:
    if not _is_debug_log_enabled():
        return
    if payload:
        logger.info(
            "RequirementBoard[%s] %s",
            event,
            _serialize_log_payload(payload),
        )
        return
    logger.info("RequirementBoard[%s]", event)


def _to_bool(value: Any, default: bool = False) -> bool:
    if value is None:
        return default
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return bool(value)
    text = str(value).strip().lower()
    if not text:
        return default
    return text in {"1", "true", "yes", "y", "on"}


def _clean_text(value: Any) -> str:
    return str(value or "").strip()


def _normalize_text_list(values: Any) -> list[str]:
    if values is None:
        return []
    raw_values = values if isinstance(values, list) else [values]
    result: list[str] = []
    seen: set[str] = set()
    for item in raw_values:
        text = _clean_text(item)
        if not text or text in seen:
            continue
        seen.add(text)
        result.append(text)
    return result


def _strip_owner_suffix(value: str) -> str:
    if not value:
        return ""
    return re.sub(r"[（(].*?[)）]$", "", value).strip()


def _extract_owner_username(value: str) -> str:
    text = value.strip()
    if not text:
        return ""
    if any(char.isspace() for char in text):
        parts = [item for item in re.split(r"\s+", text) if item]
        return _strip_owner_suffix(parts[-1]) if parts else ""
    if "(" in text or "（" in text:
        match = _OWNER_PAREN_RE.search(text)
        if match:
            candidate = match.group(1).strip()
            if candidate:
                return candidate
    return _strip_owner_suffix(text)


def _normalize_owner_list(values: Any) -> list[str]:
    if values is None:
        return []
    raw_values = values if isinstance(values, list) else [values]
    result: list[str] = []
    seen: set[str] = set()
    for item in raw_values:
        text = _clean_text(item)
        if not text:
            continue
        for part in _OWNER_SPLIT_RE.split(text):
            username = _extract_owner_username(part)
            if not username or username in seen:
                continue
            seen.add(username)
            result.append(username)
    return result


def _normalize_categories(values: Any) -> list[str]:
    normalized = _normalize_text_list(values)
    if not normalized:
        return list(CATEGORY_ORDER)

    result: list[str] = []
    for item in normalized:
        category = item.upper()
        if category not in CATEGORY_ORDER:
            raise HttpError(422, f"非法需求类型: {item}")
        if category not in result:
            result.append(category)
    return result


def _normalize_verification_policies(values: Any) -> list[str]:
    normalized = _normalize_text_list(values)
    if not normalized:
        return []

    result: list[str] = []
    for item in normalized:
        policy = str(item).strip()
        if policy not in VERIFICATION_POLICY_LABELS:
            raise HttpError(422, f"非法验证策略: {item}")
        if policy not in result:
            result.append(policy)
    return result


def _normalize_time_field(value: Any) -> str | None:
    text = _clean_text(value)
    if not text:
        return None
    if text not in TIME_FIELD_OPTIONS:
        raise HttpError(422, f"非法时间维度: {text}")
    return text


def _to_float(value: Any) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = _clean_text(value).replace(",", "")
    if not text:
        return 0.0
    try:
        return float(text)
    except Exception:
        return 0.0


def _parse_positive_int(value: Any, default: int) -> int:
    try:
        parsed = int(value)
    except Exception:
        return default
    return parsed if parsed > 0 else default


def _round_metric(value: Any) -> float:
    return round(float(value or 0.0), 2)


def _ensure_aware(value: datetime.datetime) -> datetime.datetime:
    if timezone.is_aware(value):
        return value
    return timezone.make_aware(value, timezone.get_current_timezone())


def _looks_like_date_only(text: str) -> bool:
    return len(text) == 10 and text.count("-") == 2 and " " not in text and "T" not in text


def _normalize_datetime_pair(
    left: datetime.datetime | None,
    right: datetime.datetime | None,
) -> tuple[datetime.datetime | None, datetime.datetime | None]:
    if left is None or right is None:
        return left, right
    left_is_aware = timezone.is_aware(left)
    right_is_aware = timezone.is_aware(right)
    if left_is_aware == right_is_aware:
        return left, right

    current_tz = timezone.get_current_timezone()
    if left_is_aware or right_is_aware:
        if not left_is_aware:
            left = timezone.make_aware(left, current_tz)
        if not right_is_aware:
            right = timezone.make_aware(right, current_tz)
        return left, right

    return left, right


def _parse_datetime(value: Any) -> datetime.datetime | None:
    if value is None:
        return None
    if isinstance(value, datetime.datetime):
        return _ensure_aware(value)
    if isinstance(value, datetime.date):
        return _ensure_aware(
            datetime.datetime.combine(value, datetime.time.min),
        )

    text = _clean_text(value)
    if not text:
        return None

    iso_text = text.replace("Z", "+00:00")
    try:
        return _ensure_aware(datetime.datetime.fromisoformat(iso_text))
    except ValueError:
        pass

    for fmt in (
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y-%m-%dT%H:%M:%S",
        "%Y-%m-%dT%H:%M",
        "%Y-%m-%d",
    ):
        try:
            parsed = datetime.datetime.strptime(text, fmt)
            return _ensure_aware(parsed)
        except ValueError:
            continue
    return None


def _parse_range_boundary(value: Any, *, is_end: bool) -> datetime.datetime | None:
    text = _clean_text(value)
    if not text:
        return None
    if _looks_like_date_only(text):
        day = datetime.date.fromisoformat(text)
        point = datetime.time(23, 59, 59) if is_end else datetime.time.min
        return _ensure_aware(datetime.datetime.combine(day, point))
    parsed = _parse_datetime(text)
    if parsed is None:
        raise HttpError(422, f"非法时间格式: {text}")
    return parsed


def _format_datetime_text(value: Any) -> str | None:
    parsed = _parse_datetime(value)
    if parsed is not None:
        localized = timezone.localtime(_ensure_aware(parsed))
        return localized.strftime("%Y-%m-%d %H:%M:%S")
    text = _clean_text(value)
    return text or None


def _to_local_date(value: Any) -> datetime.date | None:
    parsed = _parse_datetime(value)
    if parsed is None:
        return None
    return timezone.localtime(_ensure_aware(parsed)).date()


def _normalize_status(raw_status: Any) -> tuple[str, str]:
    text = _clean_text(raw_status)
    compact = text.upper().replace("-", "").replace("_", "").replace(" ", "")
    normalized = (
        _STATUS_ALIASES.get(text.upper())
        or _STATUS_ALIASES.get(compact)
        or _STATUS_ALIASES.get(text)
    )
    if not normalized:
        normalized = "P"
    return normalized, STATUS_LABELS[normalized]


def _is_development_delayed(
    status_code: str,
    planned_test_time: Any,
    completed_time: Any,
    now: datetime.datetime,
) -> bool:
    planned_date = _to_local_date(planned_test_time)
    if planned_date is None:
        return False
    completed_date = _to_local_date(completed_time)
    if completed_date is not None and status_code in {"C", "A"}:
        return completed_date > planned_date
    today = timezone.localtime(_ensure_aware(now)).date()
    return status_code not in {"C", "A"} and today > planned_date


def _is_acceptance_delayed(
    status_code: str,
    due_date: Any,
    accepted_time: Any,
    now: datetime.datetime,
) -> bool:
    due_day = _to_local_date(due_date)
    if due_day is None:
        return False
    accepted_day = _to_local_date(accepted_time)
    if accepted_day is not None and status_code == "A":
        return accepted_day > due_day
    today = timezone.localtime(_ensure_aware(now)).date()
    return status_code != "A" and today > due_day


def _cache_key(prefix: str, payload: dict[str, Any]) -> str:
    digest = hashlib.md5(
        json.dumps(payload, ensure_ascii=False, sort_keys=True).encode("utf-8")
    ).hexdigest()
    return f"{prefix}:{digest}"


def _build_request_payload(
    design_ids: list[str],
    sub_teams: list[str],
    categories: list[str],
    verification_policies: list[str],
    page_no: int,
    page_size: int,
) -> dict[str, Any]:
    domain_field = _clean_text(_get_setting("REQUIREMENT_BOARD_DOMAIN_FIELD", "domainid"))
    payload = {
        domain_field or "domainid": design_ids,
        "sub_teams": sub_teams,
        "categories": categories,
        "page": {
            "page_no": page_no,
            "page_size": min(page_size, _UPSTREAM_PAGE_SIZE),
        },
    }
    if verification_policies:
        payload["verification_policy"] = verification_policies
    alias_field = _clean_text(
        _get_setting("REQUIREMENT_BOARD_DESIGN_ALIAS_FIELD", ""),
    )
    if alias_field and alias_field not in payload:
        payload[alias_field] = design_ids
    return payload


def _build_request_headers() -> dict[str, str]:
    headers = {"Content-Type": "application/json"}
    token = _clean_text(_get_setting("REQUIREMENT_BOARD_API_TOKEN", ""))
    if token:
        headers["Authorization"] = (
            token if token.lower().startswith("bearer ") else f"Bearer {token}"
        )

    raw_extra_headers = _clean_text(
        _get_setting("REQUIREMENT_BOARD_API_HEADERS_JSON", ""),
    )
    if raw_extra_headers:
        try:
            extra_headers = json.loads(raw_extra_headers)
        except json.JSONDecodeError as exc:
            raise HttpError(500, f"需求看板请求头配置非法: {exc}") from exc
        if not isinstance(extra_headers, dict):
            raise HttpError(500, "需求看板请求头配置必须是 JSON 对象")
        for key, value in extra_headers.items():
            headers[str(key)] = str(value)
    return headers


def _mock_fetch_page(
    design_ids: list[str],
    sub_teams: list[str],
    categories: list[str],
    verification_policies: list[str],
    page_no: int,
    page_size: int,
) -> dict[str, Any]:
    seed = json.dumps(
        {
            "design_ids": design_ids,
            "sub_teams": sub_teams,
            "categories": categories,
            "verification_policies": verification_policies,
        },
        ensure_ascii=False,
        sort_keys=True,
    )
    rng = random.Random(seed)
    team_pool = sub_teams or [UNKNOWN_TEAM_NAME]
    owner_pool = [f"z6009{index:04d}" for index in range(1, 80)]
    status_pool = ["Initial", "Defined", "In-Progress", "Completed", "Accepted"]
    status_weights = (0.08, 0.18, 0.36, 0.2, 0.18)
    policy_pool = verification_policies or list(VERIFICATION_POLICY_ORDER)
    all_items: list[dict[str, Any]] = []

    for design_id in design_ids:
        requirement_total = rng.randint(42, 88)
        for index in range(requirement_total):
            category = rng.choice(categories or list(CATEGORY_ORDER))
            team_name = rng.choice(team_pool)
            schedule_state = rng.choices(status_pool, weights=status_weights, k=1)[0]
            verification_policy = rng.choice(policy_pool)
            month = (index % 12) + 1
            day = min((index % 25) + 1, 28)
            planned_dt = datetime.datetime(2026, month, day, 18, 0, 0)
            due_dt = planned_dt + datetime.timedelta(days=7)
            completed_dt = None
            accepted_dt = None
            if schedule_state in {"Completed", "Accepted"}:
                completed_dt = planned_dt + datetime.timedelta(days=rng.randint(-2, 8))
            if schedule_state == "Accepted":
                accepted_dt = due_dt + datetime.timedelta(days=rng.randint(-1, 9))
                if completed_dt and accepted_dt < completed_dt:
                    accepted_dt = completed_dt + datetime.timedelta(days=1)

            develop_owners = ",".join(
                sorted(rng.sample(owner_pool, k=rng.randint(1, 3))),
            )
            test_owners = ",".join(
                sorted(rng.sample(owner_pool, k=rng.randint(1, 2))),
            )
            all_items.append(
                {
                    "category": category,
                    "id": f"{design_id}-{category}-{index + 1:04d}",
                    "title": f"{category} 需求 {index + 1:04d}",
                    "schedule_state": schedule_state,
                    "verification_policy": verification_policy,
                    "requirement2domain": design_id,
                    "planned_test_time": planned_dt.strftime("%Y-%m-%d %H:%M:%S"),
                    "due_date": due_dt.strftime("%Y-%m-%d %H:%M:%S"),
                    "completed_time": (
                        completed_dt.strftime("%Y-%m-%d %H:%M:%S")
                        if completed_dt
                        else None
                    ),
                    "accepted_time": (
                        accepted_dt.strftime("%Y-%m-%d %H:%M:%S") if accepted_dt else None
                    ),
                    "workload_kloc": round(rng.uniform(0.3, 28.0), 2),
                    "workload_man_day": round(rng.uniform(1.0, 24.0), 2),
                    "service_name": team_name,
                    "develop_owner": develop_owners,
                    "test_owner": test_owners,
                }
            )

    start = max(page_no - 1, 0) * page_size
    end = start + page_size
    total = len(all_items)
    return {
        "code": 200,
        "data": {
            "result": all_items[start:end],
            "page": {
                "page_sum": total,
                "page_no": page_no,
                "page_size": page_size,
                "total": total,
            },
        },
        "message": "success",
    }


def _fetch_raw_page(
    design_ids: list[str],
    sub_teams: list[str],
    categories: list[str],
    verification_policies: list[str],
    page_no: int,
    page_size: int,
) -> dict[str, Any]:
    url = _clean_text(_get_setting("REQUIREMENT_BOARD_API_URL", ""))
    request_payload = _build_request_payload(
        design_ids,
        sub_teams,
        categories,
        verification_policies,
        page_no,
        page_size,
    )
    if not url:
        _debug_log(
            "upstream_request",
            mode="mock",
            method="POST",
            url="mock://requirement-board",
            payload=request_payload,
        )
        raw_payload = _mock_fetch_page(
            design_ids,
            sub_teams,
            categories,
            verification_policies,
            page_no,
            page_size,
        )
        raw_items, raw_page = _extract_raw_page(raw_payload)
        _debug_log(
            "upstream_response",
            mode="mock",
            page=raw_page,
            item_count=len(raw_items),
            code=raw_payload.get("code"),
            message=raw_payload.get("message"),
        )
        return raw_payload

    headers = _build_request_headers()
    timeout = float(_get_setting("REQUIREMENT_BOARD_API_TIMEOUT", 15))
    verify = _to_bool(_get_setting("REQUIREMENT_BOARD_API_VERIFY_SSL", True), True)
    _debug_log(
        "upstream_request",
        mode="http",
        method="POST",
        url=url,
        payload=request_payload,
    )

    try:
        response = requests.post(
            url,
            json=request_payload,
            headers=headers,
            timeout=timeout,
            verify=verify,
        )
    except requests.RequestException as exc:
        _debug_log(
            "upstream_request_error",
            mode="http",
            method="POST",
            url=url,
            error=str(exc),
        )
        raise HttpError(502, f"请求数据湖失败: {exc}") from exc

    if response.status_code != 200:
        _debug_log(
            "upstream_response_error",
            mode="http",
            url=url,
            status_code=response.status_code,
            body=response.text[:1000],
        )
        raise HttpError(502, f"数据湖返回异常状态码: {response.status_code}")

    try:
        response_payload = response.json()
    except ValueError as exc:
        _debug_log(
            "upstream_response_error",
            mode="http",
            url=url,
            status_code=response.status_code,
            error="invalid_json",
            body=response.text[:1000],
        )
        raise HttpError(502, "数据湖返回的不是合法 JSON") from exc

    raw_items, raw_page = _extract_raw_page(response_payload)
    _debug_log(
        "upstream_response",
        mode="http",
        url=url,
        status_code=response.status_code,
        code=response_payload.get("code"),
        message=response_payload.get("message"),
        page=raw_page,
        item_count=len(raw_items),
    )

    if int(response_payload.get("code") or 0) != 200:
        raise HttpError(502, response_payload.get("message") or "数据湖返回失败")
    return response_payload


def _extract_raw_page(raw_payload: dict[str, Any]) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    data = raw_payload.get("data") or {}
    result = data.get("result")
    if result is None:
        result = data.get("items") or []
    if not isinstance(result, list):
        raise HttpError(502, "数据湖返回的需求结果格式错误")
    page = data.get("page") or {}
    if not isinstance(page, dict):
        page = {}
    return result, page


def _resolve_total(
    page: dict[str, Any],
    page_no: int,
    page_size: int,
    item_count: int,
) -> int:
    for key in ("page_sum", "total", "total_count", "count", "row_sum", "record_sum"):
        parsed = _parse_positive_int(page.get(key), 0)
        if parsed > 0:
            return parsed

    return max((page_no - 1) * page_size + item_count, item_count)


def _resolve_page_sum(
    page: dict[str, Any],
    page_no: int,
    page_size: int,
    item_count: int,
) -> int:
    total = _resolve_total(page, page_no, page_size, item_count)
    if total <= 0:
        return 0
    return math.ceil(total / page_size) if page_size else 0


def _resolve_scan_page_size(setting_name: str, default: int = _UPSTREAM_PAGE_SIZE) -> int:
    return min(
        _parse_positive_int(_get_setting(setting_name, default), default),
        _UPSTREAM_PAGE_SIZE,
    )


def _resolve_max_scan_pages(setting_name: str) -> int:
    return max(
        _parse_positive_int(
            _get_setting(setting_name, _MAX_SCAN_PAGES),
            _MAX_SCAN_PAGES,
        ),
        1,
    )


def _should_stop_scan(
    page_payload: dict[str, Any],
    page_no: int,
    fetched_count: int,
) -> bool:
    items = page_payload.get("items") or []
    item_count = len(items)
    total = _parse_positive_int(page_payload.get("total"), 0)
    page_size = _parse_positive_int(page_payload.get("page_size"), 0)
    page_sum = _parse_positive_int(page_payload.get("page_sum"), 0)
    if total > 0 and fetched_count >= total:
        return True
    if item_count <= 0:
        return True
    if page_size > 0 and item_count < page_size:
        return True
    return page_sum > 0 and page_no >= page_sum


def _iterate_remote_pages(
    context: dict[str, Any],
    *,
    page_size: int,
    max_pages: int,
    limit_error_message: str,
):
    fetched_count = 0
    page_no = 1
    while True:
        if page_no > max_pages:
            raise HttpError(502, limit_error_message)
        page_payload = _load_remote_page(context, page_no=page_no, page_size=page_size)
        fetched_count += len(page_payload.get("items") or [])
        yield page_no, page_payload, fetched_count
        if _should_stop_scan(page_payload, page_no, fetched_count):
            break
        page_no += 1


def _project_payload(project: Project | None, design_id: str) -> tuple[str, str]:
    if project is None:
        return "", design_id or "未匹配项目"
    return str(project.id), project.name


def _standardize_requirement_items(
    items: list[dict[str, Any]],
    design_project_map: dict[str, Project],
) -> list[dict[str, Any]]:
    standardized: list[dict[str, Any]] = []
    fallback_project = next(iter(design_project_map.values()), None)
    now = timezone.now()

    for item in items:
        design_id = _clean_text(
            item.get("requirement2domain")
            or item.get("domainid")
            or item.get("design_id")
        )
        project = design_project_map.get(design_id)
        if project is None and len(design_project_map) == 1:
            project = fallback_project
        project_id, project_name = _project_payload(project, design_id)

        status_code, status_label = _normalize_status(item.get("schedule_state"))
        category = _clean_text(item.get("category")).upper()
        if category not in CATEGORY_ORDER:
            category = category or CATEGORY_ORDER[0]
        verification_policy = _clean_text(item.get("verification_policy"))
        verification_policy_label = VERIFICATION_POLICY_LABELS.get(
            verification_policy,
            verification_policy or "--",
        )

        team_name = _clean_text(item.get("service_name") or item.get("servioce_name"))
        if not team_name:
            team_name = UNKNOWN_TEAM_NAME

        requirement_id = _clean_text(item.get("id") or item.get("requirement_id"))
        if not requirement_id:
            requirement_id = f"{design_id or 'unknown'}-{category}"

        planned_test_time = _format_datetime_text(item.get("planned_test_time"))
        due_date = _format_datetime_text(item.get("due_date"))
        completed_time = _format_datetime_text(item.get("completed_time"))
        accepted_time = _format_datetime_text(item.get("accepted_time"))
        develop_users = _normalize_owner_list(
            item.get("develop_owner") or item.get("develop_user"),
        )
        test_users = _normalize_owner_list(
            item.get("test_owner") or item.get("test_user"),
        )
        develop_user_display = ", ".join(develop_users)
        test_user_display = ", ".join(test_users)

        standardized.append(
            {
                "requirement_id": requirement_id,
                "title": _clean_text(item.get("title")) or requirement_id,
                "category": category,
                "verification_policy": verification_policy,
                "verification_policy_label": verification_policy_label,
                "status_code": status_code,
                "status_label": status_label,
                "raw_status": _clean_text(item.get("schedule_state")),
                "project_id": project_id,
                "project_name": project_name,
                "design_id": design_id or None,
                "team_name": team_name,
                "planned_test_time": planned_test_time,
                "due_date": due_date,
                "completed_time": completed_time,
                "accepted_time": accepted_time,
                "is_dev_delayed": _is_development_delayed(
                    status_code,
                    planned_test_time,
                    completed_time,
                    now,
                ),
                "is_test_delayed": _is_acceptance_delayed(
                    status_code,
                    due_date,
                    accepted_time,
                    now,
                ),
                "workload_kloc": _round_metric(_to_float(item.get("workload_kloc"))),
                "workload_man_day": _round_metric(
                    _to_float(item.get("workload_man_day")),
                ),
                "develop_users": develop_users,
                "test_users": test_users,
                "develop_user_display": develop_user_display,
                "test_user_display": test_user_display,
                "develop_user": develop_user_display,
                "test_user": test_user_display,
            }
        )
    return standardized


def _normalize_time_filters(
    time_field: Any,
    time_start: Any,
    time_end: Any,
    accepted_time_start: Any,
    accepted_time_end: Any,
) -> dict[str, Any]:
    normalized_field = _normalize_time_field(time_field)
    normalized_start = _clean_text(time_start) or _clean_text(accepted_time_start)
    normalized_end = _clean_text(time_end) or _clean_text(accepted_time_end)
    if (normalized_start or normalized_end) and not normalized_field:
        normalized_field = "accepted_time"

    start_dt = (
        _parse_range_boundary(normalized_start, is_end=False)
        if normalized_start
        else None
    )
    end_dt = (
        _parse_range_boundary(normalized_end, is_end=True) if normalized_end else None
    )
    if start_dt and end_dt and start_dt > end_dt:
        raise HttpError(422, "时间区间开始不能晚于结束")

    return {
        "time_field": normalized_field,
        "time_start": normalized_start,
        "time_end": normalized_end,
        "time_start_dt": start_dt,
        "time_end_dt": end_dt,
    }


def _resolve_query_context(
    project_ids: list[str],
    sub_teams: list[str] | None = None,
    categories: list[str] | None = None,
    verification_policies: list[str] | None = None,
    develop_users: list[str] | None = None,
    test_users: list[str] | None = None,
    time_field: str | None = None,
    time_start: str | None = None,
    time_end: str | None = None,
    accepted_time_start: str | None = None,
    accepted_time_end: str | None = None,
) -> dict[str, Any]:
    ordered_project_ids = _normalize_text_list(project_ids)
    if not ordered_project_ids:
        raise HttpError(422, "请至少选择一个项目")

    project_qs = Project.objects.filter(id__in=ordered_project_ids, is_deleted=False)
    project_map = {str(item.id): item for item in project_qs}
    missing_ids = [item for item in ordered_project_ids if item not in project_map]
    if missing_ids:
        raise HttpError(422, f"部分项目不存在或已删除: {', '.join(missing_ids)}")

    ordered_projects = [project_map[item] for item in ordered_project_ids]
    design_project_map: dict[str, Project] = {}
    design_ids: list[str] = []
    allowed_teams: list[str] = []
    allowed_team_set: set[str] = set()
    invalid_projects: list[str] = []

    for project in ordered_projects:
        design_id = _clean_text(project.design_id)
        project_teams = _normalize_text_list(project.sub_teams)
        if not design_id or not project_teams:
            invalid_projects.append(project.name)
            continue

        conflict = design_project_map.get(design_id)
        if conflict and conflict.id != project.id:
            raise HttpError(
                422,
                f"项目 {conflict.name} 与 {project.name} 的 design_id 重复，无法匹配需求归属",
            )

        design_project_map[design_id] = project
        design_ids.append(design_id)
        for team in project_teams:
            if team in allowed_team_set:
                continue
            allowed_team_set.add(team)
            allowed_teams.append(team)

    if invalid_projects:
        raise HttpError(
            422,
            f"以下项目未完成需求数据源配置: {', '.join(invalid_projects)}",
        )

    selected_teams = _normalize_text_list(sub_teams)
    if selected_teams:
        invalid_teams = [team for team in selected_teams if team not in allowed_team_set]
        if invalid_teams:
            raise HttpError(422, f"存在无效责任团队: {', '.join(invalid_teams)}")
    else:
        selected_teams = allowed_teams[:]

    selected_categories = _normalize_categories(categories)
    selected_verification_policies = _normalize_verification_policies(
        verification_policies,
    )
    normalized_develop_users = _normalize_owner_list(develop_users)
    normalized_test_users = _normalize_owner_list(test_users)
    normalized_time = _normalize_time_filters(
        time_field,
        time_start,
        time_end,
        accepted_time_start,
        accepted_time_end,
    )
    requires_local_filter = bool(
        normalized_develop_users
        or normalized_test_users
        or normalized_time["time_start_dt"]
        or normalized_time["time_end_dt"]
    )

    remote_cache_payload = {
        "project_ids": ordered_project_ids,
        "design_ids": design_ids,
        "sub_teams": selected_teams,
        "categories": selected_categories,
        "verification_policies": selected_verification_policies,
    }
    cache_payload = {
        **remote_cache_payload,
        "develop_users": normalized_develop_users,
        "test_users": normalized_test_users,
        "time_field": normalized_time["time_field"] or "",
        "time_start": normalized_time["time_start"] or "",
        "time_end": normalized_time["time_end"] or "",
    }

    return {
        "projects": ordered_projects,
        "design_project_map": design_project_map,
        "design_ids": design_ids,
        "sub_teams": selected_teams,
        "categories": selected_categories,
        "verification_policies": selected_verification_policies,
        "develop_users": normalized_develop_users,
        "test_users": normalized_test_users,
        **normalized_time,
        "requires_local_filter": requires_local_filter,
        "remote_cache_payload": remote_cache_payload,
        "cache_payload": cache_payload,
    }


def _wait_for_cached_payload(cache_key: str, *, minimum_items_key: str | None = None) -> Any:
    for _ in range(10):
        time.sleep(0.3)
        cached = cache.get(cache_key)
        if minimum_items_key is None and cached is not None:
            return cached
        if (
            isinstance(cached, dict)
            and minimum_items_key
            and isinstance(cached.get(minimum_items_key), list)
        ):
            return cached
    return None


def _load_remote_page(context: dict[str, Any], page_no: int, page_size: int) -> dict[str, Any]:
    normalized_page_size = min(_parse_positive_int(page_size, 20), _UPSTREAM_PAGE_SIZE)
    payload = {
        **context["remote_cache_payload"],
        "page_no": page_no,
        "page_size": normalized_page_size,
    }
    cache_key = _cache_key("pm:requirement-board:page:v4", payload)
    cached = cache.get(cache_key)
    if isinstance(cached, dict) and isinstance(cached.get("items"), list):
        _debug_log(
            "data_page_cache_hit",
            cache_key=cache_key,
            page_no=page_no,
            page_size=normalized_page_size,
            total=cached.get("total"),
            page_sum=cached.get("page_sum"),
        )
        return cached

    raw_payload = _fetch_raw_page(
        design_ids=context["design_ids"],
        sub_teams=context["sub_teams"],
        categories=context["categories"],
        verification_policies=context["verification_policies"],
        page_no=page_no,
        page_size=normalized_page_size,
    )
    raw_items, raw_page = _extract_raw_page(raw_payload)
    current_page_no = _parse_positive_int(raw_page.get("page_no"), page_no)
    current_page_size = normalized_page_size
    items = _standardize_requirement_items(raw_items, context["design_project_map"])
    result = {
        "items": items,
        "total": _resolve_total(raw_page, current_page_no, current_page_size, len(items)),
        "page_no": current_page_no,
        "page_size": current_page_size,
        "page_sum": _resolve_page_sum(
            raw_page,
            current_page_no,
            current_page_size,
            len(items),
        ),
    }
    _debug_log(
        "data_page_loaded",
        page_no=current_page_no,
        page_size=current_page_size,
        page_sum=result["page_sum"],
        total=result["total"],
        item_count=len(items),
        upstream_page_sum=_parse_positive_int(raw_page.get("page_sum"), 0),
        upstream_page_size=_parse_positive_int(raw_page.get("page_size"), 0),
    )
    cache.set(cache_key, result, _DATA_CACHE_TTL_SECONDS)
    return result


def _item_matches_local_filters(item: dict[str, Any], context: dict[str, Any]) -> bool:
    develop_users = context["develop_users"]
    if develop_users and not set(develop_users).intersection(item.get("develop_users") or []):
        return False

    test_users = context["test_users"]
    if test_users and not set(test_users).intersection(item.get("test_users") or []):
        return False

    time_field = context.get("time_field")
    if time_field and (context.get("time_start_dt") or context.get("time_end_dt")):
        value_dt = _parse_datetime(item.get(time_field))
        if value_dt is None:
            return False
        if context.get("time_start_dt") and value_dt < context["time_start_dt"]:
            return False
        if context.get("time_end_dt") and value_dt > context["time_end_dt"]:
            return False

    return True


def _scan_all_filtered_items(
    context: dict[str, Any],
    *,
    page_size_setting_name: str = "REQUIREMENT_BOARD_SCAN_PAGE_SIZE",
    max_pages_setting_name: str = "REQUIREMENT_BOARD_SUMMARY_MAX_PAGES",
    limit_error_message: str = "需求扫描页数过多，请缩小筛选范围",
) -> list[dict[str, Any]]:
    cache_key = _cache_key("pm:requirement-board:filtered:v4", context["cache_payload"])
    cached = cache.get(cache_key)
    if isinstance(cached, dict) and isinstance(cached.get("items"), list):
        _debug_log(
            "filtered_result_cache_hit",
            cache_key=cache_key,
            matched_total=len(cached.get("items") or []),
        )
        return cached["items"]

    lock_key = f"{cache_key}:lock"
    lock_acquired = cache.add(lock_key, "1", _LOCK_TTL_SECONDS)
    if not lock_acquired:
        waiting = _wait_for_cached_payload(cache_key, minimum_items_key="items")
        if isinstance(waiting, dict) and isinstance(waiting.get("items"), list):
            _debug_log(
                "filtered_result_wait_hit",
                cache_key=cache_key,
                matched_total=len(waiting.get("items") or []),
            )
            return waiting["items"]

    page_size = _resolve_scan_page_size(page_size_setting_name)
    max_pages = _resolve_max_scan_pages(max_pages_setting_name)
    matched_items: list[dict[str, Any]] = []
    scanned_pages = 0
    _debug_log(
        "local_filter_scan_start",
        filters=context["cache_payload"],
        page_size=page_size,
        max_pages=max_pages,
    )

    try:
        for page_no, page_payload, fetched_count in _iterate_remote_pages(
            context,
            page_size=page_size,
            max_pages=max_pages,
            limit_error_message=limit_error_message,
        ):
            scanned_pages = page_no
            items = page_payload["items"]
            matched_before = len(matched_items)
            for item in items:
                if _item_matches_local_filters(item, context):
                    matched_items.append(item)

            _debug_log(
                "local_filter_scan_page",
                page_no=page_no,
                fetched_count=len(items),
                fetched_total=fetched_count,
                matched_count=len(matched_items) - matched_before,
                matched_total=len(matched_items),
                page_sum=page_payload.get("page_sum"),
                upstream_total=page_payload.get("total"),
            )

        cache.set(
            cache_key,
            {"items": matched_items},
            _FILTERED_SCAN_CACHE_TTL_SECONDS,
        )
        _debug_log(
            "local_filter_scan_done",
            matched_total=len(matched_items),
            scanned_pages=scanned_pages,
        )
        return matched_items
    finally:
        if lock_acquired:
            cache.delete(lock_key)


def _paginate_items(items: list[dict[str, Any]], page_no: int, page_size: int) -> dict[str, Any]:
    safe_page_no = _parse_positive_int(page_no, 1)
    safe_page_size = min(_parse_positive_int(page_size, 20), _UPSTREAM_PAGE_SIZE)
    total = len(items)
    page_sum = math.ceil(total / safe_page_size) if total and safe_page_size else 0
    start = max(safe_page_no - 1, 0) * safe_page_size
    end = start + safe_page_size
    return {
        "items": items[start:end],
        "total": total,
        "page_no": safe_page_no,
        "page_size": safe_page_size,
        "page_sum": page_sum,
    }


def get_filter_options() -> dict[str, Any]:
    projects = list(
        Project.objects.filter(is_deleted=False)
        .order_by("is_closed", "name")
        .only("id", "name", "code", "domain", "type", "design_id", "sub_teams", "is_closed")
    )
    project_items = [
        {
            "id": str(project.id),
            "name": project.name,
            "code": _clean_text(project.code),
            "domain": _clean_text(project.domain),
            "type": _clean_text(project.type),
            "design_id": _clean_text(project.design_id) or None,
            "sub_teams": _normalize_text_list(project.sub_teams),
            "config_complete": bool(
                _clean_text(project.design_id)
                and _normalize_text_list(project.sub_teams)
            ),
        }
        for project in projects
    ]
    project_items.sort(key=lambda item: (not item["config_complete"], item["name"]))
    return {
        "projects": project_items,
    }


def get_requirement_board_page(data: RequirementBoardDataQuerySchema) -> dict[str, Any]:
    context = _resolve_query_context(
        data.project_ids,
        data.sub_teams,
        data.categories,
        data.verification_policies,
        data.develop_users,
        data.test_users,
        data.time_field,
        data.time_start,
        data.time_end,
        data.accepted_time_start,
        data.accepted_time_end,
    )
    page_no = _parse_positive_int(data.page_no, 1)
    page_size = min(_parse_positive_int(data.page_size, 20), _UPSTREAM_PAGE_SIZE)
    _debug_log(
        "data_query_context",
        requires_local_filter=context["requires_local_filter"],
        filters=context["cache_payload"],
        page_no=page_no,
        page_size=page_size,
    )
    if context["requires_local_filter"]:
        items = _scan_all_filtered_items(context)
        result = _paginate_items(items, page_no=page_no, page_size=page_size)
        _debug_log(
            "data_query_result",
            mode="local_filter",
            page_no=result["page_no"],
            page_size=result["page_size"],
            page_sum=result["page_sum"],
            total=result["total"],
            item_count=len(result["items"]),
        )
        return result
    result = _load_remote_page(context, page_no=page_no, page_size=page_size)
    _debug_log(
        "data_query_result",
        mode="remote_page",
        page_no=result["page_no"],
        page_size=result["page_size"],
        page_sum=result["page_sum"],
        total=result["total"],
        item_count=len(result["items"]),
    )
    return result


def _create_empty_completion_payload() -> dict[str, Any]:
    return {
        "count": 0,
        "workload_man_day": 0.0,
        "workload_kloc": 0.0,
        "count_rate": 0.0,
        "workload_man_day_rate": 0.0,
        "workload_kloc_rate": 0.0,
    }


def _create_user_payload() -> dict[str, Any]:
    return {
        "task_count": 0,
        "workload_man_day": 0.0,
        "workload_kloc": 0.0,
    }


def _create_summary_accumulator() -> dict[str, Any]:
    return {
        "total_count": 0,
        "total_workload_man_day": 0.0,
        "total_workload_kloc": 0.0,
        "status_summary": {
            status: {
                "count": 0,
                "workload_man_day": 0.0,
                "workload_kloc": 0.0,
            }
            for status in STATUS_ORDER
        },
        "type_summary": {},
        "project_summary": {},
        "team_summary": {},
        "user_summary": {"develop_users": {}, "test_users": {}},
        "dispatch_rate": {
            "p_total": 0,
            "develop_owner_count": 0,
            "test_owner_count": 0,
        },
        "plan_refresh_rate": {
            "planned_test_time_count": 0,
            "due_date_count": 0,
        },
        "delay_summary": {
            "development": {"count": 0, "preview_items": []},
            "acceptance": {"count": 0, "preview_items": []},
        },
        "development_delivery_trend": {},
        "acceptance_delivery_trend": {},
    }


def _update_completion_payload(payload: dict[str, Any], item: dict[str, Any]) -> None:
    payload["count"] += 1
    payload["workload_man_day"] += _to_float(item.get("workload_man_day"))
    payload["workload_kloc"] += _to_float(item.get("workload_kloc"))


def _update_user_summary(
    user_summary: dict[str, dict[str, Any]],
    usernames: list[str],
    item: dict[str, Any],
) -> None:
    workload_man_day = _to_float(item.get("workload_man_day"))
    workload_kloc = _to_float(item.get("workload_kloc"))
    for username in usernames:
        row = user_summary.setdefault(username, _create_user_payload())
        row["task_count"] += 1
        row["workload_man_day"] += workload_man_day
        row["workload_kloc"] += workload_kloc


def _month_bucket(value: Any) -> str | None:
    parsed = _parse_datetime(value)
    if parsed is None:
        return None
    localized = timezone.localtime(_ensure_aware(parsed))
    return localized.strftime("%Y-%m")


def _push_delay_preview(
    preview_items: list[dict[str, Any]],
    item: dict[str, Any],
    *,
    sort_field: str,
) -> None:
    preview_items.append(item)

    def _preview_key(row: dict[str, Any]):
        parsed = _parse_datetime(row.get(sort_field))
        fallback = datetime.datetime.max.replace(tzinfo=datetime.timezone.utc)
        return parsed or fallback, row.get("requirement_id") or ""

    preview_items.sort(key=_preview_key)
    del preview_items[_DELAY_PREVIEW_LIMIT:]


def _aggregate_item(summary: dict[str, Any], item: dict[str, Any]) -> None:
    status_code = item["status_code"]
    category = item["category"]
    project_id = item.get("project_id") or ""
    project_name = item.get("project_name") or "未匹配项目"
    team_name = item.get("team_name") or UNKNOWN_TEAM_NAME
    workload_man_day = _to_float(item.get("workload_man_day"))
    workload_kloc = _to_float(item.get("workload_kloc"))
    develop_users = item.get("develop_users") or []
    test_users = item.get("test_users") or []

    summary["total_count"] += 1
    summary["total_workload_man_day"] += workload_man_day
    summary["total_workload_kloc"] += workload_kloc

    status_row = summary["status_summary"][status_code]
    status_row["count"] += 1
    status_row["workload_man_day"] += workload_man_day
    status_row["workload_kloc"] += workload_kloc

    type_row = summary["type_summary"].setdefault(
        category,
        {
            "category": category,
            "total_count": 0,
            "total_workload_man_day": 0.0,
            "total_workload_kloc": 0.0,
        },
    )
    type_row["total_count"] += 1
    type_row["total_workload_man_day"] += workload_man_day
    type_row["total_workload_kloc"] += workload_kloc

    project_key = project_id or project_name
    project_row = summary["project_summary"].setdefault(
        project_key,
        {
            "project_id": project_id,
            "project_name": project_name,
            "total_count": 0,
            "total_workload_man_day": 0.0,
            "total_workload_kloc": 0.0,
        },
    )
    project_row["total_count"] += 1
    project_row["total_workload_man_day"] += workload_man_day
    project_row["total_workload_kloc"] += workload_kloc

    team_row = summary["team_summary"].setdefault(
        team_name,
        {
            "team_name": team_name,
            "total_count": 0,
            "total_workload_man_day": 0.0,
            "total_workload_kloc": 0.0,
            "i_count": 0,
            "d_count": 0,
            "p_count": 0,
            "c_count": 0,
            "a_count": 0,
            "dev_done": _create_empty_completion_payload(),
            "acceptance_done": _create_empty_completion_payload(),
        },
    )
    team_row["total_count"] += 1
    team_row["total_workload_man_day"] += workload_man_day
    team_row["total_workload_kloc"] += workload_kloc
    team_row[_STATUS_FIELD_MAP[status_code]] += 1

    if status_code in {"C", "A"}:
        _update_completion_payload(team_row["dev_done"], item)
    if status_code == "A":
        _update_completion_payload(team_row["acceptance_done"], item)

    _update_user_summary(
        summary["user_summary"]["develop_users"],
        develop_users,
        item,
    )
    _update_user_summary(
        summary["user_summary"]["test_users"],
        test_users,
        item,
    )

    if status_code == "P":
        dispatch = summary["dispatch_rate"]
        dispatch["p_total"] += 1
        if develop_users:
            dispatch["develop_owner_count"] += 1
        if test_users:
            dispatch["test_owner_count"] += 1

    plan_refresh = summary["plan_refresh_rate"]
    if item.get("planned_test_time"):
        plan_refresh["planned_test_time_count"] += 1
    if item.get("due_date"):
        plan_refresh["due_date_count"] += 1

    if item.get("is_dev_delayed"):
        summary["delay_summary"]["development"]["count"] += 1
        _push_delay_preview(
            summary["delay_summary"]["development"]["preview_items"],
            item,
            sort_field="planned_test_time",
        )
    if item.get("is_test_delayed"):
        summary["delay_summary"]["acceptance"]["count"] += 1
        _push_delay_preview(
            summary["delay_summary"]["acceptance"]["preview_items"],
            item,
            sort_field="due_date",
        )

    development_month = _month_bucket(item.get("planned_test_time"))
    if development_month:
        row = summary["development_delivery_trend"].setdefault(
            development_month,
            {"planned_count": 0, "actual_count": 0},
        )
        row["planned_count"] += 1
    completed_month = _month_bucket(item.get("completed_time"))
    if completed_month:
        row = summary["development_delivery_trend"].setdefault(
            completed_month,
            {"planned_count": 0, "actual_count": 0},
        )
        row["actual_count"] += 1

    acceptance_month = _month_bucket(item.get("due_date"))
    if acceptance_month:
        row = summary["acceptance_delivery_trend"].setdefault(
            acceptance_month,
            {"planned_count": 0, "actual_count": 0},
        )
        row["planned_count"] += 1
    accepted_month = _month_bucket(item.get("accepted_time"))
    if accepted_month:
        row = summary["acceptance_delivery_trend"].setdefault(
            accepted_month,
            {"planned_count": 0, "actual_count": 0},
        )
        row["actual_count"] += 1


def _finalize_completion_payload(
    payload: dict[str, Any],
    total_count: int,
    total_workload_man_day: float,
    total_workload_kloc: float,
) -> dict[str, Any]:
    count_rate = payload["count"] / total_count if total_count else 0.0
    man_day_rate = (
        payload["workload_man_day"] / total_workload_man_day
        if total_workload_man_day
        else 0.0
    )
    kloc_rate = (
        payload["workload_kloc"] / total_workload_kloc if total_workload_kloc else 0.0
    )
    return {
        "count": int(payload["count"]),
        "workload_man_day": _round_metric(payload["workload_man_day"]),
        "workload_kloc": _round_metric(payload["workload_kloc"]),
        "count_rate": round(count_rate, 4),
        "workload_man_day_rate": round(man_day_rate, 4),
        "workload_kloc_rate": round(kloc_rate, 4),
    }


def _finalize_user_summary(user_summary: dict[str, dict[str, Any]]) -> list[dict[str, Any]]:
    result = [
        {
            "username": username,
            "task_count": int(stats["task_count"]),
            "workload_man_day": _round_metric(stats["workload_man_day"]),
            "workload_kloc": _round_metric(stats["workload_kloc"]),
        }
        for username, stats in user_summary.items()
    ]
    result.sort(
        key=lambda item: (
            -int(item["task_count"]),
            -float(item["workload_man_day"]),
            -float(item["workload_kloc"]),
            item["username"],
        ),
    )
    return result


def _finalize_trend_summary(trend_summary: dict[str, dict[str, Any]]) -> list[dict[str, Any]]:
    return [
        {
            "month": month,
            "planned_count": int(values.get("planned_count") or 0),
            "actual_count": int(values.get("actual_count") or 0),
        }
        for month, values in sorted(trend_summary.items(), key=lambda item: item[0])
    ]


def _finalize_summary_payload(summary: dict[str, Any]) -> dict[str, Any]:
    total_count = int(summary["total_count"])
    total_workload_man_day = float(summary["total_workload_man_day"])
    total_workload_kloc = float(summary["total_workload_kloc"])

    status_summary = []
    for status in STATUS_ORDER:
        row = summary["status_summary"][status]
        status_summary.append(
            {
                "status_code": status,
                "status_label": STATUS_LABELS[status],
                "count": int(row["count"]),
                "count_rate": round(
                    (row["count"] / total_count) if total_count else 0.0,
                    4,
                ),
                "workload_man_day": _round_metric(row["workload_man_day"]),
                "workload_kloc": _round_metric(row["workload_kloc"]),
            }
        )

    type_summary = sorted(
        (
            {
                "category": row["category"],
                "total_count": int(row["total_count"]),
                "total_workload_man_day": _round_metric(row["total_workload_man_day"]),
                "total_workload_kloc": _round_metric(row["total_workload_kloc"]),
            }
            for row in summary["type_summary"].values()
        ),
        key=lambda item: (
            -int(item["total_count"]),
            -float(item["total_workload_man_day"]),
            item["category"],
        ),
    )

    project_summary = sorted(
        (
            {
                "project_id": row["project_id"],
                "project_name": row["project_name"],
                "total_count": int(row["total_count"]),
                "total_workload_man_day": _round_metric(row["total_workload_man_day"]),
                "total_workload_kloc": _round_metric(row["total_workload_kloc"]),
            }
            for row in summary["project_summary"].values()
        ),
        key=lambda item: (
            -int(item["total_count"]),
            -float(item["total_workload_man_day"]),
            item["project_name"],
        ),
    )

    team_summary = []
    for row in summary["team_summary"].values():
        team_summary.append(
            {
                "team_name": row["team_name"],
                "total_count": int(row["total_count"]),
                "total_workload_man_day": _round_metric(row["total_workload_man_day"]),
                "total_workload_kloc": _round_metric(row["total_workload_kloc"]),
                "i_count": int(row["i_count"]),
                "d_count": int(row["d_count"]),
                "p_count": int(row["p_count"]),
                "c_count": int(row["c_count"]),
                "a_count": int(row["a_count"]),
                "dev_done": _finalize_completion_payload(
                    row["dev_done"],
                    int(row["total_count"]),
                    float(row["total_workload_man_day"]),
                    float(row["total_workload_kloc"]),
                ),
                "acceptance_done": _finalize_completion_payload(
                    row["acceptance_done"],
                    int(row["total_count"]),
                    float(row["total_workload_man_day"]),
                    float(row["total_workload_kloc"]),
                ),
            }
        )
    team_summary.sort(
        key=lambda item: (
            -int(item["total_count"]),
            -float(item["total_workload_man_day"]),
            item["team_name"],
        ),
    )

    delay_summary = {
        "development": {
            "count": int(summary["delay_summary"]["development"]["count"]),
            "rate": round(
                (
                    summary["delay_summary"]["development"]["count"] / total_count
                    if total_count
                    else 0.0
                ),
                4,
            ),
            "preview_items": summary["delay_summary"]["development"]["preview_items"],
        },
        "acceptance": {
            "count": int(summary["delay_summary"]["acceptance"]["count"]),
            "rate": round(
                (
                    summary["delay_summary"]["acceptance"]["count"] / total_count
                    if total_count
                    else 0.0
                ),
                4,
            ),
            "preview_items": summary["delay_summary"]["acceptance"]["preview_items"],
        },
    }

    dispatch = summary["dispatch_rate"]
    p_total = int(dispatch["p_total"])
    develop_owner_count = int(dispatch["develop_owner_count"])
    test_owner_count = int(dispatch["test_owner_count"])

    plan_refresh = summary["plan_refresh_rate"]
    planned_test_time_count = int(plan_refresh["planned_test_time_count"])
    due_date_count = int(plan_refresh["due_date_count"])

    return {
        "total_count": total_count,
        "total_workload_man_day": _round_metric(total_workload_man_day),
        "total_workload_kloc": _round_metric(total_workload_kloc),
        "status_summary": status_summary,
        "type_summary": type_summary,
        "project_summary": project_summary,
        "team_summary": team_summary,
        "user_summary": {
            "develop_users": _finalize_user_summary(
                summary["user_summary"]["develop_users"],
            ),
            "test_users": _finalize_user_summary(
                summary["user_summary"]["test_users"],
            ),
        },
        "dispatch_rate": {
            "p_total": p_total,
            "develop_owner_count": develop_owner_count,
            "develop_owner_rate": round(
                (develop_owner_count / p_total) if p_total else 0.0,
                4,
            ),
            "test_owner_count": test_owner_count,
            "test_owner_rate": round(
                (test_owner_count / p_total) if p_total else 0.0,
                4,
            ),
        },
        "plan_refresh_rate": {
            "planned_test_time_count": planned_test_time_count,
            "planned_test_time_rate": round(
                (planned_test_time_count / total_count) if total_count else 0.0,
                4,
            ),
            "due_date_count": due_date_count,
            "due_date_rate": round(
                (due_date_count / total_count) if total_count else 0.0,
                4,
            ),
        },
        "delay_summary": delay_summary,
        "development_delivery_trend": _finalize_trend_summary(
            summary["development_delivery_trend"],
        ),
        "acceptance_delivery_trend": _finalize_trend_summary(
            summary["acceptance_delivery_trend"],
        ),
    }


def _compute_summary_from_items(items: list[dict[str, Any]]) -> dict[str, Any]:
    summary = _create_summary_accumulator()
    for item in items:
        _aggregate_item(summary, item)
    return _finalize_summary_payload(summary)


def _compute_summary(context: dict[str, Any]) -> dict[str, Any]:
    _debug_log(
        "summary_compute_start",
        mode="local_filter" if context["requires_local_filter"] else "remote_scan",
        filters=context["cache_payload"],
    )
    if context["requires_local_filter"]:
        result = _compute_summary_from_items(_scan_all_filtered_items(context))
        _debug_log(
            "summary_compute_done",
            mode="local_filter",
            total_count=result.get("total_count"),
            team_count=len(result.get("team_summary") or []),
        )
        return result

    page_size = _resolve_scan_page_size("REQUIREMENT_BOARD_SUMMARY_PAGE_SIZE")
    max_pages = _resolve_max_scan_pages("REQUIREMENT_BOARD_SUMMARY_MAX_PAGES")
    summary = _create_summary_accumulator()
    scanned_pages = 0

    for page_no, page_payload, fetched_count in _iterate_remote_pages(
        context,
        page_size=page_size,
        max_pages=max_pages,
        limit_error_message="需求总结扫描页数过多，请缩小筛选范围",
    ):
        scanned_pages = page_no
        items = page_payload["items"]
        for item in items:
            _aggregate_item(summary, item)

        _debug_log(
            "summary_scan_page",
            page_no=page_no,
            item_count=len(items),
            fetched_total=fetched_count,
            accumulated_total=summary["total_count"],
            page_sum=page_payload.get("page_sum"),
            upstream_total=page_payload.get("total"),
        )

    result = _finalize_summary_payload(summary)
    _debug_log(
        "summary_compute_done",
        mode="remote_scan",
        total_count=result.get("total_count"),
        team_count=len(result.get("team_summary") or []),
        scanned_pages=scanned_pages,
    )
    return result


def _build_export_row(item: dict[str, Any]) -> list[Any]:
    return [
        _clean_text(item.get("project_name")),
        _clean_text(item.get("team_name")) or UNKNOWN_TEAM_NAME,
        _clean_text(item.get("category")),
        _clean_text(item.get("verification_policy_label")),
        _clean_text(item.get("requirement_id")),
        _clean_text(item.get("title")),
        _clean_text(item.get("status_code")),
        _clean_text(item.get("status_label")),
        _clean_text(item.get("planned_test_time")),
        _clean_text(item.get("due_date")),
        _clean_text(item.get("completed_time")),
        _clean_text(item.get("accepted_time")),
        "是" if _to_bool(item.get("is_dev_delayed"), False) else "否",
        "是" if _to_bool(item.get("is_test_delayed"), False) else "否",
        _round_metric(_to_float(item.get("workload_man_day"))),
        _round_metric(_to_float(item.get("workload_kloc"))),
        _clean_text(item.get("develop_user_display")),
        _clean_text(item.get("test_user_display")),
    ]


def _build_export_response(workbook: openpyxl.Workbook) -> HttpResponse:
    timestamp = timezone.localtime(_ensure_aware(timezone.now())).strftime(
        "%Y%m%d-%H%M%S"
    )
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = (
        f'attachment; filename="requirement-board-{timestamp}.xlsx"'
    )
    workbook.save(response)
    return response


def _export_items_to_workbook(items: list[dict[str, Any]]) -> HttpResponse:
    workbook = openpyxl.Workbook(write_only=True)
    worksheet = workbook.create_sheet(title=_EXPORT_SHEET_TITLE)
    worksheet.append(list(_EXPORT_HEADERS))
    for item in items:
        worksheet.append(_build_export_row(item))
    return _build_export_response(workbook)


def get_requirement_board_summary(
    data: RequirementBoardSummaryQuerySchema,
) -> dict[str, Any]:
    context = _resolve_query_context(
        data.project_ids,
        data.sub_teams,
        data.categories,
        data.verification_policies,
        data.develop_users,
        data.test_users,
        data.time_field,
        data.time_start,
        data.time_end,
        data.accepted_time_start,
        data.accepted_time_end,
    )
    summary_key = _cache_key("pm:requirement-board:summary:v4", context["cache_payload"])
    cached = cache.get(summary_key)
    if isinstance(cached, dict) and isinstance(cached.get("team_summary"), list):
        _debug_log("summary_cache_hit", cache_key=summary_key)
        return cached

    lock_key = f"{summary_key}:lock"
    lock_acquired = cache.add(lock_key, "1", _LOCK_TTL_SECONDS)
    if not lock_acquired:
        waiting = _wait_for_cached_payload(summary_key, minimum_items_key="team_summary")
        if isinstance(waiting, dict) and isinstance(waiting.get("team_summary"), list):
            _debug_log("summary_wait_hit", cache_key=summary_key)
            return waiting

    try:
        summary = _compute_summary(context)
        cache.set(summary_key, summary, _SUMMARY_CACHE_TTL_SECONDS)
        _debug_log(
            "summary_cached",
            cache_key=summary_key,
            total_count=summary.get("total_count"),
            team_count=len(summary.get("team_summary") or []),
        )
        return summary
    finally:
        if lock_acquired:
            cache.delete(lock_key)


def export_requirement_board_data(
    data: RequirementBoardExportQuerySchema,
) -> HttpResponse:
    context = _resolve_query_context(
        data.project_ids,
        data.sub_teams,
        data.categories,
        data.verification_policies,
        data.develop_users,
        data.test_users,
        data.time_field,
        data.time_start,
        data.time_end,
        data.accepted_time_start,
        data.accepted_time_end,
    )

    if context["requires_local_filter"]:
        items = _scan_all_filtered_items(
            context,
            page_size_setting_name="REQUIREMENT_BOARD_EXPORT_PAGE_SIZE",
            max_pages_setting_name="REQUIREMENT_BOARD_EXPORT_MAX_PAGES",
            limit_error_message="需求导出扫描页数过多，请缩小筛选范围",
        )
        _debug_log(
            "export_local_filter",
            total_count=len(items),
            filters=context["cache_payload"],
        )
        return _export_items_to_workbook(items)

    page_size = _resolve_scan_page_size("REQUIREMENT_BOARD_EXPORT_PAGE_SIZE")
    max_pages = _resolve_max_scan_pages("REQUIREMENT_BOARD_EXPORT_MAX_PAGES")
    workbook = openpyxl.Workbook(write_only=True)
    worksheet = workbook.create_sheet(title=_EXPORT_SHEET_TITLE)
    worksheet.append(list(_EXPORT_HEADERS))
    exported_count = 0
    scanned_pages = 0
    _debug_log(
        "export_remote_scan_start",
        filters=context["cache_payload"],
        page_size=page_size,
        max_pages=max_pages,
    )

    for page_no, page_payload, fetched_count in _iterate_remote_pages(
        context,
        page_size=page_size,
        max_pages=max_pages,
        limit_error_message="需求导出扫描页数过多，请缩小筛选范围",
    ):
        scanned_pages = page_no
        items = page_payload["items"]
        for item in items:
            worksheet.append(_build_export_row(item))
            exported_count += 1
        _debug_log(
            "export_remote_scan_page",
            page_no=page_no,
            fetched_count=len(items),
            fetched_total=fetched_count,
            exported_total=exported_count,
            page_sum=page_payload.get("page_sum"),
            upstream_total=page_payload.get("total"),
        )

    _debug_log(
        "export_remote_scan_done",
        exported_total=exported_count,
        scanned_pages=scanned_pages,
        filters=context["cache_payload"],
    )
    return _build_export_response(workbook)
