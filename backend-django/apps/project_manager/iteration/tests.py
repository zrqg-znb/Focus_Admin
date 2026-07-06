from datetime import date
from io import BytesIO
from unittest import mock

from django.test import RequestFactory, TestCase
from openpyxl import load_workbook

from apps.project_manager.iteration import iteration_api, iteration_service
from apps.project_manager.iteration.iteration_model import Iteration, IterationMetric
from apps.project_manager.project.project_model import Project


class IterationDetailExportTests(TestCase):
    def setUp(self):
        self.factory = RequestFactory()
        self.project = Project.objects.create(
            name="健康项目",
            domain="车控",
            type="量产",
            code="health-project",
            enable_iteration=True,
            sub_teams=["Team-A"],
        )
        self.iteration = Iteration.objects.create(
            project=self.project,
            name="迭代一",
            code="IT-001",
            start_date=date(2026, 7, 1),
            end_date=date(2026, 7, 31),
            is_current=True,
            is_healthy=True,
        )
        IterationMetric.objects.create(
            iteration=self.iteration,
            record_date=date(2026, 7, 6),
            sr_num=2,
            dr_num=3,
            ar_num=1,
            need_break_sr_num=2,
            need_break_dr_num=2,
            need_break_but_un_break_sr_num=1,
            need_break_but_un_break_dr_num=0,
            a_state_ar_num=1,
            c_state_ar_num=0,
            a_state_dr_num=2,
            c_state_dr_num=1,
            test_automation_rate=0.8,
            test_case_execution_rate=0.9,
            bug_fix_rate=0.7,
            code_review_rate=0.6,
            code_coverage_rate=0.5,
        )

    def _load_export_workbook(self, requirements: list[dict]):
        with mock.patch.object(
            iteration_service,
            "get_cached_iteration_requirements",
            return_value=requirements,
        ):
            response = iteration_api.export_iteration_detail(
                self.factory.get(
                    f"/api/project-manager/iterations/iteration/{self.iteration.id}/export-detail"
                ),
                str(self.iteration.id),
            )

        workbook = load_workbook(BytesIO(response.content), read_only=True)
        return response, workbook

    def test_export_detail_creates_expected_sheets_and_content_type(self):
        """导出文件应包含详情页各类数据的独立 sheet。"""
        response, workbook = self._load_export_workbook(
            [
                {
                    "requirement_id": "REQ-001",
                    "title": "开发责任人优先",
                    "requirement_type": "dr",
                    "idpca_status": "P",
                    "owner_team": "Team-A",
                    "develop_owner": "dev-a",
                    "owner": "legacy-owner",
                    "need_breakdown": True,
                    "is_decomposed": False,
                    "workload_man_filled": True,
                    "workload_loc_filled": False,
                }
            ]
        )

        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertEqual(
            workbook.sheetnames,
            ["迭代基础信息", "入口指标", "出口指标", "需求IDPCA状态", "未分解需求"],
        )
        base_rows = list(workbook["迭代基础信息"].iter_rows(values_only=True))
        self.assertEqual(base_rows[1][0], "健康项目")
        self.assertEqual(base_rows[1][1], "迭代一")

    def test_export_detail_uses_develop_owner_priority_and_legacy_fallback(self):
        """需求责任人导出应优先使用开发责任人，并兼容历史 owner 字段。"""
        _, workbook = self._load_export_workbook(
            [
                {
                    "requirement_id": "REQ-001",
                    "title": "显式开发责任人",
                    "requirement_type": "dr",
                    "idpca_status": "P",
                    "owner_team": "Team-A",
                    "develop_owner": "dev-a",
                    "owner": "legacy-owner",
                    "need_breakdown": True,
                    "is_decomposed": True,
                },
                {
                    "requirement_id": "REQ-002",
                    "title": "列表开发责任人",
                    "requirement_type": "ar",
                    "idpca_status": "C",
                    "owner_team": "Team-B",
                    "develop_users": ["dev-b", "dev-c"],
                    "owner": "legacy-owner",
                    "need_breakdown": False,
                    "is_decomposed": True,
                },
                {
                    "requirement_id": "REQ-003",
                    "title": "历史责任人",
                    "requirement_type": "sr",
                    "idpca_status": "I",
                    "owner_team": "Team-C",
                    "owner": "legacy-owner",
                    "need_breakdown": True,
                    "is_decomposed": False,
                },
            ]
        )

        rows = list(workbook["需求IDPCA状态"].iter_rows(min_row=2, values_only=True))
        develop_owner_by_id = {row[0]: row[5] for row in rows}
        self.assertEqual(develop_owner_by_id["REQ-001"], "dev-a")
        self.assertEqual(develop_owner_by_id["REQ-002"], "dev-b, dev-c")
        self.assertEqual(develop_owner_by_id["REQ-003"], "legacy-owner")

    def test_export_detail_unresolved_sheet_only_contains_unresolved_items(self):
        """未分解需求 sheet 只导出需分解且未分解的需求。"""
        _, workbook = self._load_export_workbook(
            [
                {
                    "requirement_id": "REQ-001",
                    "title": "未分解",
                    "requirement_type": "dr",
                    "idpca_status": "P",
                    "need_breakdown": True,
                    "is_decomposed": False,
                    "develop_user": "dev-a",
                },
                {
                    "requirement_id": "REQ-002",
                    "title": "已分解",
                    "requirement_type": "dr",
                    "idpca_status": "C",
                    "need_breakdown": True,
                    "is_decomposed": True,
                    "develop_user": "dev-b",
                },
                {
                    "requirement_id": "REQ-003",
                    "title": "无需分解",
                    "requirement_type": "ar",
                    "idpca_status": "A",
                    "need_breakdown": False,
                    "is_decomposed": True,
                    "develop_user": "dev-c",
                },
            ]
        )

        rows = list(workbook["未分解需求"].iter_rows(min_row=2, values_only=True))
        self.assertEqual([row[0] for row in rows], ["REQ-001"])
