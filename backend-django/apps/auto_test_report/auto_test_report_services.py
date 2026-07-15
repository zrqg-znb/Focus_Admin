import io
from collections import Counter, defaultdict
from datetime import datetime, timedelta
from typing import Optional

import openpyxl
from django.conf import settings
from django.db import IntegrityError, transaction
from django.db.models import F, OuterRef, Q, Subquery, Window
from django.db.models.functions import RowNumber
from django.http import HttpResponse
from django.utils import timezone
from ninja.errors import HttpError
from scheduler.module.executor import scheduler_task

from .auto_test_report_model import (
    DOMAIN_COCKPIT,
    DOMAIN_COCKPIT_SOC,
    DOMAIN_VEHICLE,
    DailyExecutionBatch,
    DailyExecutionResult,
    DownstreamCommit,
    DownstreamCommitUsage,
    DOWNSTREAM_TRIGGER_MANUAL,
    DOWNSTREAM_TRIGGER_SCHEDULED,
    FAILURE_CATEGORY_CASE,
    FAILURE_CATEGORY_CHOICES,
    FAILURE_CATEGORY_ENVIRONMENT,
    FAILURE_CATEGORY_VERSION,
    McuPlatform,
    TestCase,
    VehicleModel,
    VIU_CODE_VALUES,
    RESULT_FAILED,
    RESULT_SUCCESS,
    RESULT_TIMEOUT,
    RESULT_SKIP,
)
from .auto_test_report_schemas import (
    DailyHistoryPage,
    DailyOverviewResponse,
    DailyOverviewRow,
    DailyOverviewSummary,
    DailyHistoryRow,
    DailyResultItemOut,
    DailySummaryOut,
    DownstreamCommitOut,
    DownstreamCommitUsageOut,
    DownstreamTriggerOut,
    ImportErrorRow,
    ImportResultOut,
    SummaryStat,
)


RESULT_LABELS = {
    RESULT_SUCCESS: '成功',
    RESULT_FAILED: '失败',
    RESULT_TIMEOUT: '超时',
    RESULT_SKIP: '跳过',
    'missing': '未执行',
}
RESULT_MISSING = 'missing'
MANUAL_REASON_RESULTS = {RESULT_FAILED, RESULT_TIMEOUT, RESULT_SKIP}
VALID_RESULT_VALUES = {
    RESULT_SUCCESS,
    RESULT_FAILED,
    RESULT_TIMEOUT,
    RESULT_SKIP,
}
VALID_DOMAINS = {DOMAIN_COCKPIT, DOMAIN_COCKPIT_SOC, DOMAIN_VEHICLE}
DOMAIN_ERROR_MESSAGE = '领域仅支持 cockpit、cockpit_soc 或 vehicle'
NON_SUCCESS_RESULTS = {RESULT_FAILED, RESULT_TIMEOUT, RESULT_SKIP}
VALID_FAILURE_CATEGORIES = {value for value, _ in FAILURE_CATEGORY_CHOICES}
NON_VERSION_FAILURE_CATEGORIES = {
    FAILURE_CATEGORY_ENVIRONMENT,
    FAILURE_CATEGORY_CASE,
}
TESTCASE_LOG_SUFFIX = 'testcase.html'


def _get_latest_result_order_by():
    return [
        F('start_time').desc(),
        F('reported_at').desc(),
        F('sys_create_datetime').desc(),
    ]


def build_latest_daily_results_queryset(
    *,
    execute_date=None,
    test_case_id: Optional[str] = None,
    vehicle: Optional[VehicleModel] = None,
    vehicle_id: Optional[str] = None,
):
    queryset = DailyExecutionResult.objects.filter(
        is_deleted=False,
        test_case__is_deleted=False,
        vehicle__is_deleted=False,
    )
    if vehicle is not None:
        queryset = queryset.filter(vehicle=vehicle)
    if vehicle_id:
        queryset = queryset.filter(vehicle_id=vehicle_id)
    if execute_date is not None:
        queryset = queryset.filter(execute_date=execute_date)
    if test_case_id:
        queryset = queryset.filter(test_case_id=test_case_id)
    return queryset.annotate(
        latest_rank=Window(
            expression=RowNumber(),
            partition_by=[F('vehicle_id'), F('execute_date'), F('test_case_id')],
            order_by=_get_latest_result_order_by(),
        ),
    ).filter(latest_rank=1)


def _apply_audit_fields(instance, user, *, is_create: bool = False):
    if not user:
        return
    if is_create and hasattr(instance, 'sys_creator'):
        instance.sys_creator = user
    if hasattr(instance, 'sys_modifier'):
        instance.sys_modifier = user


def _is_daily_overview_abnormal(active_case_count: int, latest_results: list[DailyExecutionResult]) -> bool:
    """按注册用例口径判断车型是否异常，漏传结果也必须暴露为异常。"""
    missing_result_count = max(active_case_count - len(latest_results), 0)
    if missing_result_count > 0:
        return True
    if active_case_count <= 0:
        return True
    return any(item.result != RESULT_SUCCESS for item in latest_results)


def _parse_domain_filter(domain: Optional[str]):
    value = (domain or '').strip().lower()
    if not value:
        return None
    if value not in VALID_DOMAINS:
        raise HttpError(422, DOMAIN_ERROR_MESSAGE)
    return value


def _resolve_domain_value(domain: Optional[str], *, default: str = DOMAIN_COCKPIT):
    value = (domain or '').strip().lower()
    if not value:
        return default
    if value not in VALID_DOMAINS:
        raise HttpError(422, DOMAIN_ERROR_MESSAGE)
    return value


def derive_car_log_url(log_url: Optional[str]) -> Optional[str]:
    """从运行日志 URL 中派生车机日志 URL。"""
    value = (log_url or '').strip()
    if not value or not value.endswith(TESTCASE_LOG_SUFFIX):
        return None
    car_log_url = value[: -len(TESTCASE_LOG_SUFFIX)]
    return car_log_url or None


def _normalize_failure_category(value: Optional[str]) -> Optional[str]:
    """标准化失败根因大类，空值表示尚未分类。"""
    normalized = (value or '').strip().lower()
    if not normalized:
        return None
    if normalized not in VALID_FAILURE_CATEGORIES:
        raise HttpError(422, '失败根因大类仅支持 version、environment、case')
    return normalized


def _is_non_version_failure(result: DailyExecutionResult) -> bool:
    """判断一条非成功结果是否属于不阻塞人工触发的根因。"""
    return result.result in NON_SUCCESS_RESULTS and result.failure_category in NON_VERSION_FAILURE_CATEGORIES


def _normalize_viu_codes(viu_codes, *, require_non_empty: bool = False):
    normalized = []
    for raw_code in viu_codes or []:
        code = str(raw_code or '').strip().lower()
        if not code:
            continue
        if code not in VIU_CODE_VALUES:
            raise HttpError(422, f'VIU编号仅支持: {", ".join(VIU_CODE_VALUES)}')
        if code not in normalized:
            normalized.append(code)
    if require_non_empty and not normalized:
        raise HttpError(422, '车控车型至少需要配置一个VIU编号')
    return normalized


def _normalize_case_viu_code(vehicle: VehicleModel, viu_code: Optional[str]):
    domain = vehicle.platform.domain
    if domain != DOMAIN_VEHICLE:
        return ''
    normalized = (viu_code or '').strip().lower()
    if not normalized:
        raise HttpError(422, '车控领域用例必须配置VIU编号')
    allowed_viu_codes = set(vehicle.viu_codes or [])
    if normalized not in allowed_viu_codes:
        raise HttpError(422, f'车型 {vehicle.name} 未配置 VIU 编号: {normalized}')
    return normalized


def _normalize_case_module(vehicle: VehicleModel, module: Optional[str]):
    """标准化 SOC 用例模块，只有座舱 SOC 领域强制填写模块。"""
    normalized = (module or '').strip()
    if vehicle.platform.domain == DOMAIN_COCKPIT_SOC and not normalized:
        raise HttpError(422, '座舱SOC用例必须填写模块')
    return normalized if vehicle.platform.domain == DOMAIN_COCKPIT_SOC else ''


def list_platforms(domain: Optional[str] = None):
    queryset = McuPlatform.objects.filter(is_deleted=False)
    parsed_domain = _parse_domain_filter(domain)
    if parsed_domain:
        queryset = queryset.filter(domain=parsed_domain)
    queryset = queryset.order_by('domain', '-sort', 'name')
    return [serialize_platform(item) for item in queryset]


def serialize_platform(item: McuPlatform):
    return {
        'id': str(item.id),
        'name': item.name,
        'version_code': item.version_code,
        'domain': item.domain,
        'sort': item.sort,
        'is_active': item.is_active,
        'remark': item.remark,
        'vehicle_count': item.vehicles.filter(is_deleted=False).count(),
        'sys_create_datetime': item.sys_create_datetime,
        'sys_update_datetime': item.sys_update_datetime,
    }


def create_platform(user, payload):
    instance = McuPlatform(
        name=payload.name.strip(),
        version_code=payload.version_code.strip(),
        domain=_resolve_domain_value(getattr(payload, 'domain', None)),
        sort=payload.sort,
        is_active=payload.is_active,
        remark=(payload.remark or '').strip() or None,
    )
    _apply_audit_fields(instance, user, is_create=True)
    instance.save()
    return serialize_platform(instance)


def update_platform(user, platform_id: str, payload):
    instance = get_platform(platform_id)
    instance.name = payload.name.strip()
    instance.version_code = payload.version_code.strip()
    instance.domain = _resolve_domain_value(
        getattr(payload, 'domain', None),
        default=instance.domain,
    )
    instance.sort = payload.sort
    instance.is_active = payload.is_active
    instance.remark = (payload.remark or '').strip() or None
    _apply_audit_fields(instance, user)
    instance.save()
    return serialize_platform(instance)


def delete_platform(platform_id: str):
    instance = get_platform(platform_id)
    if instance.vehicles.filter(is_deleted=False).exists():
        raise HttpError(400, '该平台下存在车型，无法删除')
    instance.soft_delete()
    return True


def get_platform(platform_id: str) -> McuPlatform:
    instance = McuPlatform.objects.filter(id=platform_id, is_deleted=False).first()
    if not instance:
        raise HttpError(404, '平台不存在')
    return instance


def list_vehicles(domain: Optional[str] = None, platform_id: Optional[str] = None, keyword: str = ''):
    queryset = VehicleModel.objects.select_related('platform').filter(is_deleted=False)
    parsed_domain = _parse_domain_filter(domain)
    if parsed_domain:
        queryset = queryset.filter(platform__domain=parsed_domain)
    if platform_id:
        queryset = queryset.filter(platform_id=platform_id)
    keyword = (keyword or '').strip()
    if keyword:
        queryset = queryset.filter(
            Q(name__icontains=keyword)
            | Q(vehicle_code__icontains=keyword)
            | Q(cdc_platform__icontains=keyword)
            | Q(execution_machine__icontains=keyword)
        )
    queryset = queryset.order_by('-sort', 'platform__sort', 'name')
    return [serialize_vehicle(item) for item in queryset]


def serialize_vehicle(item: VehicleModel):
    return {
        'id': str(item.id),
        'platform_id': str(item.platform_id),
        'platform_name': item.platform.name,
        'viu_codes': list(item.viu_codes or []),
        'name': item.name,
        'vehicle_code': item.vehicle_code,
        'cdc_platform': item.cdc_platform,
        'execution_machine': item.execution_machine,
        'sort': item.sort,
        'is_active': item.is_active,
        'remark': item.remark,
        'sys_create_datetime': item.sys_create_datetime,
        'sys_update_datetime': item.sys_update_datetime,
    }


def list_vehicle_options(domain: Optional[str] = None):
    queryset = VehicleModel.objects.select_related('platform').filter(
        is_deleted=False,
        is_active=True,
        platform__is_deleted=False,
        platform__is_active=True,
    ).order_by('platform__sort', 'platform__name', 'name')
    parsed_domain = _parse_domain_filter(domain)
    if parsed_domain:
        queryset = queryset.filter(platform__domain=parsed_domain)
    return [
        {
            'id': str(item.id),
            'name': item.name,
            'vehicle_code': item.vehicle_code,
            'platform_id': str(item.platform_id),
            'platform_name': item.platform.name,
            'viu_codes': list(item.viu_codes or []),
        }
        for item in queryset
    ]


def create_vehicle(user, payload):
    platform = get_platform(payload.platform_id)
    parsed_viu_codes = _normalize_viu_codes(payload.viu_codes, require_non_empty=platform.domain == DOMAIN_VEHICLE)
    if platform.domain != DOMAIN_VEHICLE:
        parsed_viu_codes = []
    instance = VehicleModel(
        platform=platform,
        name=payload.name.strip(),
        vehicle_code=payload.vehicle_code.strip(),
        cdc_platform=payload.cdc_platform.strip(),
        execution_machine=payload.execution_machine.strip(),
        viu_codes=parsed_viu_codes,
        sort=payload.sort,
        is_active=payload.is_active,
        remark=(payload.remark or '').strip() or None,
    )
    _apply_audit_fields(instance, user, is_create=True)
    instance.save()
    return serialize_vehicle(instance)


def update_vehicle(user, vehicle_id: str, payload):
    instance = get_vehicle(vehicle_id)
    instance.platform = get_platform(payload.platform_id)
    parsed_viu_codes = _normalize_viu_codes(
        payload.viu_codes,
        require_non_empty=instance.platform.domain == DOMAIN_VEHICLE,
    )
    if instance.platform.domain != DOMAIN_VEHICLE:
        parsed_viu_codes = []
    instance.name = payload.name.strip()
    instance.vehicle_code = payload.vehicle_code.strip()
    instance.cdc_platform = payload.cdc_platform.strip()
    instance.execution_machine = payload.execution_machine.strip()
    instance.viu_codes = parsed_viu_codes
    instance.sort = payload.sort
    instance.is_active = payload.is_active
    instance.remark = (payload.remark or '').strip() or None
    _apply_audit_fields(instance, user)
    instance.save()
    return serialize_vehicle(instance)


def delete_vehicle(vehicle_id: str):
    instance = get_vehicle(vehicle_id)
    instance.soft_delete()
    return True


def get_vehicle(vehicle_id: str) -> VehicleModel:
    instance = VehicleModel.objects.select_related('platform').filter(id=vehicle_id, is_deleted=False).first()
    if not instance:
        raise HttpError(404, '车型不存在')
    return instance


def list_test_cases(filters):
    latest_execute_time_subquery = (
        DailyExecutionResult.objects.filter(
            test_case_id=OuterRef('pk'),
            is_deleted=False,
        )
        .order_by(
            '-execute_date',
            '-start_time',
            '-reported_at',
            '-sys_create_datetime',
        )
        .values('start_time')[:1]
    )
    queryset = (
        TestCase.objects.select_related('vehicle', 'vehicle__platform')
        .filter(is_deleted=False, vehicle__is_deleted=False)
        .annotate(latest_execute_time=Subquery(latest_execute_time_subquery))
    )
    parsed_domain = _parse_domain_filter(getattr(filters, 'domain', None))
    if parsed_domain:
        queryset = queryset.filter(vehicle__platform__domain=parsed_domain)
    if filters.platform_id:
        queryset = queryset.filter(vehicle__platform_id=filters.platform_id)
    if filters.vehicle_id:
        queryset = queryset.filter(vehicle_id=filters.vehicle_id)
    if filters.viu_code:
        queryset = queryset.filter(viu_code=(filters.viu_code or '').strip().lower())
    if filters.is_active is not None:
        queryset = queryset.filter(is_active=filters.is_active)
    keyword = (filters.keyword or '').strip()
    if keyword:
        queryset = queryset.filter(
            Q(case_no__icontains=keyword)
            | Q(case_name__icontains=keyword)
            | Q(module__icontains=keyword)
            | Q(remark__icontains=keyword)
            | Q(viu_code__icontains=keyword)
        )
    queryset = queryset.order_by('-sort', 'viu_code', 'case_no')
    return [serialize_test_case(item) for item in queryset]


def serialize_test_case(item: TestCase):
    latest_execute_time = getattr(item, 'latest_execute_time', None)
    if latest_execute_time is None:
        latest_result = (
            item.daily_results.filter(is_deleted=False)
            .order_by(
                '-execute_date',
                '-start_time',
                '-reported_at',
                '-sys_create_datetime',
            )
            .first()
        )
        latest_execute_time = latest_result.start_time if latest_result else None
    return {
        'id': str(item.id),
        'vehicle_id': str(item.vehicle_id),
        'vehicle_name': item.vehicle.name,
        'vehicle_code': item.vehicle.vehicle_code,
        'platform_name': item.vehicle.platform.name,
        'viu_code': item.viu_code,
        'module': item.module,
        'case_no': item.case_no,
        'case_name': item.case_name,
        'remark': item.remark,
        'sort': item.sort,
        'is_active': item.is_active,
        'latest_execute_time': latest_execute_time,
        'sys_create_datetime': item.sys_create_datetime,
        'sys_update_datetime': item.sys_update_datetime,
    }


def create_test_case(user, payload):
    vehicle = get_vehicle(payload.vehicle_id)
    viu_code = _normalize_case_viu_code(vehicle, getattr(payload, 'viu_code', None))
    module = _normalize_case_module(vehicle, getattr(payload, 'module', None))
    instance = TestCase(
        vehicle=vehicle,
        viu_code=viu_code,
        module=module,
        case_no=payload.case_no.strip(),
        case_name=payload.case_name.strip(),
        remark=(payload.remark or '').strip() or None,
        sort=payload.sort,
        is_active=payload.is_active,
    )
    _apply_audit_fields(instance, user, is_create=True)
    instance.save()
    return serialize_test_case(instance)


def update_test_case(user, case_id: str, payload):
    instance = get_test_case(case_id)
    vehicle = get_vehicle(payload.vehicle_id)
    instance.vehicle = vehicle
    instance.viu_code = _normalize_case_viu_code(vehicle, getattr(payload, 'viu_code', None))
    instance.module = _normalize_case_module(vehicle, getattr(payload, 'module', None))
    instance.case_no = payload.case_no.strip()
    instance.case_name = payload.case_name.strip()
    instance.remark = (payload.remark or '').strip() or None
    instance.sort = payload.sort
    instance.is_active = payload.is_active
    _apply_audit_fields(instance, user)
    instance.save()
    return serialize_test_case(instance)


def update_test_case_remark(user, case_id: str, remark: Optional[str]):
    instance = get_test_case(case_id)
    instance.remark = (remark or '').strip() or None
    _apply_audit_fields(instance, user)
    instance.save()
    return serialize_test_case(instance)


def delete_test_case(case_id: str):
    instance = get_test_case(case_id)
    instance.soft_delete()
    return True


def batch_delete_test_cases(case_ids):
    queryset = TestCase.objects.filter(id__in=case_ids, is_deleted=False)
    count = 0
    for item in queryset:
        item.soft_delete()
        count += 1
    return count


def get_test_case(case_id: str) -> TestCase:
    instance = TestCase.objects.select_related('vehicle', 'vehicle__platform').filter(id=case_id, is_deleted=False).first()
    if not instance:
        raise HttpError(404, '测试用例不存在')
    return instance


@transaction.atomic
def import_test_cases(user, payload) -> ImportResultOut:
    vehicle = get_vehicle(payload.vehicle_id)
    require_viu_code = vehicle.platform.domain == DOMAIN_VEHICLE
    require_module = vehicle.platform.domain == DOMAIN_COCKPIT_SOC
    created_count = 0
    updated_count = 0
    ignored_count = 0
    errors = []
    seen_case_keys = set()

    for index, row in enumerate(payload.rows, start=1):
        raw_viu_code = (row.viu_code or '').strip().lower()
        module = (row.module or '').strip()
        if require_viu_code and not raw_viu_code:
            errors.append(ImportErrorRow(row_no=index, message='车控车型导入时VIU编号不能为空'))
            continue
        if require_viu_code and raw_viu_code not in set(vehicle.viu_codes or []):
            errors.append(ImportErrorRow(row_no=index, message=f'车型 {vehicle.name} 未配置 VIU 编号: {raw_viu_code}'))
            continue
        if require_module and not module:
            errors.append(ImportErrorRow(row_no=index, message='座舱SOC车型导入时模块不能为空'))
            continue
        if not require_module:
            module = ''
        viu_code = raw_viu_code if require_viu_code else ''
        case_no = (row.case_no or '').strip()
        case_name = (row.case_name or '').strip()
        remark = (row.remark or '').strip() or None
        if not case_no:
            errors.append(ImportErrorRow(row_no=index, message='用例编号不能为空'))
            continue
        if not case_name:
            errors.append(ImportErrorRow(row_no=index, message='用例名称不能为空'))
            continue
        case_key = (viu_code, case_no)
        if case_key in seen_case_keys:
            errors.append(ImportErrorRow(row_no=index, message='Excel 内VIU编号+用例编号重复'))
            continue
        seen_case_keys.add(case_key)

        instance = TestCase.objects.filter(
            vehicle=vehicle,
            viu_code=viu_code,
            case_no=case_no,
            is_deleted=False,
        ).first()
        if not instance:
            instance = TestCase(
                vehicle=vehicle,
                viu_code=viu_code,
                module=module,
                case_no=case_no,
                case_name=case_name,
                remark=remark,
                is_active=True,
            )
            _apply_audit_fields(instance, user, is_create=True)
            instance.save()
            created_count += 1
            continue
        if instance.case_name != case_name or instance.module != module or instance.remark != remark:
            instance.case_name = case_name
            instance.module = module
            instance.remark = remark
            _apply_audit_fields(instance, user)
            instance.save(update_fields=['case_name', 'module', 'remark', 'sys_modifier', 'sys_update_datetime'])
            updated_count += 1
        else:
            ignored_count += 1

    return ImportResultOut(
        created_count=created_count,
        updated_count=updated_count,
        ignored_count=ignored_count,
        errors=errors,
    )


def _read_excel_rows(file_obj) -> list[tuple]:
    """读取 Excel 活动工作表，统一处理文件损坏和空表错误。"""
    try:
        content = file_obj.read()
        workbook = openpyxl.load_workbook(
            filename=io.BytesIO(content),
            read_only=True,
            data_only=True,
        )
    except Exception as exc:
        raise HttpError(400, f'Excel 解析失败: {exc}')

    rows = list(workbook.active.iter_rows(values_only=True))
    if not rows:
        raise HttpError(400, 'Excel 内容为空')
    return rows


def parse_excel_rows(file_obj, *, require_module: bool = False, require_viu_code: bool = False) -> list[dict]:
    """解析兼容旧接口的单车型用例 Excel。"""
    rows = _read_excel_rows(file_obj)

    header = [str(item or '').strip() for item in rows[0]]
    try:
        case_no_index = header.index('用例编号')
        case_name_index = header.index('用例名称')
    except ValueError:
        raise HttpError(400, 'Excel 模板缺少必填列：用例编号、用例名称')
    viu_code_index = header.index('VIU编号') if 'VIU编号' in header else None
    if require_viu_code and viu_code_index is None:
        raise HttpError(400, 'Excel 模板缺少必填列：VIU编号')
    module_index = header.index('模块') if '模块' in header else None
    if require_module and module_index is None:
        raise HttpError(400, 'Excel 模板缺少必填列：模块')
    remark_index = header.index('备注') if '备注' in header else None

    parsed_rows = []
    for row in rows[1:]:
        if not row:
            continue
        case_no = str(row[case_no_index] or '').strip()
        case_name = str(row[case_name_index] or '').strip()
        if not case_no and not case_name:
            continue
        parsed_rows.append({
            'viu_code': str(row[viu_code_index] or '').strip() if viu_code_index is not None else '',
            'module': str(row[module_index] or '').strip() if module_index is not None else '',
            'case_no': case_no,
            'case_name': case_name,
            'remark': str(row[remark_index] or '').strip() if remark_index is not None else '',
        })
    return parsed_rows


def parse_full_test_case_excel_rows(file_obj, domain: str) -> list[dict]:
    """解析一站式导入模板，并保留真实 Excel 行号用于错误定位。"""
    rows = _read_excel_rows(file_obj)
    header = [str(item or '').strip() for item in rows[0]]
    required_headers = [
        '版本名称',
        '版本标识',
        '车型名称',
        '车型编号',
        '执行机器',
        '用例编号',
        '用例名称',
        '备注',
    ]
    if domain != DOMAIN_VEHICLE:
        required_headers.append('CDC平台')
    if domain == DOMAIN_VEHICLE:
        required_headers.append('VIU编号')
    if domain == DOMAIN_COCKPIT_SOC:
        required_headers.append('模块')
    missing_headers = [name for name in required_headers if name not in header]
    if missing_headers:
        raise HttpError(400, f'Excel 模板缺少必填列：{"、".join(missing_headers)}')

    indexes = {name: header.index(name) for name in required_headers}

    def read_cell(row, name: str) -> str:
        """按列名读取单元格，缺失或空值统一转换为空字符串。"""
        index = indexes.get(name)
        if index is None or index >= len(row):
            return ''
        return str(row[index] or '').strip()

    parsed_rows = []
    for row_no, row in enumerate(rows[1:], start=2):
        if not row or not any(str(item or '').strip() for item in row):
            continue
        parsed_rows.append({
            'row_no': row_no,
            'platform_name': read_cell(row, '版本名称'),
            'version_code': read_cell(row, '版本标识'),
            'vehicle_name': read_cell(row, '车型名称'),
            'vehicle_code': read_cell(row, '车型编号'),
            'cdc_platform': read_cell(row, 'CDC平台'),
            'execution_machine': read_cell(row, '执行机器'),
            'viu_code': read_cell(row, 'VIU编号').lower(),
            'module': read_cell(row, '模块'),
            'case_no': read_cell(row, '用例编号'),
            'case_name': read_cell(row, '用例名称'),
            'remark': read_cell(row, '备注'),
        })
    if not parsed_rows:
        raise HttpError(400, 'Excel 中没有可导入的数据行')
    return parsed_rows


def _validate_full_import_rows(rows: list[dict], domain: str) -> dict[int, list[str]]:
    """校验必填字段及文件内配置冲突，冲突涉及的所有行都标记失败。"""
    errors: dict[int, list[str]] = defaultdict(list)
    platform_groups = defaultdict(list)
    vehicle_groups = defaultdict(list)
    case_groups = defaultdict(list)

    for row in rows:
        row_no = row['row_no']
        required_fields = {
            '版本名称': row['platform_name'],
            '版本标识': row['version_code'],
            '车型名称': row['vehicle_name'],
            '车型编号': row['vehicle_code'],
            '执行机器': row['execution_machine'],
        }
        if domain != DOMAIN_VEHICLE:
            required_fields['CDC平台'] = row['cdc_platform']
        for label, value in required_fields.items():
            if not value:
                errors[row_no].append(f'{label}不能为空')

        has_case_no = bool(row['case_no'])
        has_case_name = bool(row['case_name'])
        if has_case_no != has_case_name:
            errors[row_no].append('用例编号和用例名称必须同时填写或同时留空')
        if domain == DOMAIN_VEHICLE:
            if row['viu_code'] and row['viu_code'] not in VIU_CODE_VALUES:
                errors[row_no].append(f'VIU编号仅支持: {", ".join(VIU_CODE_VALUES)}')
            if has_case_no and not row['viu_code']:
                errors[row_no].append('车控用例必须填写VIU编号')
        if domain == DOMAIN_COCKPIT_SOC and has_case_no and not row['module']:
            errors[row_no].append('座舱SOC用例必须填写模块')

        if row['version_code']:
            platform_groups[row['version_code']].append((row_no, row['platform_name']))
        if row['vehicle_code']:
            vehicle_signature = (
                row['version_code'],
                row['vehicle_name'],
                row['cdc_platform'] if domain != DOMAIN_VEHICLE else '',
                row['execution_machine'],
            )
            vehicle_groups[row['vehicle_code']].append((row_no, vehicle_signature))
        if has_case_no and has_case_name and row['vehicle_code']:
            viu_code = row['viu_code'] if domain == DOMAIN_VEHICLE else ''
            case_key = (row['vehicle_code'], viu_code, row['case_no'])
            case_signature = (
                row['case_name'],
                row['module'] if domain == DOMAIN_COCKPIT_SOC else '',
                row['remark'],
            )
            case_groups[case_key].append((row_no, case_signature))

    for version_code, values in platform_groups.items():
        if len({signature for _, signature in values}) > 1:
            for row_no, _ in values:
                errors[row_no].append(f'版本标识 {version_code} 在文件内对应多个版本名称')
    for vehicle_code, values in vehicle_groups.items():
        if len({signature for _, signature in values}) > 1:
            for row_no, _ in values:
                errors[row_no].append(f'车型编号 {vehicle_code} 在文件内配置不一致')
    for case_key, values in case_groups.items():
        if len({signature for _, signature in values}) > 1:
            for row_no, _ in values:
                errors[row_no].append(f'用例 {case_key[2]} 在文件内配置不一致')
    return errors


def _upsert_full_import_platform(user, domain: str, row: dict):
    """按版本标识新增、恢复或更新平台，返回实例与处理状态。"""
    platform = McuPlatform.objects.filter(version_code=row['version_code']).first()
    if platform and platform.domain != domain:
        raise ValueError(f'版本标识 {row["version_code"]} 已属于其他领域')
    name_conflict = McuPlatform.objects.filter(name=row['platform_name'])
    if platform:
        name_conflict = name_conflict.exclude(id=platform.id)
    if name_conflict.exists():
        raise ValueError(f'版本名称 {row["platform_name"]} 已被其他版本使用')

    if not platform:
        platform = McuPlatform(
            name=row['platform_name'],
            version_code=row['version_code'],
            domain=domain,
            is_active=True,
        )
        _apply_audit_fields(platform, user, is_create=True)
        platform.save()
        return platform, 'created'

    changed = False
    if platform.name != row['platform_name']:
        platform.name = row['platform_name']
        changed = True
    if platform.is_deleted:
        platform.is_deleted = False
        platform.is_active = True
        changed = True
    if changed:
        _apply_audit_fields(platform, user)
        platform.save()
        return platform, 'updated'
    return platform, 'ignored'


def _upsert_full_import_vehicle(user, domain: str, platform: McuPlatform, row: dict):
    """按车型编号新增、恢复或更新车型，并增量合并车控 VIU 配置。"""
    vehicle = VehicleModel.objects.filter(vehicle_code=row['vehicle_code']).first()
    if vehicle and vehicle.platform.domain != domain:
        raise ValueError(f'车型编号 {row["vehicle_code"]} 已属于其他领域')
    name_conflict = VehicleModel.objects.filter(platform=platform, name=row['vehicle_name'])
    if vehicle:
        name_conflict = name_conflict.exclude(id=vehicle.id)
    if name_conflict.exists():
        raise ValueError(f'车型名称 {row["vehicle_name"]} 在该版本下已存在')

    viu_codes = list(vehicle.viu_codes or []) if vehicle else []
    if domain == DOMAIN_VEHICLE and row['viu_code'] and row['viu_code'] not in viu_codes:
        viu_codes.append(row['viu_code'])
    cdc_platform = row['cdc_platform'] if domain != DOMAIN_VEHICLE else ''
    if not vehicle:
        vehicle = VehicleModel(
            platform=platform,
            name=row['vehicle_name'],
            vehicle_code=row['vehicle_code'],
            cdc_platform=cdc_platform,
            execution_machine=row['execution_machine'],
            viu_codes=viu_codes,
            is_active=True,
        )
        _apply_audit_fields(vehicle, user, is_create=True)
        vehicle.save()
        return vehicle, 'created'

    changed = any([
        vehicle.platform_id != platform.id,
        vehicle.name != row['vehicle_name'],
        vehicle.cdc_platform != cdc_platform,
        vehicle.execution_machine != row['execution_machine'],
        list(vehicle.viu_codes or []) != viu_codes,
        vehicle.is_deleted,
    ])
    if changed:
        vehicle.platform = platform
        vehicle.name = row['vehicle_name']
        vehicle.cdc_platform = cdc_platform
        vehicle.execution_machine = row['execution_machine']
        vehicle.viu_codes = viu_codes
        if vehicle.is_deleted:
            vehicle.is_deleted = False
            vehicle.is_active = True
        _apply_audit_fields(vehicle, user)
        vehicle.save()
        return vehicle, 'updated'
    return vehicle, 'ignored'


def _upsert_full_import_case(user, domain: str, vehicle: VehicleModel, row: dict):
    """按领域用例唯一键新增、恢复或更新用例。"""
    viu_code = row['viu_code'] if domain == DOMAIN_VEHICLE else ''
    module = row['module'] if domain == DOMAIN_COCKPIT_SOC else ''
    remark = row['remark'] or None
    instance = TestCase.objects.filter(
        vehicle=vehicle,
        viu_code=viu_code,
        case_no=row['case_no'],
    ).first()
    if not instance:
        instance = TestCase(
            vehicle=vehicle,
            viu_code=viu_code,
            module=module,
            case_no=row['case_no'],
            case_name=row['case_name'],
            remark=remark,
            is_active=True,
        )
        _apply_audit_fields(instance, user, is_create=True)
        instance.save()
        return 'created'

    changed = any([
        instance.case_name != row['case_name'],
        instance.module != module,
        instance.remark != remark,
        instance.is_deleted,
    ])
    if changed:
        instance.case_name = row['case_name']
        instance.module = module
        instance.remark = remark
        if instance.is_deleted:
            instance.is_deleted = False
            instance.is_active = True
        _apply_audit_fields(instance, user)
        instance.save()
        return 'updated'
    return 'ignored'


def import_full_test_case_excel(user, domain: str, file_obj) -> ImportResultOut:
    """从一个 Excel 按行批量导入平台、车型及可选用例。"""
    parsed_domain = _parse_domain_filter(domain)
    if not parsed_domain:
        raise HttpError(422, '领域不能为空')
    rows = parse_full_test_case_excel_rows(file_obj, parsed_domain)
    validation_errors = _validate_full_import_rows(rows, parsed_domain)
    platform_created_ids = set()
    platform_updated_ids = set()
    vehicle_created_ids = set()
    vehicle_updated_ids = set()
    case_counts = Counter()
    configuration_row_count = 0

    for row in rows:
        row_no = row['row_no']
        if validation_errors.get(row_no):
            continue
        try:
            # 每行独立保存点，避免用例失败后留下半条平台或车型配置。
            with transaction.atomic():
                platform, platform_status = _upsert_full_import_platform(
                    user, parsed_domain, row
                )
                vehicle, vehicle_status = _upsert_full_import_vehicle(
                    user, parsed_domain, platform, row
                )
                case_status = None
                if row['case_no']:
                    case_status = _upsert_full_import_case(
                        user, parsed_domain, vehicle, row
                    )
        except (IntegrityError, ValueError) as exc:
            validation_errors[row_no].append(str(exc))
            continue

        if platform_status == 'created':
            platform_created_ids.add(platform.id)
        elif platform_status == 'updated' and platform.id not in platform_created_ids:
            platform_updated_ids.add(platform.id)
        if vehicle_status == 'created':
            vehicle_created_ids.add(vehicle.id)
        elif vehicle_status == 'updated' and vehicle.id not in vehicle_created_ids:
            vehicle_updated_ids.add(vehicle.id)
        if case_status:
            case_counts[case_status] += 1
        else:
            configuration_row_count += 1

    error_rows = [
        ImportErrorRow(row_no=row_no, message='；'.join(messages))
        for row_no, messages in sorted(validation_errors.items())
        if messages
    ]
    return ImportResultOut(
        created_count=case_counts['created'],
        updated_count=case_counts['updated'],
        ignored_count=case_counts['ignored'],
        platform_created_count=len(platform_created_ids),
        platform_updated_count=len(platform_updated_ids),
        vehicle_created_count=len(vehicle_created_ids),
        vehicle_updated_count=len(vehicle_updated_ids),
        configuration_row_count=configuration_row_count,
        errors=error_rows,
    )


def build_test_case_template_response(domain: Optional[str] = None):
    """生成可一次导入多个平台、车型及用例的一站式模板。"""
    parsed_domain = _parse_domain_filter(domain) or DOMAIN_COCKPIT
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = '平台车型用例'
    header = ['版本名称', '版本标识', '车型名称', '车型编号']
    if parsed_domain == DOMAIN_VEHICLE:
        header.extend(['执行机器', 'VIU编号', '用例编号', '用例名称', '备注'])
        example = [
            'VIU版本 2026.07', 'viu-2026.07', '示例车控车型', 'VEH-CTRL-001',
            '10.0.0.20', 'viu0', 'CASE-001', '示例车控用例', '固定备注示例',
        ]
    elif parsed_domain == DOMAIN_COCKPIT_SOC:
        header.extend(['CDC平台', '执行机器', '模块', '用例编号', '用例名称', '备注'])
        example = [
            'SOC版本 2026.07', 'soc-2026.07', '示例SOC车型', 'VEH-SOC-001',
            'CDC-SOC-01', '10.0.0.21', '座舱域控', 'CASE-001',
            '示例座舱SOC用例', '固定备注示例',
        ]
    else:
        header.extend(['CDC平台', '执行机器', '用例编号', '用例名称', '备注'])
        example = [
            'MCU版本 2026.07', 'mcu-2026.07', '示例MCU车型', 'VEH-MCU-001',
            'CDC-MCU-01', '10.0.0.22', 'CASE-001', '示例座舱MCU用例',
            '固定备注示例',
        ]
    sheet.append(header)
    sheet.append(example)
    sheet.freeze_panes = 'A2'
    for column in sheet.columns:
        letter = column[0].column_letter
        max_length = max(len(str(cell.value or '')) for cell in column)
        sheet.column_dimensions[letter].width = min(max(max_length + 2, 12), 32)
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = (
        f"attachment; filename*=UTF-8''auto_test_case_template_{parsed_domain}.xlsx"
    )
    workbook.save(response)
    return response


def build_test_case_export_response(filters):
    rows = list_test_cases(filters)
    parsed_domain = _parse_domain_filter(getattr(filters, 'domain', None))
    include_viu_code = parsed_domain == DOMAIN_VEHICLE or any(
        (item.get('viu_code') or '').strip() for item in rows
    )
    include_module = parsed_domain == DOMAIN_COCKPIT_SOC or any(
        (item.get('module') or '').strip() for item in rows
    )
    workbook = openpyxl.Workbook(write_only=True)
    sheet = workbook.create_sheet('测试用例')
    header = [
        '平台',
        '车型',
        '车型编号',
    ]
    if include_viu_code:
        header.append('VIU编号')
    if include_module:
        header.append('模块')
    header.extend([
        '用例编号',
        '用例名称',
        '备注',
        '最近执行时间',
        '更新时间',
    ])
    sheet.append(header)
    for item in rows:
        row = [
            item['platform_name'],
            item['vehicle_name'],
            item['vehicle_code'],
        ]
        if include_viu_code:
            row.append(item.get('viu_code') or '')
        if include_module:
            row.append(item.get('module') or '')
        row.extend([
            item['case_no'],
            item['case_name'],
            item.get('remark') or '',
            item.get('latest_execute_time') or '',
            item.get('sys_update_datetime') or '',
        ])
        sheet.append(row)
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = (
        f"attachment; filename*=UTF-8''auto_test_cases_{(parsed_domain or 'all')}.xlsx"
    )
    workbook.save(response)
    return response


@transaction.atomic
def report_daily_results(payload):
    vehicle = VehicleModel.objects.filter(
        vehicle_code=payload.vehicle_code,
        is_deleted=False,
        is_active=True,
    ).first()
    if not vehicle:
        raise HttpError(404, '车型不存在')

    vehicle_domain = vehicle.platform.domain
    allowed_viu_codes = {
        str(code or '').strip().lower()
        for code in (vehicle.viu_codes or [])
        if str(code or '').strip()
    }
    active_cases = {
        (
            (item.viu_code or '').strip().lower(),
            str(item.case_no or '').strip(),
        ): item
        for item in TestCase.objects.filter(
            vehicle=vehicle,
            is_deleted=False,
            is_active=True,
        )
    }
    created_count = 0
    ignored_count = 0
    errors = []
    now = datetime.now()
    for index, item in enumerate(payload.results, start=1):
        case_no = (item.case_no or '').strip()
        result = (item.result or '').strip().lower()
        if not case_no:
            errors.append(ImportErrorRow(row_no=index, message='用例编号不能为空'))
            ignored_count += 1
            continue
        if result not in VALID_RESULT_VALUES:
            errors.append(
                ImportErrorRow(
                    row_no=index,
                    message=f'执行结果仅支持: {", ".join(sorted(VALID_RESULT_VALUES))}',
                ),
            )
            ignored_count += 1
            continue

        viu_code = (item.viu_code or '').strip().lower()
        if vehicle_domain == DOMAIN_VEHICLE:
            if not viu_code:
                errors.append(ImportErrorRow(row_no=index, message='车控领域上报结果需要填写VIU编号'))
                ignored_count += 1
                continue
            if viu_code not in allowed_viu_codes:
                errors.append(
                    ImportErrorRow(
                        row_no=index,
                        message=f'车型 {vehicle.name} 未配置 VIU 编号: {viu_code}',
                    ),
                )
                ignored_count += 1
                continue
        else:
            viu_code = ''

        test_case = active_cases.get((viu_code, case_no))
        if not test_case:
            lookup_key = f'{viu_code + " / " if viu_code else ""}{case_no}'
            errors.append(
                ImportErrorRow(
                    row_no=index,
                    message=f'未找到匹配用例: {lookup_key}',
                ),
            )
            ignored_count += 1
            continue
        DailyExecutionResult.objects.create(
            vehicle=vehicle,
            execute_date=payload.execute_date,
            test_case=test_case,
            start_time=item.start_time,
            duration_seconds=max(int(item.duration_seconds or 0), 0),
            result=result,
            failure_reason=None,
            log_url=(item.log_url or '').strip() or None,
        )
        created_count += 1

    if created_count > 0:
        recalculate_daily_batch(vehicle.id, payload.execute_date, now)
    return {
        'vehicle_id': str(vehicle.id),
        'execute_date': payload.execute_date,
        'created_count': created_count,
        'updated_count': created_count,
        'ignored_count': ignored_count,
        'errors': errors,
    }


def recalculate_daily_batch(vehicle_id: str, execute_date, last_report_at=None):
    vehicle = get_vehicle(vehicle_id)
    results = list(build_latest_daily_results_queryset(vehicle=vehicle, execute_date=execute_date))
    counter = Counter(item.result for item in results)
    total_duration_seconds = sum(max(item.duration_seconds or 0, 0) for item in results)
    total_count = len(results)
    skip_count = counter.get(RESULT_SKIP, 0)

    batch, _ = DailyExecutionBatch.objects.get_or_create(
        vehicle=vehicle,
        execute_date=execute_date,
    )
    batch.total_count = total_count
    batch.success_count = counter.get(RESULT_SUCCESS, 0)
    batch.failed_count = counter.get(RESULT_FAILED, 0)
    batch.timeout_count = counter.get(RESULT_TIMEOUT, 0)
    batch.skip_count = skip_count
    batch.total_duration_seconds = total_duration_seconds
    if last_report_at is not None:
        batch.last_report_at = last_report_at
    batch.save()
    return batch


def get_suggested_failure_reason(vehicle_id: str, test_case_id: str, execute_date):
    """获取上一条人工维护的异常原因，辅助测试人员快速回填。"""
    item = (
        DailyExecutionResult.objects.filter(
            vehicle_id=vehicle_id,
            test_case_id=test_case_id,
            is_deleted=False,
            result__in=list(MANUAL_REASON_RESULTS),
        )
        .exclude(execute_date=execute_date)
        .exclude(failure_reason__isnull=True)
        .exclude(failure_reason__exact='')
        .order_by(
            '-execute_date',
            '-start_time',
            '-reported_at',
            '-sys_create_datetime',
        )
        .first()
    )
    return item.failure_reason if item else None


def _get_active_case_count(vehicle: VehicleModel) -> int:
    """统计车型下当前仍生效的用例数，用于判断是否存在缺失执行。"""
    return TestCase.objects.filter(
        vehicle=vehicle,
        is_deleted=False,
        is_active=True,
    ).count()


def _list_latest_active_results(vehicle: VehicleModel, execute_date):
    """获取车型当日活跃用例的最新执行结果。"""
    return list(
        build_latest_daily_results_queryset(vehicle=vehicle, execute_date=execute_date)
        .filter(test_case__is_active=True)
        .select_related('test_case')
    )


def _list_active_cases(vehicle: VehicleModel):
    """按看板展示顺序获取车型当前生效的注册用例。"""
    return list(
        TestCase.objects.filter(
            vehicle=vehicle,
            is_deleted=False,
            is_active=True,
        ).order_by('-sort', 'viu_code', 'case_no')
    )


def _build_result_category_counts(results: list[DailyExecutionResult]):
    """按根因分类统计当日最新非成功结果。"""
    version_count = 0
    non_version_count = 0
    uncategorized_count = 0
    for result in results:
        if result.result not in NON_SUCCESS_RESULTS:
            continue
        if result.failure_category == FAILURE_CATEGORY_VERSION:
            version_count += 1
        elif _is_non_version_failure(result):
            non_version_count += 1
        else:
            uncategorized_count += 1
    return {
        'version_failure_count': version_count,
        'non_version_failure_count': non_version_count,
        'uncategorized_failure_count': uncategorized_count,
    }


def _normalize_commit_id(value: Optional[str]) -> str:
    """标准化 CI 上报的 commit-id，当前只做去空白和必填校验。"""
    commit_id = (value or '').strip()
    if not commit_id:
        raise HttpError(422, 'commit-id 不能为空')
    return commit_id


def _serialize_downstream_commit(item: DownstreamCommit) -> DownstreamCommitOut:
    """序列化 commit-id 记录，供列表、选择器和上传接口复用。"""
    return DownstreamCommitOut(
        id=str(item.id),
        commit_id=item.commit_id,
        first_uploaded_at=item.first_uploaded_at,
        last_uploaded_at=item.last_uploaded_at,
        upload_count=item.upload_count,
        last_used_at=item.last_used_at,
        use_count=item.use_count,
    )


def _get_user_display_name(user) -> str:
    """获取触发人展示名，兼容后台用户模型的不同字段。"""
    if not user:
        return ''
    return getattr(user, 'name', None) or getattr(user, 'username', None) or str(user)


def _serialize_downstream_commit_usage(item: DownstreamCommitUsage) -> DownstreamCommitUsageOut:
    """序列化 commit-id 使用记录。"""
    return DownstreamCommitUsageOut(
        id=str(item.id),
        commit_id=item.commit.commit_id,
        execute_date=item.execute_date,
        trigger_type=item.trigger_type,
        trigger_user_name=_get_user_display_name(item.trigger_user),
        success=item.success,
        dry_run=item.dry_run,
        message=item.message,
        used_at=item.used_at,
    )


def report_downstream_commit(payload) -> DownstreamCommitOut:
    """接收 CI 上报的 commit-id，重复上传只更新计数和最近上传信息。"""
    commit_id = _normalize_commit_id(payload.commit_id)
    now = timezone.now()
    defaults = {
        'first_uploaded_at': now,
        'last_uploaded_at': now,
        'upload_count': 1,
    }
    with transaction.atomic():
        item, created = DownstreamCommit.objects.select_for_update().get_or_create(
            commit_id=commit_id,
            defaults=defaults,
        )
        if not created:
            item.last_uploaded_at = now
            item.upload_count = max(item.upload_count or 0, 0) + 1
            item.save(
                update_fields=[
                    'last_uploaded_at',
                    'upload_count',
                    'sys_update_datetime',
                ],
            )
    return _serialize_downstream_commit(item)


def _parse_commit_uploaded_bound(value: str, *, end_of_day: bool = False):
    """解析 commit-id 上传时间筛选边界，支持前端传日期或日期时间字符串。"""
    raw_value = (value or '').strip()
    if not raw_value:
        return None
    try:
        parsed = datetime.fromisoformat(raw_value.replace('Z', '+00:00'))
    except ValueError:
        raise HttpError(422, '上传时间筛选格式不正确')
    if len(raw_value) <= 10 and end_of_day:
        parsed = parsed.replace(hour=23, minute=59, second=59, microsecond=999999)
    if settings.USE_TZ and parsed.tzinfo is None:
        parsed = timezone.make_aware(parsed, timezone.get_current_timezone())
    if not settings.USE_TZ and parsed.tzinfo is not None:
        parsed = timezone.make_naive(parsed, timezone.get_current_timezone())
    return parsed


def list_downstream_commits(
    *,
    keyword: str = '',
    uploaded_start: str = '',
    uploaded_end: str = '',
    page: int = 1,
    page_size: int = 20,
):
    """分页查询 commit-id 历史，支持按 commit-id 和上传时间搜索。"""
    safe_page = max(int(page or 1), 1)
    safe_size = min(max(int(page_size or 20), 1), 100)
    queryset = DownstreamCommit.objects.filter(is_deleted=False)
    keyword_value = (keyword or '').strip()
    if keyword_value:
        queryset = queryset.filter(commit_id__icontains=keyword_value)
    start_time = _parse_commit_uploaded_bound(uploaded_start)
    end_time = _parse_commit_uploaded_bound(uploaded_end, end_of_day=True)
    if start_time:
        queryset = queryset.filter(last_uploaded_at__gte=start_time)
    if end_time:
        queryset = queryset.filter(last_uploaded_at__lte=end_time)
    total = queryset.count()
    start = (safe_page - 1) * safe_size
    items = [
        _serialize_downstream_commit(item)
        for item in queryset.order_by('-last_uploaded_at', '-sys_create_datetime')[start:start + safe_size]
    ]
    return {'items': items, 'total': total, 'page': safe_page, 'page_size': safe_size}


def list_downstream_commit_usages(commit_record_id: str, *, page: int = 1, page_size: int = 10):
    """分页查询某个 commit-id 的下游触发使用记录。"""
    commit = DownstreamCommit.objects.filter(id=commit_record_id, is_deleted=False).first()
    if not commit:
        raise HttpError(404, 'commit-id 记录不存在')
    safe_page = max(int(page or 1), 1)
    safe_size = min(max(int(page_size or 10), 1), 100)
    queryset = DownstreamCommitUsage.objects.select_related('commit', 'trigger_user').filter(
        commit=commit,
        is_deleted=False,
    )
    total = queryset.count()
    start = (safe_page - 1) * safe_size
    items = [
        _serialize_downstream_commit_usage(item)
        for item in queryset.order_by('-used_at', '-sys_create_datetime')[start:start + safe_size]
    ]
    return {'items': items, 'total': total, 'page': safe_page, 'page_size': safe_size}


def _get_downstream_commit_by_commit_id(commit_id: str) -> DownstreamCommit:
    """按 commit-id 获取可用记录，人工触发必须选择已上报的 commit-id。"""
    normalized = _normalize_commit_id(commit_id)
    item = DownstreamCommit.objects.filter(commit_id=normalized, is_deleted=False).first()
    if not item:
        raise HttpError(404, 'commit-id 记录不存在')
    return item


def _get_latest_unused_downstream_commit() -> Optional[DownstreamCommit]:
    """定时任务使用最新且未使用过的 commit-id，避免重复触发同一提交。"""
    return (
        DownstreamCommit.objects.filter(is_deleted=False, use_count=0)
        .order_by('-last_uploaded_at', '-sys_create_datetime')
        .first()
    )


def _record_downstream_commit_usage(
    commit: DownstreamCommit,
    *,
    user,
    execute_date,
    trigger_type: str,
    success: bool,
    dry_run: bool,
    message: str,
) -> DownstreamCommitUsage:
    """记录每次下游触发尝试，并同步 commit-id 的最近使用信息。"""
    now = timezone.now()
    with transaction.atomic():
        usage = DownstreamCommitUsage.objects.create(
            commit=commit,
            execute_date=execute_date,
            trigger_type=trigger_type,
            trigger_user=user if user and getattr(user, 'is_authenticated', True) else None,
            success=success,
            dry_run=dry_run,
            message=(message or '').strip() or None,
            used_at=now,
        )
        commit.last_used_at = now
        commit.use_count = max(commit.use_count or 0, 0) + 1
        commit.save(update_fields=['last_used_at', 'use_count', 'sys_update_datetime'])
    return usage


def build_cockpit_downstream_gate(execute_date):
    """构建座舱下游任务人工触发门禁，供 API、看板和定时任务复用。"""
    vehicles = VehicleModel.objects.select_related('platform').filter(
        is_deleted=False,
        is_active=True,
        platform__is_deleted=False,
        platform__is_active=True,
        platform__domain=DOMAIN_COCKPIT,
    ).order_by('platform__sort', 'name')

    vehicle_count = 0
    total_case_count = 0
    success_count = 0
    failed_count = 0
    timeout_count = 0
    skip_count = 0
    version_failure_count = 0
    non_version_failure_count = 0
    uncategorized_failure_count = 0
    missing_result_count = 0

    for vehicle in vehicles:
        vehicle_count += 1
        active_case_count = _get_active_case_count(vehicle)
        latest_results = _list_latest_active_results(vehicle, execute_date)
        counter = Counter(result.result for result in latest_results)
        category_counts = _build_result_category_counts(latest_results)

        total_case_count += active_case_count
        success_count += counter.get(RESULT_SUCCESS, 0)
        failed_count += counter.get(RESULT_FAILED, 0)
        timeout_count += counter.get(RESULT_TIMEOUT, 0)
        skip_count += counter.get(RESULT_SKIP, 0)
        version_failure_count += category_counts['version_failure_count']
        non_version_failure_count += category_counts['non_version_failure_count']
        uncategorized_failure_count += category_counts['uncategorized_failure_count']
        missing_result_count += max(active_case_count - len(latest_results), 0)

    block_reasons = []
    if vehicle_count == 0:
        block_reasons.append('暂无可用座舱车型')
    if total_case_count == 0:
        block_reasons.append('暂无可用座舱用例')
    if missing_result_count > 0:
        block_reasons.append(f'还有 {missing_result_count} 条座舱用例缺少当日执行结果')
    if uncategorized_failure_count > 0:
        block_reasons.append(f'还有 {uncategorized_failure_count} 条非成功用例未填写根因大类')
    if version_failure_count > 0:
        block_reasons.append(f'存在 {version_failure_count} 条版本问题用例')

    return {
        'execute_date': execute_date,
        'vehicle_count': vehicle_count,
        'total_case_count': total_case_count,
        'success_count': success_count,
        'failed_count': failed_count,
        'timeout_count': timeout_count,
        'skip_count': skip_count,
        'version_failure_count': version_failure_count,
        'non_version_failure_count': non_version_failure_count,
        'uncategorized_failure_count': uncategorized_failure_count,
        'missing_result_count': missing_result_count,
        'block_reasons': block_reasons,
        'enabled': not block_reasons,
    }


def invoke_cockpit_downstream_ci(
    gate: dict,
    *,
    trigger_type: str,
    commit: DownstreamCommit,
) -> DownstreamTriggerOut:
    """占位调用座舱下游 CI，生产环境可替换为真实 HTTP 请求实现。"""
    # TODO: 生产环境在这里对接真实 CI 接口，请求体必须携带 commit.commit_id。
    return DownstreamTriggerOut(
        triggered=True,
        dry_run=True,
        message=f'座舱下游任务已完成占位触发（{trigger_type}，commit-id: {commit.commit_id}）',
        execute_date=gate['execute_date'],
        commit_id=commit.commit_id,
        commit_record_id=str(commit.id),
        vehicle_count=gate['vehicle_count'],
        total_case_count=gate['total_case_count'],
        success_count=gate['success_count'],
        failed_count=gate['failed_count'],
        timeout_count=gate['timeout_count'],
        skip_count=gate['skip_count'],
        non_version_failure_count=gate['non_version_failure_count'],
        version_failure_count=gate['version_failure_count'],
        uncategorized_failure_count=gate['uncategorized_failure_count'],
        missing_result_count=gate['missing_result_count'],
        block_reasons=[],
    )


def trigger_cockpit_downstream(user, execute_date, commit_id: str) -> DownstreamTriggerOut:
    """人工触发座舱下游任务，后端强制执行完整门禁校验。"""
    commit = _get_downstream_commit_by_commit_id(commit_id)
    gate = build_cockpit_downstream_gate(execute_date)
    if not gate['enabled']:
        message = '；'.join(gate['block_reasons'])
        _record_downstream_commit_usage(
            commit,
            user=user,
            execute_date=execute_date,
            trigger_type=DOWNSTREAM_TRIGGER_MANUAL,
            success=False,
            dry_run=True,
            message=message,
        )
        raise HttpError(400, message)
    try:
        result = invoke_cockpit_downstream_ci(
            gate,
            trigger_type=DOWNSTREAM_TRIGGER_MANUAL,
            commit=commit,
        )
    except Exception as exc:
        _record_downstream_commit_usage(
            commit,
            user=user,
            execute_date=execute_date,
            trigger_type=DOWNSTREAM_TRIGGER_MANUAL,
            success=False,
            dry_run=True,
            message=str(exc),
        )
        raise
    usage = _record_downstream_commit_usage(
        commit,
        user=user,
        execute_date=execute_date,
        trigger_type=DOWNSTREAM_TRIGGER_MANUAL,
        success=result.triggered,
        dry_run=result.dry_run,
        message=result.message,
    )
    result.usage_id = str(usage.id)
    return result


@scheduler_task
def run_scheduled_cockpit_downstream_check(date_offset: int = 0, **kwargs):
    """定时检查座舱每日执行结果，只有全部成功才自动触发占位 CI。"""
    offset = int(date_offset or 0)
    execute_date = datetime.now().date() - timedelta(days=offset)
    gate = build_cockpit_downstream_gate(execute_date)
    if gate['total_case_count'] <= 0:
        return f'{execute_date} 暂无座舱用例，跳过下游任务。'
    if (
        gate['missing_result_count'] > 0
        or gate['failed_count'] > 0
        or gate['timeout_count'] > 0
        or gate['skip_count'] > 0
    ):
        return (
            f"{execute_date} 座舱结果未全部成功，跳过下游任务："
            f"缺失 {gate['missing_result_count']}，失败 {gate['failed_count']}，"
            f"超时 {gate['timeout_count']}，跳过 {gate['skip_count']}。"
        )
    commit = _get_latest_unused_downstream_commit()
    if not commit:
        return f'{execute_date} 暂无未使用的 commit-id，跳过下游任务。'
    try:
        result = invoke_cockpit_downstream_ci(
            gate,
            trigger_type=DOWNSTREAM_TRIGGER_SCHEDULED,
            commit=commit,
        )
    except Exception as exc:
        _record_downstream_commit_usage(
            commit,
            user=None,
            execute_date=execute_date,
            trigger_type=DOWNSTREAM_TRIGGER_SCHEDULED,
            success=False,
            dry_run=True,
            message=str(exc),
        )
        raise
    usage = _record_downstream_commit_usage(
        commit,
        user=None,
        execute_date=execute_date,
        trigger_type=DOWNSTREAM_TRIGGER_SCHEDULED,
        success=result.triggered,
        dry_run=result.dry_run,
        message=result.message,
    )
    result.usage_id = str(usage.id)
    return result.message


def get_daily_summary(vehicle_id: str, execute_date, domain: Optional[str] = None) -> DailySummaryOut:
    """获取单车型每日执行汇总。"""
    vehicle = get_vehicle(vehicle_id)
    parsed_domain = _parse_domain_filter(domain)
    if parsed_domain and vehicle.platform.domain != parsed_domain:
        raise HttpError(404, '车型不存在')
    batch = recalculate_daily_batch(vehicle.id, execute_date)
    active_case_count = _get_active_case_count(vehicle)
    latest_results = _list_latest_active_results(vehicle, execute_date)
    missing_result_count = max(active_case_count - len(latest_results), 0)
    total = max(active_case_count, 0)

    def stat(key, count):
        ratio = round((count / total), 4) if total else 0
        return SummaryStat(key=key, label=RESULT_LABELS[key], count=count, ratio=ratio)

    return DailySummaryOut(
        vehicle_id=str(vehicle.id),
        vehicle_name=vehicle.name,
        vehicle_code=vehicle.vehicle_code,
        execute_date=execute_date,
        total_count=active_case_count,
        success_count=batch.success_count,
        failed_count=batch.failed_count,
        timeout_count=batch.timeout_count,
        skip_count=batch.skip_count,
        missing_result_count=missing_result_count,
        total_duration_seconds=batch.total_duration_seconds,
        stats=[
            stat(RESULT_SUCCESS, batch.success_count),
            stat(RESULT_FAILED, batch.failed_count),
            stat(RESULT_TIMEOUT, batch.timeout_count),
            stat(RESULT_SKIP, batch.skip_count),
            stat(RESULT_MISSING, missing_result_count),
        ],
        last_report_at=batch.last_report_at,
    )


def get_daily_overview(query) -> DailyOverviewResponse:
    """获取每日全量概览，并携带座舱下游触发门禁摘要。"""
    vehicles = VehicleModel.objects.select_related('platform').filter(
        is_deleted=False,
        is_active=True,
        platform__is_deleted=False,
        platform__is_active=True,
    )
    parsed_domain = _parse_domain_filter(getattr(query, 'domain', None))
    if parsed_domain:
        vehicles = vehicles.filter(platform__domain=parsed_domain)
    if query.platform_id:
        vehicles = vehicles.filter(platform_id=query.platform_id)

    rows: list[DailyOverviewRow] = []
    total_case_count = 0
    success_count = 0
    failed_count = 0
    timeout_count = 0
    skip_count = 0
    non_version_failure_count = 0
    version_failure_count = 0
    uncategorized_failure_count = 0
    missing_result_count = 0
    total_duration_seconds = 0
    abnormal_vehicle_count = 0
    latest_report_at = None

    for vehicle in vehicles:
        batch = recalculate_daily_batch(vehicle.id, query.execute_date)
        active_case_count = _get_active_case_count(vehicle)
        latest_results = _list_latest_active_results(vehicle, query.execute_date)
        is_abnormal = _is_daily_overview_abnormal(active_case_count, latest_results)
        if query.abnormal_only and not is_abnormal:
            continue

        category_counts = _build_result_category_counts(latest_results)
        vehicle_missing_result_count = max(active_case_count - len(latest_results), 0)

        row = DailyOverviewRow(
            vehicle_id=str(vehicle.id),
            vehicle_name=vehicle.name,
            vehicle_code=vehicle.vehicle_code,
            platform_id=str(vehicle.platform_id),
            platform_name=vehicle.platform.name,
            total_count=active_case_count,
            success_count=batch.success_count,
            failed_count=batch.failed_count,
            timeout_count=batch.timeout_count,
            skip_count=batch.skip_count,
            non_version_failure_count=category_counts['non_version_failure_count'],
            version_failure_count=category_counts['version_failure_count'],
            uncategorized_failure_count=category_counts['uncategorized_failure_count'],
            missing_result_count=vehicle_missing_result_count,
            total_duration_seconds=batch.total_duration_seconds,
            last_report_at=batch.last_report_at,
            is_abnormal=is_abnormal,
        )
        rows.append(row)
        total_case_count += active_case_count
        success_count += batch.success_count
        failed_count += batch.failed_count
        timeout_count += batch.timeout_count
        skip_count += batch.skip_count
        non_version_failure_count += category_counts['non_version_failure_count']
        version_failure_count += category_counts['version_failure_count']
        uncategorized_failure_count += category_counts['uncategorized_failure_count']
        missing_result_count += vehicle_missing_result_count
        total_duration_seconds += batch.total_duration_seconds
        abnormal_vehicle_count += int(is_abnormal)
        if batch.last_report_at and (
            latest_report_at is None or batch.last_report_at > latest_report_at
        ):
            latest_report_at = batch.last_report_at

    rows.sort(
        key=lambda item: (
            0 if item.is_abnormal else 1,
            -item.failed_count,
            -item.timeout_count,
            item.vehicle_name,
        )
    )

    visible_vehicle_count = len(rows)

    def stat(key, count):
        ratio = round((count / total_case_count), 4) if total_case_count else 0
        return SummaryStat(key=key, label=RESULT_LABELS[key], count=count, ratio=ratio)

    summary = DailyOverviewSummary(
        execute_date=query.execute_date,
        vehicle_count=visible_vehicle_count,
        abnormal_vehicle_count=abnormal_vehicle_count,
        total_case_count=total_case_count,
        success_count=success_count,
        failed_count=failed_count,
        timeout_count=timeout_count,
        skip_count=skip_count,
        non_version_failure_count=non_version_failure_count,
        version_failure_count=version_failure_count,
        uncategorized_failure_count=uncategorized_failure_count,
        missing_result_count=missing_result_count,
        downstream_trigger_enabled=False,
        downstream_trigger_block_reasons=[],
        total_duration_seconds=total_duration_seconds,
        stats=[
            stat(RESULT_SUCCESS, success_count),
            stat(RESULT_FAILED, failed_count),
            stat(RESULT_TIMEOUT, timeout_count),
            stat(RESULT_SKIP, skip_count),
            stat(RESULT_MISSING, missing_result_count),
        ],
        last_report_at=latest_report_at,
    )
    if parsed_domain == DOMAIN_COCKPIT:
        if query.platform_id:
            summary.downstream_trigger_enabled = False
            summary.downstream_trigger_block_reasons = ['触发下游任务前请先切换为全部平台']
        else:
            gate = build_cockpit_downstream_gate(query.execute_date)
            summary.downstream_trigger_enabled = gate['enabled']
            summary.downstream_trigger_block_reasons = gate['block_reasons']
    return DailyOverviewResponse(items=rows, summary=summary)


def list_daily_results(vehicle_id: str, execute_date, domain: Optional[str] = None):
    """获取车型每日明细，未上传结果的活跃用例会合成为未执行行。"""
    vehicle = get_vehicle(vehicle_id)
    parsed_domain = _parse_domain_filter(domain)
    if parsed_domain and vehicle.platform.domain != parsed_domain:
        raise HttpError(404, '车型不存在')
    active_cases = _list_active_cases(vehicle)
    latest_results = {
        str(result.test_case_id): result
        for result in _list_latest_active_results(vehicle, execute_date)
    }
    items = []
    for case in active_cases:
        result = latest_results.get(str(case.id))
        if result is None:
            # 未执行行只用于看板暴露漏执行用例，不对应数据库执行结果。
            items.append(
                DailyResultItemOut(
                    result_id=None,
                    case_id=str(case.id),
                    viu_code=case.viu_code,
                    module=case.module,
                    case_no=case.case_no,
                    case_name=case.case_name,
                    remark=case.remark,
                    status=RESULT_MISSING,
                    duration_seconds=0,
                )
            )
            continue
        items.append(
            DailyResultItemOut(
                result_id=str(result.id),
                case_id=str(case.id),
                viu_code=case.viu_code,
                module=case.module,
                case_no=case.case_no,
                case_name=case.case_name,
                remark=case.remark,
                status=result.result,
                failure_reason=result.failure_reason,
                failure_category=result.failure_category,
                suggested_failure_reason=(
                    get_suggested_failure_reason(str(vehicle.id), str(case.id), execute_date)
                    if result.result in MANUAL_REASON_RESULTS and not result.failure_reason
                    else None
                ),
                start_time=result.start_time,
                duration_seconds=result.duration_seconds,
                log_url=result.log_url,
                car_log_url=derive_car_log_url(result.log_url),
                reported_at=result.reported_at,
            )
        )
    return items


def update_daily_result_failure_reason(user, result_id: str, failure_reason: Optional[str], failure_category: Optional[str] = None):
    """更新非成功结果的异常原因与根因大类。"""
    instance = DailyExecutionResult.objects.filter(id=result_id, is_deleted=False).first()
    if not instance:
        raise HttpError(404, '执行结果不存在')
    if instance.result not in MANUAL_REASON_RESULTS:
        raise HttpError(400, '仅失败、超时或跳过结果支持填写异常原因')
    instance.failure_reason = (failure_reason or '').strip() or None
    instance.failure_category = _normalize_failure_category(failure_category)
    _apply_audit_fields(instance, user)
    instance.save()
    return True


def get_test_case_history(case_id: str, page: int = 1, page_size: int = 10) -> DailyHistoryPage:
    """获取测试用例历史执行记录。"""
    test_case = get_test_case(case_id)
    queryset = DailyExecutionResult.objects.filter(
        test_case=test_case,
        vehicle_id=test_case.vehicle_id,
        is_deleted=False,
    ).order_by(
        '-execute_date',
        '-start_time',
        '-reported_at',
        '-sys_create_datetime',
    )
    total = queryset.count()
    start = max(page - 1, 0) * page_size
    rows = [
        DailyHistoryRow(
            id=str(item.id),
            execute_date=item.execute_date,
            viu_code=item.test_case.viu_code,
            module=item.test_case.module,
            status=item.result,
            failure_reason=item.failure_reason,
            failure_category=item.failure_category,
            start_time=item.start_time,
            duration_seconds=item.duration_seconds,
            log_url=item.log_url,
            car_log_url=derive_car_log_url(item.log_url),
            reported_at=item.reported_at,
        )
        for item in queryset[start:start + page_size]
    ]
    return DailyHistoryPage(items=rows, total=total, page=page, page_size=page_size)
