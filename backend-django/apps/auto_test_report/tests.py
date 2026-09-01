import io
from datetime import date, timedelta
from unittest.mock import patch

import openpyxl
from django.test import TestCase
from django.utils import timezone
from core.user.user_model import User

from apps.auto_test_report import auto_test_report_services as services
from apps.auto_test_report.auto_test_report_model import (
    DailyExecutionResult,
    DOMAIN_COCKPIT_SOC,
    DownstreamCommitUsage,
    FAILURE_CATEGORY_CASE,
    FAILURE_CATEGORY_ENVIRONMENT,
    FAILURE_CATEGORY_NON_MCU,
    FAILURE_CATEGORY_VERSION,
    McuPlatform,
    TestCase as AutoTestCase,
    VehicleModel,
    DOMAIN_COCKPIT,
    DOMAIN_VEHICLE,
    DOMAIN_VEHICLE_IO,
    RESULT_FAILED,
    RESULT_SUCCESS,
    RESULT_SKIP,
    RESULT_TIMEOUT,
)
from apps.auto_test_report.auto_test_report_schemas import (
    DailyOverviewQuery,
    DownstreamCommitIn,
    ImportCasePayload,
    ImportCaseRow,
    ReportDailyResultsIn,
    ReportResultItemIn,
    TestCaseFilter,
    VehicleIn,
)


class AutoTestReportFullExcelImportTests(TestCase):
    MCU_HEADER = [
        '版本名称', '版本标识', '车型名称', '车型编号', 'CDC平台',
        '执行机器', '用例编号', '用例名称', '备注',
    ]
    SOC_HEADER = [
        '版本名称', '版本标识', '车型名称', '车型编号', 'CDC平台',
        '执行机器', '模块', '用例编号', '用例名称', '备注',
    ]
    VEHICLE_HEADER = [
        '版本名称', '版本标识', '车型名称', '车型编号', '执行机器',
        'VIU编号', '用例编号', '用例名称', '备注',
    ]

    def _build_excel(self, header, rows):
        """构造与下载模板一致的内存 Excel，供批量导入测试复用。"""
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(header)
        for row in rows:
            sheet.append(row)
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        return output

    def test_full_import_creates_multiple_platforms_vehicles_cases_and_config_row(self):
        file_obj = self._build_excel(self.MCU_HEADER, [
            ['MCU V1', 'mcu-v1', '车型A', 'MCU-A', 'CDC-A', 'host-a', 'CASE-1', '用例1', ''],
            ['MCU V1', 'mcu-v1', '车型B', 'MCU-B', 'CDC-B', 'host-b', 'CASE-2', '用例2', ''],
            ['MCU V2', 'mcu-v2', '车型C', 'MCU-C', 'CDC-C', 'host-c', '', '', ''],
        ])

        result = services.import_full_test_case_excel(None, DOMAIN_COCKPIT, file_obj)

        self.assertEqual(result.platform_created_count, 2)
        self.assertEqual(result.vehicle_created_count, 3)
        self.assertEqual(result.created_count, 2)
        self.assertEqual(result.configuration_row_count, 1)
        self.assertEqual(result.errors, [])
        self.assertEqual(McuPlatform.objects.filter(domain=DOMAIN_COCKPIT).count(), 2)
        self.assertEqual(VehicleModel.objects.count(), 3)
        self.assertEqual(AutoTestCase.objects.count(), 2)

    def test_full_import_updates_existing_records_without_removing_other_cases(self):
        first_file = self._build_excel(self.MCU_HEADER, [
            ['MCU V1', 'mcu-v1', '车型A', 'MCU-A', 'CDC-A', 'host-a', 'CASE-1', '用例1', '旧备注'],
            ['MCU V1', 'mcu-v1', '车型A', 'MCU-A', 'CDC-A', 'host-a', 'CASE-2', '保留用例', ''],
        ])
        services.import_full_test_case_excel(None, DOMAIN_COCKPIT, first_file)
        second_file = self._build_excel(self.MCU_HEADER, [
            ['MCU V1 新名称', 'mcu-v1', '车型A新名称', 'MCU-A', 'CDC-NEW', 'host-new', 'CASE-1', '用例1新名称', '新备注'],
        ])

        result = services.import_full_test_case_excel(None, DOMAIN_COCKPIT, second_file)

        self.assertEqual(result.platform_updated_count, 1)
        self.assertEqual(result.vehicle_updated_count, 1)
        self.assertEqual(result.updated_count, 1)
        self.assertEqual(AutoTestCase.objects.filter(is_deleted=False).count(), 2)
        vehicle = VehicleModel.objects.get(vehicle_code='MCU-A')
        self.assertEqual(vehicle.name, '车型A新名称')
        self.assertEqual(vehicle.cdc_platform, 'CDC-NEW')
        self.assertEqual(vehicle.execution_machine, 'host-new')
        case = AutoTestCase.objects.get(vehicle=vehicle, case_no='CASE-1')
        self.assertEqual(case.case_name, '用例1新名称')
        self.assertEqual(case.remark, '新备注')

    def test_soc_case_missing_module_reports_excel_row_and_writes_nothing(self):
        file_obj = self._build_excel(self.SOC_HEADER, [
            ['SOC V1', 'soc-v1', 'SOC车型', 'SOC-A', 'CDC-SOC', 'host-soc', '', 'CASE-1', 'SOC用例', ''],
        ])

        result = services.import_full_test_case_excel(None, DOMAIN_COCKPIT_SOC, file_obj)

        self.assertEqual(result.created_count, 0)
        self.assertEqual(result.errors[0].row_no, 2)
        self.assertIn('座舱SOC用例必须填写模块', result.errors[0].message)
        self.assertFalse(McuPlatform.objects.filter(version_code='soc-v1').exists())
        self.assertFalse(VehicleModel.objects.filter(vehicle_code='SOC-A').exists())

    def test_vehicle_import_does_not_require_cdc_and_merges_viu_codes(self):
        file_obj = self._build_excel(self.VEHICLE_HEADER, [
            ['VIU V1', 'viu-v1', '车控车型', 'CTRL-A', 'host-ctrl', '', '', '', ''],
            ['VIU V1', 'viu-v1', '车控车型', 'CTRL-A', 'host-ctrl', 'viu1', 'CASE-1', '车控用例', ''],
        ])

        result = services.import_full_test_case_excel(None, DOMAIN_VEHICLE, file_obj)

        self.assertEqual(result.vehicle_created_count, 1)
        self.assertEqual(result.configuration_row_count, 1)
        self.assertEqual(result.created_count, 1)
        vehicle = VehicleModel.objects.get(vehicle_code='CTRL-A')
        self.assertEqual(vehicle.cdc_platform, '')
        self.assertEqual(vehicle.viu_codes, ['viu1'])
        self.assertTrue(
            AutoTestCase.objects.filter(vehicle=vehicle, viu_code='viu1', case_no='CASE-1').exists()
        )

    def test_conflicting_vehicle_rows_are_all_rejected(self):
        file_obj = self._build_excel(self.MCU_HEADER, [
            ['MCU V1', 'mcu-v1', '车型A', 'MCU-A', 'CDC-A', 'host-a', 'CASE-1', '用例1', ''],
            ['MCU V1', 'mcu-v1', '车型A', 'MCU-A', 'CDC-A', 'host-b', 'CASE-2', '用例2', ''],
        ])

        result = services.import_full_test_case_excel(None, DOMAIN_COCKPIT, file_obj)

        self.assertEqual([item.row_no for item in result.errors], [2, 3])
        self.assertTrue(all('配置不一致' in item.message for item in result.errors))
        self.assertFalse(McuPlatform.objects.filter(version_code='mcu-v1').exists())

    def test_download_templates_have_domain_specific_hierarchy_columns(self):
        mcu_response = services.build_test_case_template_response(DOMAIN_COCKPIT)
        soc_response = services.build_test_case_template_response(DOMAIN_COCKPIT_SOC)
        vehicle_response = services.build_test_case_template_response(DOMAIN_VEHICLE)
        vehicle_io_response = services.build_test_case_template_response(DOMAIN_VEHICLE_IO)

        def header_of(response):
            workbook = openpyxl.load_workbook(io.BytesIO(response.content), read_only=True)
            return [cell.value for cell in next(workbook.active.iter_rows())]

        self.assertIn('CDC平台', header_of(mcu_response))
        self.assertIn('模块', header_of(soc_response))
        self.assertNotIn('CDC平台', header_of(vehicle_response))
        self.assertIn('VIU编号', header_of(vehicle_response))
        self.assertNotIn('CDC平台', header_of(vehicle_io_response))
        self.assertIn('VIU编号', header_of(vehicle_io_response))

    def test_vehicle_io_import_reuses_vehicle_viu_structure(self):
        file_obj = self._build_excel(self.VEHICLE_HEADER, [
            ['IO V1', 'io-v1', '车控IO车型', 'IO-A', 'host-io', 'viu2', 'CASE-1', 'IO用例', ''],
        ])

        result = services.import_full_test_case_excel(None, DOMAIN_VEHICLE_IO, file_obj)

        self.assertEqual(result.errors, [])
        vehicle = VehicleModel.objects.get(vehicle_code='IO-A')
        self.assertEqual(vehicle.platform.domain, DOMAIN_VEHICLE_IO)
        self.assertEqual(vehicle.cdc_platform, '')
        self.assertEqual(vehicle.viu_codes, ['viu2'])
        self.assertTrue(
            AutoTestCase.objects.filter(vehicle=vehicle, viu_code='viu2', case_no='CASE-1').exists()
        )


class AutoTestReportOverviewTests(TestCase):
    def setUp(self) -> None:
        self.execute_date = date(2026, 4, 26)
        self.platform = McuPlatform.objects.create(
            name='Test Platform',
            version_code='test-platform',
            domain=DOMAIN_COCKPIT,
            is_active=True,
        )

    def _create_vehicle(
        self,
        suffix: str,
        *,
        domain: str = DOMAIN_COCKPIT,
        viu_codes: list[str] | None = None,
    ) -> VehicleModel:
        platform = self.platform
        if platform.domain != domain:
            platform = McuPlatform.objects.create(
                name=f'Platform {suffix}',
                version_code=f'platform-{suffix}',
                domain=domain,
                is_active=True,
            )
        return VehicleModel.objects.create(
            platform=platform,
            name=f'Vehicle {suffix}',
            vehicle_code=f'VEH-{suffix}',
            cdc_platform='CDC',
            execution_machine=f'machine-{suffix}',
            viu_codes=viu_codes or [],
            is_active=True,
        )

    def _create_cases(self, vehicle: VehicleModel, count: int) -> list[AutoTestCase]:
        return [
            AutoTestCase.objects.create(
                vehicle=vehicle,
                module='',
                case_no=f'CASE-{vehicle.vehicle_code}-{index + 1}',
                case_name=f'Case {index + 1}',
                is_active=True,
            )
            for index in range(count)
        ]

    def _report_result(
        self,
        vehicle: VehicleModel,
        test_case: AutoTestCase,
        result: str,
        *,
        minutes_offset: int = 0,
        failure_category: str | None = None,
        log_url: str | None = None,
    ) -> DailyExecutionResult:
        return DailyExecutionResult.objects.create(
            vehicle=vehicle,
            test_case=test_case,
            execute_date=self.execute_date,
            start_time=timezone.now() + timedelta(minutes=minutes_offset),
            duration_seconds=60,
            result=result,
            failure_category=failure_category,
            log_url=log_url,
        )

    def _get_row(self, overview, vehicle: VehicleModel):
        return next(item for item in overview.items if item.vehicle_id == str(vehicle.id))

    def _build_overview(self, *, abnormal_only: bool = False):
        return services.get_daily_overview(
            DailyOverviewQuery(
                execute_date=self.execute_date,
                abnormal_only=abnormal_only,
            )
        )

    def test_overview_marks_all_success_uploaded_vehicle_as_normal(self):
        vehicle = self._create_vehicle('success')
        cases = self._create_cases(vehicle, 2)

        for index, case in enumerate(cases):
            self._report_result(vehicle, case, RESULT_SUCCESS, minutes_offset=index)

        overview = self._build_overview()
        row = self._get_row(overview, vehicle)

        self.assertFalse(row.is_abnormal)
        self.assertEqual(row.success_count, 2)
        self.assertEqual(row.skip_count, 0)
        self.assertEqual(overview.summary.abnormal_vehicle_count, 0)

    def test_overview_marks_failed_vehicle_as_abnormal(self):
        vehicle = self._create_vehicle('failed')
        cases = self._create_cases(vehicle, 2)
        self._report_result(vehicle, cases[0], RESULT_SUCCESS)
        self._report_result(vehicle, cases[1], RESULT_FAILED, minutes_offset=1)

        overview = self._build_overview()
        row = self._get_row(overview, vehicle)

        self.assertTrue(row.is_abnormal)
        self.assertEqual(row.failed_count, 1)
        self.assertEqual(overview.summary.abnormal_vehicle_count, 1)

    def test_overview_marks_timeout_vehicle_as_abnormal(self):
        vehicle = self._create_vehicle('timeout')
        cases = self._create_cases(vehicle, 2)
        self._report_result(vehicle, cases[0], RESULT_SUCCESS)
        self._report_result(vehicle, cases[1], RESULT_TIMEOUT, minutes_offset=1)

        overview = self._build_overview()
        row = self._get_row(overview, vehicle)

        self.assertTrue(row.is_abnormal)
        self.assertEqual(row.timeout_count, 1)
        self.assertEqual(overview.summary.abnormal_vehicle_count, 1)

    def test_overview_marks_partially_uploaded_vehicle_as_abnormal(self):
        vehicle = self._create_vehicle('partial')
        cases = self._create_cases(vehicle, 2)
        self._report_result(vehicle, cases[0], RESULT_SUCCESS)

        overview = self._build_overview()
        row = self._get_row(overview, vehicle)

        self.assertTrue(row.is_abnormal)
        self.assertEqual(row.success_count, 1)
        self.assertEqual(row.skip_count, 0)
        self.assertEqual(row.total_count, 2)
        self.assertEqual(row.missing_result_count, 1)
        self.assertEqual(overview.summary.total_case_count, 2)
        self.assertEqual(overview.summary.missing_result_count, 1)
        self.assertEqual(overview.summary.abnormal_vehicle_count, 1)

        summary = services.get_daily_summary(vehicle.id, self.execute_date, DOMAIN_COCKPIT)
        self.assertEqual(summary.total_count, 2)
        self.assertEqual(summary.success_count, 1)
        self.assertEqual(summary.missing_result_count, 1)
        self.assertEqual(
            {item.key: item.count for item in summary.stats},
            {
                'success': 1,
                'failed': 0,
                'timeout': 0,
                'skip': 0,
                'missing': 1,
            },
        )

    def test_overview_marks_no_upload_vehicle_as_abnormal(self):
        vehicle = self._create_vehicle('no-upload')
        self._create_cases(vehicle, 2)

        overview = self._build_overview()
        row = self._get_row(overview, vehicle)

        self.assertTrue(row.is_abnormal)
        self.assertEqual(row.success_count, 0)
        self.assertEqual(row.skip_count, 0)
        self.assertEqual(row.total_count, 2)
        self.assertEqual(row.missing_result_count, 2)
        self.assertEqual(overview.summary.total_case_count, 2)
        self.assertEqual(overview.summary.missing_result_count, 2)
        self.assertEqual(overview.summary.abnormal_vehicle_count, 1)

        items = services.list_daily_results(vehicle.id, self.execute_date, DOMAIN_COCKPIT)
        self.assertEqual(len(items), 2)
        self.assertEqual({item.status for item in items}, {'missing'})
        self.assertTrue(all(item.result_id is None for item in items))
        self.assertTrue(all(item.log_url is None and item.car_log_url is None for item in items))

    def test_overview_abnormal_only_includes_skip_missing_partial_and_failed_vehicles(self):
        success_vehicle = self._create_vehicle('all-success')
        skip_vehicle = self._create_vehicle('skip-only')
        no_upload_vehicle = self._create_vehicle('no-upload-only')
        partial_vehicle = self._create_vehicle('partial-only')
        failed_vehicle = self._create_vehicle('failed-only')

        success_cases = self._create_cases(success_vehicle, 2)
        skip_cases = self._create_cases(skip_vehicle, 2)
        self._create_cases(no_upload_vehicle, 2)
        partial_cases = self._create_cases(partial_vehicle, 2)
        failed_cases = self._create_cases(failed_vehicle, 2)

        for index, case in enumerate(success_cases):
            self._report_result(success_vehicle, case, RESULT_SUCCESS, minutes_offset=index)

        self._report_result(skip_vehicle, skip_cases[0], RESULT_SUCCESS)
        self._report_result(skip_vehicle, skip_cases[1], RESULT_SKIP, minutes_offset=1)
        self._report_result(partial_vehicle, partial_cases[0], RESULT_SUCCESS)
        self._report_result(failed_vehicle, failed_cases[0], RESULT_SUCCESS)
        self._report_result(failed_vehicle, failed_cases[1], RESULT_FAILED, minutes_offset=1)

        overview = self._build_overview(abnormal_only=True)
        vehicle_codes = {item.vehicle_code for item in overview.items}

        self.assertSetEqual(
            vehicle_codes,
            {
                skip_vehicle.vehicle_code,
                no_upload_vehicle.vehicle_code,
                partial_vehicle.vehicle_code,
                failed_vehicle.vehicle_code,
            },
        )
        self.assertEqual(overview.summary.vehicle_count, 4)
        self.assertEqual(overview.summary.abnormal_vehicle_count, 4)

    def test_derive_car_log_url_handles_valid_and_invalid_values(self):
        self.assertEqual(
            services.derive_car_log_url('https://example.com/logs/testcase.html'),
            'https://example.com/logs/',
        )
        self.assertIsNone(services.derive_car_log_url(''))
        self.assertIsNone(services.derive_car_log_url('   '))
        self.assertIsNone(services.derive_car_log_url('https://example.com/logs/index.html'))
        self.assertIsNone(services.derive_car_log_url('testcase.html'))

    def test_list_daily_results_derives_car_log_url_for_cockpit_result(self):
        vehicle = self._create_vehicle('car-log')
        case = self._create_cases(vehicle, 1)[0]
        self._report_result(
            vehicle,
            case,
            RESULT_SUCCESS,
            log_url='https://example.com/artifacts/testcase.html',
        )

        items = services.list_daily_results(vehicle.id, self.execute_date, DOMAIN_COCKPIT)

        self.assertEqual(len(items), 1)
        self.assertEqual(items[0].log_url, 'https://example.com/artifacts/testcase.html')
        self.assertEqual(items[0].car_log_url, 'https://example.com/artifacts/')

    def test_failure_category_can_be_updated_for_all_non_success_results(self):
        vehicle = self._create_vehicle('failure-category')
        cases = self._create_cases(vehicle, 4)
        failed = self._report_result(vehicle, cases[0], RESULT_FAILED)
        timeout = self._report_result(vehicle, cases[1], RESULT_TIMEOUT, minutes_offset=1)
        skipped = self._report_result(vehicle, cases[2], RESULT_SKIP, minutes_offset=2)
        non_mcu = self._report_result(vehicle, cases[3], RESULT_FAILED, minutes_offset=3)

        services.update_daily_result_failure_reason(None, str(failed.id), '失败原因', FAILURE_CATEGORY_VERSION)
        services.update_daily_result_failure_reason(None, str(timeout.id), '环境不稳定', FAILURE_CATEGORY_ENVIRONMENT)
        services.update_daily_result_failure_reason(None, str(skipped.id), '用例需调整', FAILURE_CATEGORY_CASE)
        services.update_daily_result_failure_reason(None, str(non_mcu.id), '外部依赖异常', FAILURE_CATEGORY_NON_MCU)

        failed.refresh_from_db()
        timeout.refresh_from_db()
        skipped.refresh_from_db()
        non_mcu.refresh_from_db()
        self.assertEqual(failed.failure_category, FAILURE_CATEGORY_VERSION)
        self.assertEqual(timeout.failure_category, FAILURE_CATEGORY_ENVIRONMENT)
        self.assertEqual(skipped.failure_category, FAILURE_CATEGORY_CASE)
        self.assertEqual(non_mcu.failure_category, FAILURE_CATEGORY_NON_MCU)

    def test_failure_category_rejects_success_and_unknown_category(self):
        vehicle = self._create_vehicle('failure-category-reject')
        cases = self._create_cases(vehicle, 2)
        success = self._report_result(vehicle, cases[0], RESULT_SUCCESS)
        failed = self._report_result(vehicle, cases[1], RESULT_FAILED, minutes_offset=1)

        with self.assertRaisesMessage(Exception, '仅失败、超时或跳过结果支持填写异常原因'):
            services.update_daily_result_failure_reason(None, str(success.id), '', FAILURE_CATEGORY_VERSION)
        with self.assertRaisesMessage(Exception, '失败根因大类仅支持'):
            services.update_daily_result_failure_reason(None, str(failed.id), '', 'unknown')

    def test_cockpit_overview_counts_downstream_gate_fields(self):
        vehicle = self._create_vehicle('gate-counts')
        cases = self._create_cases(vehicle, 6)
        self._report_result(vehicle, cases[0], RESULT_SUCCESS)
        self._report_result(
            vehicle,
            cases[1],
            RESULT_FAILED,
            minutes_offset=1,
            failure_category=FAILURE_CATEGORY_VERSION,
        )
        self._report_result(
            vehicle,
            cases[2],
            RESULT_TIMEOUT,
            minutes_offset=2,
            failure_category=FAILURE_CATEGORY_ENVIRONMENT,
        )
        self._report_result(vehicle, cases[3], RESULT_SKIP, minutes_offset=3)
        self._report_result(
            vehicle,
            cases[4],
            RESULT_FAILED,
            minutes_offset=4,
            failure_category=FAILURE_CATEGORY_NON_MCU,
        )

        overview = self._build_overview()
        row = self._get_row(overview, vehicle)

        self.assertEqual(row.version_failure_count, 1)
        self.assertEqual(row.non_version_failure_count, 2)
        self.assertEqual(row.uncategorized_failure_count, 1)
        self.assertEqual(row.missing_result_count, 1)
        self.assertFalse(overview.summary.downstream_trigger_enabled)
        self.assertEqual(overview.summary.version_failure_count, 1)
        self.assertEqual(overview.summary.non_version_failure_count, 2)
        self.assertEqual(overview.summary.uncategorized_failure_count, 1)
        self.assertEqual(overview.summary.missing_result_count, 1)

    def test_downstream_commit_upload_deduplicates_and_list_supports_keyword(self):
        first = services.report_downstream_commit(
            DownstreamCommitIn(commit_id='abc123'),
        )
        second = services.report_downstream_commit(
            DownstreamCommitIn(commit_id='abc123'),
        )

        self.assertEqual(first.id, second.id)
        self.assertEqual(second.upload_count, 2)

        page = services.list_downstream_commits(keyword='abc', page=1, page_size=20)
        self.assertEqual(page['total'], 1)
        self.assertEqual(page['items'][0].commit_id, 'abc123')

    def test_manual_downstream_trigger_rejects_unknown_commit(self):
        vehicle = self._create_vehicle('manual-unknown-commit')
        cases = self._create_cases(vehicle, 1)
        self._report_result(vehicle, cases[0], RESULT_SUCCESS)

        with self.assertRaisesMessage(Exception, 'commit-id 记录不存在'):
            services.trigger_cockpit_downstream(None, self.execute_date, 'not-exists')

        self.assertEqual(DownstreamCommitUsage.objects.count(), 0)

    def test_manual_downstream_trigger_allows_non_version_failures_when_classified(self):
        vehicle = self._create_vehicle('manual-pass')
        cases = self._create_cases(vehicle, 3)
        commit = services.report_downstream_commit(DownstreamCommitIn(commit_id='commit-manual-pass'))
        self._report_result(vehicle, cases[0], RESULT_SUCCESS)
        self._report_result(
            vehicle,
            cases[1],
            RESULT_FAILED,
            minutes_offset=1,
            failure_category=FAILURE_CATEGORY_ENVIRONMENT,
        )
        self._report_result(
            vehicle,
            cases[2],
            RESULT_TIMEOUT,
            minutes_offset=2,
            failure_category=FAILURE_CATEGORY_NON_MCU,
        )

        result = services.trigger_cockpit_downstream(None, self.execute_date, commit.commit_id)

        self.assertTrue(result.triggered)
        self.assertTrue(result.dry_run)
        self.assertEqual(result.commit_id, commit.commit_id)
        self.assertIsNotNone(result.usage_id)
        self.assertEqual(result.non_version_failure_count, 2)
        self.assertEqual(DownstreamCommitUsage.objects.count(), 1)

    def test_manual_downstream_trigger_rejects_version_uncategorized_and_missing(self):
        vehicle = self._create_vehicle('manual-block')
        cases = self._create_cases(vehicle, 4)
        commit = services.report_downstream_commit(DownstreamCommitIn(commit_id='commit-manual-block'))
        self._report_result(vehicle, cases[0], RESULT_SUCCESS)
        self._report_result(
            vehicle,
            cases[1],
            RESULT_FAILED,
            minutes_offset=1,
            failure_category=FAILURE_CATEGORY_VERSION,
        )
        self._report_result(vehicle, cases[2], RESULT_TIMEOUT, minutes_offset=2)

        with self.assertRaises(Exception) as ctx:
            services.trigger_cockpit_downstream(None, self.execute_date, commit.commit_id)

        message = str(ctx.exception)
        self.assertIn('缺少当日执行结果', message)
        self.assertIn('未填写根因大类', message)
        self.assertIn('版本问题', message)
        usage = DownstreamCommitUsage.objects.get()
        self.assertFalse(usage.success)
        self.assertIn('版本问题', usage.message)

    def test_scheduled_downstream_check_triggers_only_when_all_cockpit_cases_success(self):
        vehicle = self._create_vehicle('scheduled-pass')
        cases = self._create_cases(vehicle, 2)
        today = date.today()
        commit = services.report_downstream_commit(DownstreamCommitIn(commit_id='commit-scheduled-pass'))
        for index, case in enumerate(cases):
            DailyExecutionResult.objects.create(
                vehicle=vehicle,
                test_case=case,
                execute_date=today,
                start_time=timezone.now() + timedelta(minutes=index),
                duration_seconds=60,
                result=RESULT_SUCCESS,
            )

        result = services.run_scheduled_cockpit_downstream_check.__wrapped__(date_offset=0)

        self.assertIn('座舱下游任务已完成占位触发', result)
        self.assertIn(commit.commit_id, result)
        usage = DownstreamCommitUsage.objects.get()
        self.assertEqual(usage.trigger_type, 'scheduled')
        self.assertTrue(usage.success)

    def test_scheduled_downstream_check_skips_when_any_cockpit_case_is_not_success(self):
        vehicle = self._create_vehicle('scheduled-skip')
        cases = self._create_cases(vehicle, 2)
        today = date.today()
        DailyExecutionResult.objects.create(
            vehicle=vehicle,
            test_case=cases[0],
            execute_date=today,
            start_time=timezone.now(),
            duration_seconds=60,
            result=RESULT_SUCCESS,
        )
        DailyExecutionResult.objects.create(
            vehicle=vehicle,
            test_case=cases[1],
            execute_date=today,
            start_time=timezone.now() + timedelta(minutes=1),
            duration_seconds=60,
            result=RESULT_FAILED,
            failure_category=FAILURE_CATEGORY_ENVIRONMENT,
        )

        result = services.run_scheduled_cockpit_downstream_check.__wrapped__(date_offset=0)

        self.assertIn('未全部成功，跳过下游任务', result)
        self.assertEqual(DownstreamCommitUsage.objects.count(), 0)

    def test_scheduled_downstream_check_skips_without_unused_commit(self):
        vehicle = self._create_vehicle('scheduled-no-commit')
        cases = self._create_cases(vehicle, 1)
        today = date.today()
        DailyExecutionResult.objects.create(
            vehicle=vehicle,
            test_case=cases[0],
            execute_date=today,
            start_time=timezone.now(),
            duration_seconds=60,
            result=RESULT_SUCCESS,
        )

        result = services.run_scheduled_cockpit_downstream_check.__wrapped__(date_offset=0)

        self.assertIn('暂无未使用的 commit-id', result)

    def test_downstream_commit_usage_list_returns_trigger_history(self):
        vehicle = self._create_vehicle('manual-history')
        cases = self._create_cases(vehicle, 1)
        commit = services.report_downstream_commit(DownstreamCommitIn(commit_id='commit-history'))
        self._report_result(vehicle, cases[0], RESULT_SUCCESS)

        services.trigger_cockpit_downstream(None, self.execute_date, commit.commit_id)

        page = services.list_downstream_commit_usages(commit.id, page=1, page_size=10)
        self.assertEqual(page['total'], 1)
        self.assertEqual(page['items'][0].commit_id, commit.commit_id)
        self.assertEqual(page['items'][0].trigger_type, 'manual')
        self.assertTrue(page['items'][0].success)

    def test_vehicle_domain_report_supports_skip_and_row_errors(self):
        vehicle_platform = McuPlatform.objects.create(
            name='Vehicle Platform',
            version_code='vehicle-platform',
            domain=DOMAIN_VEHICLE,
            is_active=True,
        )
        vehicle = VehicleModel.objects.create(
            platform=vehicle_platform,
            name='Vehicle domain model',
            vehicle_code='VEH-VIU',
            cdc_platform='CDC',
            execution_machine='machine-viu',
            viu_codes=['viu0', 'viu1', 'viu2'],
            is_active=True,
        )
        case0 = AutoTestCase.objects.create(
            vehicle=vehicle,
            viu_code='viu0',
            case_no='CASE-001',
            case_name='Case 1',
            is_active=True,
        )
        case1 = AutoTestCase.objects.create(
            vehicle=vehicle,
            viu_code='viu1',
            case_no='CASE-001',
            case_name='Case 2',
            is_active=True,
        )
        AutoTestCase.objects.create(
            vehicle=vehicle,
            viu_code='viu0',
            case_no='CASE-002',
            case_name='Case 3',
            is_active=True,
        )

        payload = ReportDailyResultsIn(
            vehicle_code=vehicle.vehicle_code,
            execute_date=self.execute_date,
            results=[
                ReportResultItemIn(
                    viu_code='viu0',
                    case_no='CASE-001',
                    start_time=timezone.now(),
                    duration_seconds=60,
                    result=RESULT_SUCCESS,
                    log_url='https://example.com/viu0/testcase.html',
                ),
                ReportResultItemIn(
                    viu_code='viu1',
                    case_no='CASE-001',
                    start_time=timezone.now() + timedelta(minutes=1),
                    duration_seconds=0,
                    result=RESULT_SKIP,
                    log_url='https://example.com/viu1/testcase.html',
                ),
                ReportResultItemIn(
                    viu_code='viu2',
                    case_no='CASE-404',
                    start_time=timezone.now() + timedelta(minutes=2),
                    duration_seconds=30,
                    result=RESULT_SUCCESS,
                ),
            ],
        )

        result = services.report_daily_results(payload)
        self.assertEqual(result['created_count'], 2)
        self.assertEqual(result['ignored_count'], 1)
        self.assertEqual(len(result['errors']), 1)
        self.assertIn('未找到匹配用例', result['errors'][0].message)

        items = services.list_daily_results(
            vehicle.id,
            self.execute_date,
            DOMAIN_VEHICLE,
        )
        self.assertEqual(len(items), 3)
        self.assertEqual({item.viu_code for item in items}, {'viu0', 'viu1'})
        self.assertEqual(
            {(item.viu_code, item.case_no): item.status for item in items},
            {
                ('viu0', 'CASE-001'): RESULT_SUCCESS,
                ('viu1', 'CASE-001'): RESULT_SKIP,
                ('viu0', 'CASE-002'): 'missing',
            },
        )
        self.assertEqual(
            {
                (item.viu_code, item.case_no): item.car_log_url
                for item in items
                if item.car_log_url
            },
            {
                ('viu0', 'CASE-001'): 'https://example.com/viu0/',
                ('viu1', 'CASE-001'): 'https://example.com/viu1/',
            },
        )

        summary = services.get_daily_summary(
            vehicle.id,
            self.execute_date,
            DOMAIN_VEHICLE,
        )
        self.assertEqual(summary.total_count, 3)
        self.assertEqual(summary.success_count, 1)
        self.assertEqual(summary.skip_count, 1)
        self.assertEqual(summary.failed_count, 0)
        self.assertEqual(summary.missing_result_count, 1)

        overview = services.get_daily_overview(
            DailyOverviewQuery(
                execute_date=self.execute_date,
                domain=DOMAIN_VEHICLE,
            )
        )
        row = self._get_row(overview, vehicle)
        self.assertEqual(row.total_count, 3)
        self.assertEqual(row.success_count, 1)
        self.assertEqual(row.skip_count, 1)
        self.assertEqual(row.missing_result_count, 1)
        self.assertTrue(row.is_abnormal)

        self.assertEqual(case0.viu_code, 'viu0')
        self.assertEqual(case1.viu_code, 'viu1')

    def test_vehicle_io_daily_report_requires_and_matches_viu_code(self):
        platform = McuPlatform.objects.create(
            name='Vehicle IO Platform',
            version_code='vehicle-io-platform',
            domain=DOMAIN_VEHICLE_IO,
            is_active=True,
        )
        vehicle = VehicleModel.objects.create(
            platform=platform,
            name='Vehicle IO model',
            vehicle_code='VEH-IO',
            cdc_platform='',
            execution_machine='machine-io',
            viu_codes=['viu0'],
            is_active=True,
        )
        case = AutoTestCase.objects.create(
            vehicle=vehicle,
            viu_code='viu0',
            case_no='IO-CASE-001',
            case_name='IO Case',
            is_active=True,
        )

        result = services.report_daily_results(ReportDailyResultsIn(
            vehicle_code=vehicle.vehicle_code,
            execute_date=self.execute_date,
            results=[
                ReportResultItemIn(
                    viu_code='viu0',
                    case_no=case.case_no,
                    start_time=timezone.now(),
                    duration_seconds=30,
                    result=RESULT_SUCCESS,
                ),
            ],
        ))

        self.assertEqual(result['created_count'], 1)
        items = services.list_daily_results(vehicle.id, self.execute_date, DOMAIN_VEHICLE_IO)
        self.assertEqual(items[0].viu_code, 'viu0')

    def test_get_test_case_history_includes_car_log_url(self):
        vehicle = self._create_vehicle('history')
        case = self._create_cases(vehicle, 1)[0]
        result = self._report_result(
            vehicle,
            case,
            RESULT_FAILED,
            log_url='https://example.com/history/testcase.html',
        )

        history = services.get_test_case_history(str(case.id))

        self.assertEqual(history.total, 1)
        self.assertEqual(history.items[0].id, str(result.id))
        self.assertEqual(history.items[0].car_log_url, 'https://example.com/history/')

    def test_vehicle_domain_import_supports_duplicate_case_no_across_viu_codes(self):
        vehicle_platform = McuPlatform.objects.create(
            name='Vehicle Import Platform',
            version_code='vehicle-import-platform',
            domain=DOMAIN_VEHICLE,
            is_active=True,
        )
        vehicle = VehicleModel.objects.create(
            platform=vehicle_platform,
            name='Vehicle import model',
            vehicle_code='VEH-IMPORT',
            cdc_platform='CDC',
            execution_machine='machine-import',
            viu_codes=['viu0', 'viu1'],
            is_active=True,
        )

        payload = ImportCasePayload(
            vehicle_id=str(vehicle.id),
            rows=[
                ImportCaseRow(viu_code='viu0', case_no='CASE-001', case_name='Case A'),
                ImportCaseRow(viu_code='viu1', case_no='CASE-001', case_name='Case B'),
            ],
        )
        result = services.import_test_cases(None, payload)

        self.assertEqual(result.created_count, 2)
        self.assertEqual(result.updated_count, 0)
        self.assertEqual(result.ignored_count, 0)
        rows = services.list_test_cases(
            TestCaseFilter(domain=DOMAIN_VEHICLE, vehicle_id=str(vehicle.id))
        )
        self.assertEqual(len(rows), 2)
        self.assertEqual({item['viu_code'] for item in rows}, {'viu0', 'viu1'})

    def test_cockpit_soc_test_case_requires_module_and_returns_module_in_views(self):
        soc_platform = McuPlatform.objects.create(
            name='SOC Platform',
            version_code='soc-platform',
            domain=DOMAIN_COCKPIT_SOC,
            is_active=True,
        )
        soc_vehicle = VehicleModel.objects.create(
            platform=soc_platform,
            name='SOC Vehicle',
            vehicle_code='SOC-VEH-1',
            cdc_platform='SOC-CDC',
            execution_machine='soc-machine',
            viu_codes=[],
            is_active=True,
        )

        with self.assertRaisesMessage(Exception, '座舱SOC用例必须填写模块'):
            services.create_test_case(
                None,
                type(
                    'Payload',
                    (),
                    {
                        'vehicle_id': str(soc_vehicle.id),
                        'viu_code': '',
                        'module': '',
                        'case_no': 'SOC-CASE-001',
                        'case_name': 'SOC Case',
                        'remark': '',
                        'sort': 0,
                        'is_active': True,
                    },
                )(),
            )

        case = services.create_test_case(
            None,
            type(
                'Payload',
                (),
                {
                    'vehicle_id': str(soc_vehicle.id),
                    'viu_code': '',
                    'module': '音频',
                    'case_no': 'SOC-CASE-001',
                    'case_name': 'SOC Case',
                    'remark': '',
                    'sort': 0,
                    'is_active': True,
                },
            )(),
        )
        self.assertEqual(case['module'], '音频')

        DailyExecutionResult.objects.create(
            vehicle=soc_vehicle,
            test_case=AutoTestCase.objects.get(id=case['id']),
            execute_date=self.execute_date,
            start_time=timezone.now(),
            duration_seconds=60,
            result=RESULT_SUCCESS,
            log_url='https://example.com/soc/testcase.html',
        )

        rows = services.list_test_cases(
            TestCaseFilter(domain=DOMAIN_COCKPIT_SOC, vehicle_id=str(soc_vehicle.id))
        )
        self.assertEqual(rows[0]['module'], '音频')

        daily_items = services.list_daily_results(
            soc_vehicle.id,
            self.execute_date,
            DOMAIN_COCKPIT_SOC,
        )
        self.assertEqual(daily_items[0].module, '音频')

        history_page = services.get_test_case_history(case['id'], page=1, page_size=10)
        self.assertEqual(history_page.items[0].module, '音频')

    def test_list_test_cases_supports_independent_case_header_keywords(self):
        vehicle = self._create_vehicle('header-filter')
        AutoTestCase.objects.create(
            vehicle=vehicle,
            case_no='LOGIN-001',
            case_name='用户登录流程',
            is_active=True,
        )
        AutoTestCase.objects.create(
            vehicle=vehicle,
            case_no='LOGIN-002',
            case_name='用户退出流程',
            is_active=True,
        )
        AutoTestCase.objects.create(
            vehicle=vehicle,
            case_no='PAYMENT-001',
            case_name='支付登录联动',
            is_active=True,
        )

        by_case_no = services.list_test_cases(
            TestCaseFilter(case_no_keyword='login')
        )
        by_case_name = services.list_test_cases(
            TestCaseFilter(case_name_keyword='登录')
        )
        combined = services.list_test_cases(
            TestCaseFilter(case_no_keyword='payment', case_name_keyword='登录')
        )

        self.assertEqual({item['case_no'] for item in by_case_no}, {'LOGIN-001', 'LOGIN-002'})
        self.assertEqual({item['case_no'] for item in by_case_name}, {'LOGIN-001', 'PAYMENT-001'})
        self.assertEqual([item['case_no'] for item in combined], ['PAYMENT-001'])


class AutoTestReportResponsibleUserAndAnalysisStatsTests(TestCase):
    """验证车型责任人关联与第三方失败分析统计口径。"""

    def setUp(self):
        self.execute_date = date(2026, 8, 11)
        self.platform = McuPlatform.objects.create(
            name='Responsibility Platform',
            version_code='responsibility-platform',
            domain=DOMAIN_COCKPIT,
            is_active=True,
        )

    def _create_vehicle(self, suffix: str, *, domain=DOMAIN_COCKPIT):
        """创建归属指定领域的启用车型。"""
        platform = self.platform
        if domain != DOMAIN_COCKPIT:
            platform = McuPlatform.objects.create(
                name=f'{suffix} Platform',
                version_code=f'{suffix}-platform',
                domain=domain,
                is_active=True,
            )
        return VehicleModel.objects.create(
            platform=platform,
            name=f'{suffix} Vehicle',
            vehicle_code=f'{suffix}-vehicle',
            cdc_platform='CDC',
            execution_machine='machine',
            is_active=True,
        )

    def test_vehicle_responsible_users_are_saved_and_returned_by_related_views(self):
        owner = User.objects.create(username='owner-a', password='test', name='张三')
        fallback_owner = User.objects.create(username='owner-b', password='test', name='')
        payload = VehicleIn(
            platform_id=str(self.platform.id),
            name='Responsible Vehicle',
            vehicle_code='responsible-vehicle',
            cdc_platform='CDC',
            execution_machine='machine',
            responsible_user_ids=[str(owner.id), str(fallback_owner.id)],
        )

        created = services.create_vehicle(None, payload)
        vehicle = VehicleModel.objects.get(id=created['id'])
        case = AutoTestCase.objects.create(
            vehicle=vehicle,
            case_no='RESP-CASE',
            case_name='Responsible case',
            is_active=True,
        )

        self.assertEqual(
            set(created['responsible_user_ids']),
            {str(owner.id), str(fallback_owner.id)},
        )
        self.assertEqual(
            {item['name'] for item in created['responsible_users']},
            {'张三', 'owner-b'},
        )
        listed_case = services.list_test_cases(TestCaseFilter(vehicle_id=str(vehicle.id)))[0]
        self.assertEqual({item['id'] for item in listed_case['responsible_users']}, {str(owner.id), str(fallback_owner.id)})
        overview = services.get_daily_overview(DailyOverviewQuery(execute_date=self.execute_date))
        overview_row = next(item for item in overview.items if item.vehicle_id == str(vehicle.id))
        self.assertEqual({item.id for item in overview_row.responsible_users}, {str(owner.id), str(fallback_owner.id)})

        fallback_owner.soft_delete()
        payload.responsible_user_ids = [str(fallback_owner.id)]
        with self.assertRaisesMessage(Exception, '责任人不存在或已删除'):
            services.update_vehicle(None, str(vehicle.id), payload)
        self.assertEqual(case.vehicle_id, vehicle.id)

    def test_analysis_stats_filters_completed_vehicles_but_keeps_full_summary(self):
        vehicle = self._create_vehicle('cockpit')
        empty_vehicle = self._create_vehicle('empty')
        other_domain_vehicle = self._create_vehicle('vehicle-domain', domain=DOMAIN_VEHICLE)
        cases = [
            AutoTestCase.objects.create(
                vehicle=vehicle,
                case_no=f'CASE-{index}',
                case_name=f'Case {index}',
                is_active=True,
            )
            for index in range(3)
        ]
        DailyExecutionResult.objects.create(
            vehicle=vehicle,
            test_case=cases[0],
            execute_date=self.execute_date,
            start_time=timezone.now(),
            duration_seconds=10,
            result=RESULT_SUCCESS,
        )
        DailyExecutionResult.objects.create(
            vehicle=vehicle,
            test_case=cases[0],
            execute_date=self.execute_date,
            start_time=timezone.now() + timedelta(minutes=1),
            duration_seconds=10,
            result=RESULT_FAILED,
            failure_category=FAILURE_CATEGORY_VERSION,
        )
        DailyExecutionResult.objects.create(
            vehicle=vehicle,
            test_case=cases[1],
            execute_date=self.execute_date,
            start_time=timezone.now() + timedelta(minutes=2),
            duration_seconds=10,
            result=RESULT_TIMEOUT,
        )
        DailyExecutionResult.objects.create(
            vehicle=vehicle,
            test_case=cases[2],
            execute_date=self.execute_date,
            start_time=timezone.now() + timedelta(minutes=3),
            duration_seconds=10,
            result=RESULT_SKIP,
            failure_category=FAILURE_CATEGORY_ENVIRONMENT,
        )
        AutoTestCase.objects.create(
            vehicle=other_domain_vehicle,
            case_no='OTHER-CASE',
            case_name='Other case',
            is_active=True,
        )

        result = services.get_daily_analysis_stats(DOMAIN_COCKPIT, self.execute_date)
        row = next(item for item in result.items if item.vehicle_id == str(vehicle.id))

        self.assertEqual(result.summary.vehicle_count, 2)
        self.assertEqual(result.summary.failed_count, 3)
        self.assertEqual(result.summary.need_analysis_count, 3)
        self.assertEqual(result.summary.pending_analysis_count, 1)
        self.assertEqual(result.summary.version_failure_count, 1)
        self.assertEqual(row.failed_count, 3)
        self.assertEqual(row.need_analysis_count, 3)
        self.assertEqual(row.pending_analysis_count, 1)
        self.assertEqual(row.version_failure_count, 1)
        self.assertEqual(len(result.items), 1)
        self.assertNotIn(str(empty_vehicle.id), {item.vehicle_id for item in result.items})

    def test_analysis_stats_api_allows_anonymous_request(self):
        """接口显式关闭鉴权，第三方无需 Bearer token 即可调用。"""
        response = self.client.get(
            '/api/auto-test-report/daily-results/analysis-stats',
            {'domain': DOMAIN_COCKPIT, 'execute_date': self.execute_date.isoformat()},
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()['summary']['domain'], DOMAIN_COCKPIT)

    def test_analysis_stats_defaults_to_server_local_date(self):
        """未传执行日期时按服务端当天统计，便于第三方定时拉取。"""
        with patch.object(services.timezone, 'localdate', return_value=self.execute_date):
            result = services.get_daily_analysis_stats(DOMAIN_COCKPIT)

        self.assertEqual(result.summary.execute_date, self.execute_date)
