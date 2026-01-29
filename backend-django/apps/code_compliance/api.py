from ninja import Router, File, UploadedFile
from typing import List, Optional
from django.http import HttpResponse
from datetime import datetime
from .schemas import ComplianceRecordSchema, ComplianceUpdateSchema, OverviewSummarySchema, DetailSummarySchema
from .services import get_post_stats, get_post_users_detail, get_user_records, update_branch_status
from .models import ComplianceRecord, ComplianceBranch
from core.user.user_model import User
import openpyxl
import csv
import io
import re

router = Router()

@router.get("/stats/post", response=OverviewSummarySchema)
def list_post_stats(request):
    """
    获取岗位维度的合规风险统计
    """
    data = get_post_stats()
    return data

@router.get("/stats/post/{post_id}/users", response=DetailSummarySchema)
def list_post_users(request, post_id: str, start_date: Optional[str] = None, end_date: Optional[str] = None):
    """
    获取指定岗位下的用户风险详情
    """
    data = get_post_users_detail(post_id, start_date, end_date)
    return data

@router.get("/user/{user_id}/records", response=List[ComplianceRecordSchema])
def list_user_records(request, user_id: str):
    """
    获取用户的所有风险记录
    """
    data = get_user_records(user_id)
    return data

@router.put("/branch/{branch_id}")
def update_branch(request, branch_id: str, data: ComplianceUpdateSchema):
    """
    更新分支的处理状态
    """
    update_branch_status(branch_id, data.status, data.remark, user=request.auth)
    return {"msg": "操作成功"}

@router.post("/upload")
def upload_compliance_data(request, file: UploadedFile = File(...)):
    """
    批量上传合规风险数据 (支持 .xlsx 和 .csv)
    """
    try:
        data_rows = []
        filename = file.name.lower()
        
        if filename.endswith('.xlsx'):
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            # Iterate rows, skipping header
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or not any(row):
                    continue
                data_rows.append(list(row))
        elif filename.endswith('.csv'):
            raw_content = file.read()
            try:
                content = raw_content.decode('utf-8-sig')
            except UnicodeDecodeError:
                try:
                    content = raw_content.decode('gbk')
                except UnicodeDecodeError:
                    content = raw_content.decode('utf-8', errors='ignore')
            
            csv_reader = csv.reader(io.StringIO(content))
            header = next(csv_reader, None) # Skip header
            for row in csv_reader:
                if not row or not any(row):
                    continue
                data_rows.append(row)
        else:
            return {"msg": "仅支持 .xlsx 或 .csv 文件格式"}

        count = 0
        added_records = 0
        updated_records = 0
        
        for row in data_rows:
            # Columns: ChangeId, Title, URL, User(Username/Email), Branches, Remark
            if not row or len(row) < 1 or not row[0]:
                continue
                
            change_id = str(row[0]).strip()
            title = str(row[1]).strip() if len(row) > 1 and row[1] else ""
            url = str(row[2]).strip() if len(row) > 2 and row[2] else ""
            user_ident = str(row[3]).strip() if len(row) > 3 and row[3] else ""
            branches_str = str(row[4]).strip() if len(row) > 4 and row[4] else ""
            remark = str(row[5]).strip() if len(row) > 5 and row[5] else ""
            
            # Find User
            user = None
            if user_ident:
                # Try username first, then email
                user = User.objects.filter(username=user_ident).first()
                if not user:
                    user = User.objects.filter(email=user_ident).first()
            
            if not user:
                # Skip if no valid user found
                continue
                
            # Create/Update Record
            # Requirement: if change_id exists, identify new/old branch info.
            # If branch is duplicate, skip; if new, save.
            record = ComplianceRecord.objects.filter(change_id=change_id).first()
            
            if not record:
                # Create new record
                record = ComplianceRecord.objects.create(
                    change_id=change_id,
                    user=user,
                    title=title,
                    url=url,
                    update_time=datetime.now(),
                    remark=remark
                )
                added_records += 1
            else:
                # Record exists, optionally update basic info if needed
                # But user said "以数据库里面保存的为准", usually refers to branches.
                # We'll update basic info if they are provided but keep existing ones if not.
                if title: record.title = title
                if url: record.url = url
                if remark: record.remark = remark
                record.update_time = datetime.now()
                record.save()
                updated_records += 1
            
            # Parse and Update Branches
            if branches_str:
                # Support comma, semicolon or newline separated
                branch_names = [b.strip() for b in re.split(r'[,;\n]', branches_str) if b.strip()]
                
                # Get existing branches for this record
                existing_branch_names = set(ComplianceBranch.objects.filter(record=record).values_list('branch_name', flat=True))
                
                new_branches_added = False
                for bn in branch_names:
                    if bn not in existing_branch_names:
                        ComplianceBranch.objects.create(record=record, branch_name=bn)
                        new_branches_added = True
                
                # If new branches added, ensure record status is updated to 'pending' if it was resolved
                if new_branches_added:
                    all_branches = record.branches.all()
                    if any(b.status == 0 for b in all_branches):
                        record.status = 0
                        record.save(update_fields=['status'])
            
            count += 1
            
        return {
            "msg": f"上传成功: 处理 {count} 条记录 (新增 {added_records}, 更新 {updated_records})",
            "code": 200
        }
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return {"msg": f"上传失败: {str(e)}", "code": 500}

@router.get("/template")
def get_upload_template(request):
    """
    获取上传模板
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["ChangeId", "Title", "URL", "User(Username/Email)", "Branches(comma separated)", "Remark"])
    
    # Add some example data
    ws.append(["I123456", "Fix NPE", "http://git/...", "zhangsan", "master, release-1.0", "Urgent"])
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=compliance_template.xlsx'
    wb.save(response)
    return response
