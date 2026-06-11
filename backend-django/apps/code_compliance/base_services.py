import hashlib
import io
import json
import logging
import re
import tempfile
import threading
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Iterable, Optional

import openpyxl
from django.db import close_old_connections, connection, transaction
from django.db.models import Count, Q
from django.http import FileResponse, HttpResponse
from django.shortcuts import get_object_or_404
from django.utils import timezone
from ninja.errors import HttpError

from core.dict.dict_model import Dict
from core.dict_item.dict_item_model import DictItem
from core.pl.pl_model import PlGroup

from .base_schemas import ImportErrorRow, ImportResultOut
from .models import (
    COMPLIANCE_BRANCH_TYPE_CHOICES,
    COMPLIANCE_BRANCH_TYPE_DEVELOPMENT,
    COMPLIANCE_BRANCH_TYPE_OTHER,
    COMPLIANCE_BRANCH_TYPE_RELEASE,
    COMPLIANCE_BRANCH_TYPE_TRUNK,
    COMPLIANCE_DOMAIN_CHOICES,
    COMPLIANCE_DOMAIN_COCKPIT,
    COMPLIANCE_DOMAIN_VEHICLE,
    COMPLIANCE_MODE_CHOICES,
    COMPLIANCE_MODE_CR,
    COMPLIANCE_MODE_MR,
    ComplianceManagedBranch,
    ComplianceOrganization,
    ComplianceRepository,
    ComplianceRepositoryBranch,
    ComplianceRepositoryExportTask,
)


logger = logging.getLogger(__name__)

REPO_TYPE_DICT_CODE = "code_compliance_repo_type"
BIND_MODE_APPEND = "append"
BIND_MODE_REPLACE = "replace"
SUPPORTED_BIND_MODES = {BIND_MODE_APPEND, BIND_MODE_REPLACE}
REPOSITORY_EXPORT_SCOPE_ALL = "all"
REPOSITORY_EXPORT_SCOPE_FILTERED = "filtered"
REPOSITORY_EXPORT_SCOPES = {REPOSITORY_EXPORT_SCOPE_ALL, REPOSITORY_EXPORT_SCOPE_FILTERED}
REPOSITORY_EXPORT_ACTIVE_STATUSES = {
    ComplianceRepositoryExportTask.STATUS_PENDING,
    ComplianceRepositoryExportTask.STATUS_RUNNING,
}
REPOSITORY_EXPORT_FILE_TTL_SECONDS = 24 * 60 * 60
REPOSITORY_EXPORT_HEADERS = [
    "组织ID",
    "组织名",
    "父组织ID",
    "父组织名",
    "组织路径",
    "组织模式",
    "组织领域",
    "组织备注",
    "代码库ID",
    "代码库名",
    "代码库URL",
    "代码库模式",
    "代码库领域",
    "代码仓类型",
    "责任PL组",
    "绑定分支数",
    "代码库备注",
    "创建时间",
    "更新时间",
]

MODE_LABELS = dict(COMPLIANCE_MODE_CHOICES)
DOMAIN_LABELS = dict(COMPLIANCE_DOMAIN_CHOICES)
BRANCH_TYPE_LABELS = dict(COMPLIANCE_BRANCH_TYPE_CHOICES)


@dataclass
class BindCounters:
    created_count: int = 0
    restored_count: int = 0
    removed_count: int = 0
    ignored_count: int = 0

    def as_dict(self) -> dict:
        """把绑定计数转换成 API 输出字典。"""
        return {
            "created_count": self.created_count,
            "restored_count": self.restored_count,
            "removed_count": self.removed_count,
            "ignored_count": self.ignored_count,
        }


def _clean_text(value) -> str:
    """把 Excel/表单输入统一转换成去空格字符串。"""
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value)).strip()
    return str(value).strip()


def _optional_text(value) -> Optional[str]:
    """把空字符串归一成 None，便于落库备注字段。"""
    value = _clean_text(value)
    return value or None


def _audit_user_id(user) -> Optional[str]:
    """从 request.auth 中提取审计用户 ID。"""
    return str(getattr(user, "id", "") or "") or None


def _apply_audit_fields(instance, user, *, is_create: bool = False):
    """在服务层手动补齐 RootModel 的创建人和修改人。"""
    user_id = _audit_user_id(user)
    if not user_id:
        return
    if is_create and hasattr(instance, "sys_creator_id"):
        instance.sys_creator_id = user_id
    if hasattr(instance, "sys_modifier_id"):
        instance.sys_modifier_id = user_id


def _normalize_mode(value: Optional[str]) -> str:
    """规范化 CR/MR 模式输入。"""
    raw = _clean_text(value).upper()
    if not raw:
        return COMPLIANCE_MODE_CR
    if raw in {COMPLIANCE_MODE_CR, COMPLIANCE_MODE_MR}:
        return raw
    raise HttpError(400, f"模式仅支持 CR 或 MR: {value}")


def _normalize_domain(value: Optional[str]) -> str:
    """规范化座舱/车控领域输入，兼容中英文别名。"""
    raw = _clean_text(value).lower()
    if not raw:
        return COMPLIANCE_DOMAIN_COCKPIT
    aliases = {
        "cockpit": COMPLIANCE_DOMAIN_COCKPIT,
        "cabin": COMPLIANCE_DOMAIN_COCKPIT,
        "座舱": COMPLIANCE_DOMAIN_COCKPIT,
        "座舱领域": COMPLIANCE_DOMAIN_COCKPIT,
        "vehicle": COMPLIANCE_DOMAIN_VEHICLE,
        "car": COMPLIANCE_DOMAIN_VEHICLE,
        "车控": COMPLIANCE_DOMAIN_VEHICLE,
        "车控领域": COMPLIANCE_DOMAIN_VEHICLE,
    }
    if raw in aliases:
        return aliases[raw]
    raise HttpError(400, f"领域仅支持 cockpit/vehicle 或 座舱/车控: {value}")


def _normalize_branch_type(value: Optional[str]) -> str:
    """规范化分支类型输入，兼容常见英文别名。"""
    raw = _clean_text(value).lower()
    if not raw:
        return COMPLIANCE_BRANCH_TYPE_OTHER
    aliases = {
        "development": COMPLIANCE_BRANCH_TYPE_DEVELOPMENT,
        "develop": COMPLIANCE_BRANCH_TYPE_DEVELOPMENT,
        "dev": COMPLIANCE_BRANCH_TYPE_DEVELOPMENT,
        "开发": COMPLIANCE_BRANCH_TYPE_DEVELOPMENT,
        "trunk": COMPLIANCE_BRANCH_TYPE_TRUNK,
        "main": COMPLIANCE_BRANCH_TYPE_TRUNK,
        "master": COMPLIANCE_BRANCH_TYPE_TRUNK,
        "主干": COMPLIANCE_BRANCH_TYPE_TRUNK,
        "release": COMPLIANCE_BRANCH_TYPE_RELEASE,
        "发布": COMPLIANCE_BRANCH_TYPE_RELEASE,
        "other": COMPLIANCE_BRANCH_TYPE_OTHER,
        "其他": COMPLIANCE_BRANCH_TYPE_OTHER,
    }
    if raw in aliases:
        return aliases[raw]
    raise HttpError(400, f"分支类型仅支持 开发/主干/发布/其他: {value}")


def _normalize_branch_active(value, *, default: bool = True) -> bool:
    """规范化分支活跃状态输入，兼容 Excel 中的中英文和数字写法。"""
    if value is None or value == "":
        return default
    if isinstance(value, bool):
        return value
    raw = _clean_text(value).lower()
    truthy = {"1", "true", "yes", "y", "是", "活跃", "active", "启用"}
    falsy = {"0", "false", "no", "n", "否", "归档", "已归档", "非活跃", "inactive", "archived", "停用"}
    if raw in truthy:
        return True
    if raw in falsy:
        return False
    raise HttpError(400, f"是否活跃仅支持 是/否、true/false、1/0、活跃/归档: {value}")


def _parse_date(value) -> Optional[date]:
    """解析 Excel 或表单传入的日期值。"""
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    raw = _clean_text(value)
    if not raw:
        return None
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"):
        try:
            return datetime.strptime(raw, fmt).date()
        except ValueError:
            continue
    raise HttpError(400, f"日期格式应为 YYYY-MM-DD: {raw}")


def _split_list(value) -> list[str]:
    """把逗号、分号或换行分隔的文本拆成去重列表。"""
    raw = _clean_text(value)
    if not raw:
        return []
    values = [item.strip() for item in re.split(r"[,，;；\n\r]+", raw) if item.strip()]
    result: list[str] = []
    seen: set[str] = set()
    for item in values:
        key = item.lower()
        if key in seen:
            continue
        seen.add(key)
        result.append(item)
    return result


def _active_organizations():
    """返回未软删除的组织查询集。"""
    return ComplianceOrganization.objects.filter(is_deleted=False)


def _active_repositories():
    """返回未软删除的代码库查询集。"""
    return ComplianceRepository.objects.filter(is_deleted=False)


def _active_branches():
    """返回未软删除的分支主数据查询集。"""
    return ComplianceManagedBranch.objects.filter(is_deleted=False)


def _get_active_organization(org_id: str) -> ComplianceOrganization:
    """按系统主键读取活跃组织。"""
    return get_object_or_404(_active_organizations(), id=org_id)


def _get_active_repository(repo_id: str) -> ComplianceRepository:
    """按系统主键读取活跃代码库。"""
    return get_object_or_404(_active_repositories(), id=repo_id)


def _get_active_branch(branch_id: str) -> ComplianceManagedBranch:
    """按系统主键读取活跃分支。"""
    return get_object_or_404(_active_branches(), id=branch_id)


def _get_parent(parent_id: Optional[str]) -> Optional[ComplianceOrganization]:
    """把可空父组织 ID 转成组织对象。"""
    if not parent_id:
        return None
    return _get_active_organization(parent_id)


def _ensure_org_parent_valid(
    *,
    org_id: Optional[str],
    parent: Optional[ComplianceOrganization],
):
    """校验组织父节点，避免自引用或形成循环。"""
    if not org_id or not parent:
        return
    if str(parent.id) == str(org_id):
        raise HttpError(400, "不能将组织自身设为父组织")

    current = parent
    visited: set[str] = set()
    while current:
        current_id = str(current.id)
        if current_id in visited:
            raise HttpError(400, "组织树存在循环，无法保存")
        if current_id == str(org_id):
            raise HttpError(400, "不能将子组织设为父组织")
        visited.add(current_id)
        current = current.parent


def _repo_type_items():
    """返回启用的代码合规仓库类型字典项。"""
    return DictItem.objects.filter(
        dict__code=REPO_TYPE_DICT_CODE,
        dict__status=True,
        status=True,
    )


def _repo_type_label_map() -> dict[str, str]:
    """构建仓库类型 value 到 label 的映射。"""
    return {
        item.value: item.label or item.value
        for item in _repo_type_items()
        if item.value
    }


def _ensure_repo_type_valid(repo_type: Optional[str]) -> str:
    """校验代码仓类型是否来自启用的 core 字典项。"""
    value = _clean_text(repo_type)
    if not value:
        return ""
    if not Dict.objects.filter(code=REPO_TYPE_DICT_CODE, status=True).exists():
        raise HttpError(400, f"系统字典未配置: {REPO_TYPE_DICT_CODE}")
    if not _repo_type_items().filter(value=value).exists():
        raise HttpError(400, f"代码仓类型不存在或未启用: {value}")
    return value


def _get_pl_groups(group_ids: Iterable[str]) -> list[PlGroup]:
    """根据前端传入的 PL 组 ID 查找启用资源组。"""
    ids = [str(item) for item in group_ids if str(item or "").strip()]
    if not ids:
        return []
    groups = list(PlGroup.objects.filter(id__in=ids, status=True))
    found_ids = {str(item.id) for item in groups}
    missing_ids = [item for item in ids if item not in found_ids]
    if missing_ids:
        raise HttpError(400, f"责任PL资源组不存在或未启用: {', '.join(missing_ids)}")
    return groups


def _resolve_pl_groups_from_text(value) -> list[PlGroup]:
    """根据 Excel 中的 PL 组名称或编码解析启用资源组。"""
    labels = _split_list(value)
    if not labels:
        return []

    groups: list[PlGroup] = []
    missing: list[str] = []
    seen_ids: set[str] = set()
    for label in labels:
        group = (
            PlGroup.objects.filter(status=True, code=label).first()
            or PlGroup.objects.filter(status=True, name=label).first()
        )
        if not group:
            missing.append(label)
            continue
        if str(group.id) in seen_ids:
            continue
        seen_ids.add(str(group.id))
        groups.append(group)
    if missing:
        raise HttpError(400, f"责任PL资源组不存在或未启用: {', '.join(missing)}")
    return groups


def serialize_organization(
    item: ComplianceOrganization,
    *,
    include_children: bool = False,
) -> dict:
    """把组织模型序列化为组织树节点。"""
    children = getattr(item, "child_list", []) if include_children else []
    repository_count = getattr(item, "repository_count", 0)
    return {
        "id": str(item.id),
        "group_id": item.group_id,
        "name": item.name,
        "parent_id": str(item.parent_id) if item.parent_id else None,
        "parent_name": item.parent.name if getattr(item, "parent", None) else None,
        "mode": item.mode,
        "mode_label": MODE_LABELS.get(item.mode, item.mode),
        "domain": item.domain,
        "domain_label": DOMAIN_LABELS.get(item.domain, item.domain),
        "remark": item.remark,
        "sort": item.sort,
        "repository_count": int(repository_count or 0),
        "sys_create_datetime": item.sys_create_datetime,
        "sys_update_datetime": item.sys_update_datetime,
        "children": [
            serialize_organization(child, include_children=True)
            for child in children
        ],
    }


def serialize_repository(item: ComplianceRepository) -> dict:
    """把代码库模型序列化为列表行数据。"""
    repo_type_label = _repo_type_label_map().get(item.repo_type, item.repo_type or "")
    responsibility_groups = list(getattr(item, "prefetched_responsibility_groups", []))
    if not responsibility_groups:
        responsibility_groups = list(item.responsibility_groups.filter(status=True))
    return {
        "id": str(item.id),
        "project_id": item.project_id,
        "project_name": item.project_name,
        "project_url": item.project_url or "",
        "organization_id": str(item.organization_id),
        "organization_name": item.organization.name,
        "organization_group_id": item.organization.group_id,
        "mode": item.mode,
        "mode_label": MODE_LABELS.get(item.mode, item.mode),
        "responsibility_group_ids": [str(group.id) for group in responsibility_groups],
        "responsibility_group_names": [group.name for group in responsibility_groups],
        "repo_type": item.repo_type or "",
        "repo_type_label": repo_type_label,
        "domain": item.domain,
        "domain_label": DOMAIN_LABELS.get(item.domain, item.domain),
        "remark": item.remark,
        "sort": item.sort,
        "branch_count": int(getattr(item, "branch_count", 0) or 0),
        "sys_create_datetime": item.sys_create_datetime,
        "sys_update_datetime": item.sys_update_datetime,
    }


def serialize_branch(item: ComplianceManagedBranch) -> dict:
    """把分支主数据序列化为列表行数据。"""
    return {
        "id": str(item.id),
        "branch_name": item.branch_name,
        "created_date": item.created_date,
        "branch_type": item.branch_type,
        "branch_type_label": BRANCH_TYPE_LABELS.get(item.branch_type, item.branch_type),
        "alias": item.alias or "",
        "purpose": item.purpose or "",
        "remark": item.remark,
        "is_active": bool(item.is_active),
        "domain": item.domain,
        "domain_label": DOMAIN_LABELS.get(item.domain, item.domain),
        "sort": item.sort,
        "repository_count": int(getattr(item, "repository_count", 0) or 0),
        "sys_create_datetime": item.sys_create_datetime,
        "sys_update_datetime": item.sys_update_datetime,
    }


def list_organization_tree(exclude_id: Optional[str] = None) -> list[dict]:
    """构建组织树，并在编辑父组织时排除指定节点及后代。"""
    nodes = list(
        _active_organizations()
        .select_related("parent")
        .annotate(
            repository_count=Count(
                "repositories",
                filter=Q(repositories__is_deleted=False),
                distinct=True,
            )
        )
        .order_by("sort", "name")
    )

    node_map = {str(item.id): item for item in nodes}
    forbidden_ids = _descendant_ids(exclude_id, node_map) if exclude_id else set()
    roots: list[ComplianceOrganization] = []
    # 组织树在内存中拼装，避免每个节点递归查询数据库。
    for item in nodes:
        item.child_list = []
    for item in nodes:
        if str(item.id) in forbidden_ids:
            continue
        parent_id = str(item.parent_id) if item.parent_id else None
        if parent_id and parent_id in node_map and parent_id not in forbidden_ids:
            node_map[parent_id].child_list.append(item)
        else:
            roots.append(item)
    return [serialize_organization(item, include_children=True) for item in roots]


def _descendant_ids(
    org_id: Optional[str],
    node_map: dict[str, ComplianceOrganization],
) -> set[str]:
    """计算指定组织及其所有后代，供父组织候选项排除。"""
    if not org_id or org_id not in node_map:
        return set()
    forbidden = set()
    stack = [node_map[org_id]]
    for item in node_map.values():
        item.child_list = []
    for item in node_map.values():
        if item.parent_id and str(item.parent_id) in node_map:
            node_map[str(item.parent_id)].child_list.append(item)
    while stack:
        current = stack.pop()
        forbidden.add(str(current.id))
        stack.extend(current.child_list)
    return forbidden


@transaction.atomic
def create_organization(user, payload) -> dict:
    """创建组织；若命中已软删除外部 ID，则恢复后更新。"""
    data = payload.dict()
    group_id = _clean_text(data["group_id"])
    name = _clean_text(data["name"])
    if not group_id:
        raise HttpError(400, "组织ID不能为空")
    if not name:
        raise HttpError(400, "组织名不能为空")

    parent = _get_parent(data.get("parent_id"))
    existing = ComplianceOrganization.objects.filter(group_id=group_id).first()
    if existing and not existing.is_deleted:
        raise HttpError(400, f"组织ID已存在: {group_id}")

    item = existing or ComplianceOrganization(group_id=group_id)
    _ensure_org_parent_valid(org_id=str(item.id) if existing else None, parent=parent)
    item.group_id = group_id
    item.name = name
    item.parent = parent
    item.mode = _normalize_mode(data.get("mode"))
    item.domain = _normalize_domain(data.get("domain"))
    item.remark = _optional_text(data.get("remark"))
    item.sort = int(data.get("sort") or 0)
    item.is_deleted = False
    _apply_audit_fields(item, user, is_create=not existing)
    item.save()
    return serialize_organization(item)


@transaction.atomic
def update_organization(user, org_id: str, payload) -> dict:
    """更新组织字段，并重新校验外部 ID 和父组织合法性。"""
    item = _get_active_organization(org_id)
    data = payload.dict(exclude_unset=True)

    if "group_id" in data:
        group_id = _clean_text(data["group_id"])
        if not group_id:
            raise HttpError(400, "组织ID不能为空")
        exists = (
            ComplianceOrganization.objects.filter(group_id=group_id, is_deleted=False)
            .exclude(id=org_id)
            .exists()
        )
        if exists:
            raise HttpError(400, f"组织ID已存在: {group_id}")
        item.group_id = group_id
    if "name" in data:
        name = _clean_text(data["name"])
        if not name:
            raise HttpError(400, "组织名不能为空")
        item.name = name
    if "parent_id" in data:
        parent = _get_parent(data.get("parent_id"))
        _ensure_org_parent_valid(org_id=org_id, parent=parent)
        item.parent = parent
    if "mode" in data:
        item.mode = _normalize_mode(data.get("mode"))
    if "domain" in data:
        item.domain = _normalize_domain(data.get("domain"))
    if "remark" in data:
        item.remark = _optional_text(data.get("remark"))
    if "sort" in data:
        item.sort = int(data.get("sort") or 0)

    _apply_audit_fields(item, user)
    item.save()
    return serialize_organization(item)


@transaction.atomic
def delete_organization(org_id: str) -> dict:
    """删除组织前强校验子组织和代码库依赖。"""
    item = _get_active_organization(org_id)
    if item.children.filter(is_deleted=False).exists():
        raise HttpError(400, "该组织存在子组织，无法删除")
    if item.repositories.filter(is_deleted=False).exists():
        raise HttpError(400, "该组织下存在代码库，无法删除")
    item.soft_delete()
    return {"id": str(item.id)}


def _repository_queryset():
    """返回代码库列表基础查询，集中挂载组织和分支统计。"""
    return (
        _active_repositories()
        .select_related("organization", "organization__parent")
        .prefetch_related("responsibility_groups")
        .annotate(
            branch_count=Count(
                "branch_links",
                filter=Q(
                    branch_links__is_deleted=False,
                    branch_links__branch__is_deleted=False,
                ),
                distinct=True,
            )
        )
    )


def _apply_repository_filters(
    qs,
    *,
    organization_id: Optional[str] = None,
    keyword: Optional[str] = None,
    mode: Optional[str] = None,
    domain: Optional[str] = None,
    repo_type: Optional[str] = None,
):
    """统一应用代码库列表和导出的筛选条件。"""
    if organization_id:
        qs = qs.filter(organization_id=organization_id)
    if keyword:
        word = keyword.strip()
        qs = qs.filter(
            Q(project_id__icontains=word)
            | Q(project_name__icontains=word)
            | Q(project_url__icontains=word)
        )
    if mode:
        qs = qs.filter(mode=_normalize_mode(mode))
    if domain:
        qs = qs.filter(domain=_normalize_domain(domain))
    if repo_type:
        qs = qs.filter(repo_type=_clean_text(repo_type))
    return qs


def list_repositories(
    *,
    page: int = 1,
    page_size: int = 20,
    organization_id: Optional[str] = None,
    keyword: Optional[str] = None,
    mode: Optional[str] = None,
    domain: Optional[str] = None,
    repo_type: Optional[str] = None,
) -> dict:
    """按页面筛选条件分页查询代码库。"""
    qs = _apply_repository_filters(
        _repository_queryset(),
        organization_id=organization_id,
        keyword=keyword,
        mode=mode,
        domain=domain,
        repo_type=repo_type,
    )

    total = qs.count()
    offset = max(page - 1, 0) * page_size
    items = list(qs.order_by("sort", "project_name")[offset : offset + page_size])
    return {"items": [serialize_repository(item) for item in items], "total": total}


def _repository_export_temp_dir() -> Path:
    """返回代码库导出临时文件目录，不依赖同步请求生命周期。"""
    path = Path(tempfile.gettempdir()) / "focus_admin_code_compliance_exports"
    path.mkdir(parents=True, exist_ok=True)
    return path


def _normalize_repository_export_payload(data) -> dict:
    """把导出任务入参规范化，保证相同筛选能得到稳定指纹。"""
    raw = data.dict() if hasattr(data, "dict") else dict(data or {})
    scope = _clean_text(raw.get("scope")).lower() or REPOSITORY_EXPORT_SCOPE_ALL
    if scope not in REPOSITORY_EXPORT_SCOPES:
        raise HttpError(400, "导出范围仅支持 all 或 filtered")
    payload = {"scope": scope}
    if scope == REPOSITORY_EXPORT_SCOPE_FILTERED:
        payload.update(
            {
                "organization_id": _clean_text(raw.get("organization_id")),
                "keyword": _clean_text(raw.get("keyword")),
                "mode": _clean_text(raw.get("mode")),
                "domain": _clean_text(raw.get("domain")),
                "repo_type": _clean_text(raw.get("repo_type")),
            }
        )
    return payload


def _repository_export_fingerprint(payload: dict) -> str:
    """按规范化 payload 生成导出任务指纹。"""
    raw = json.dumps(payload, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()


def _serialize_repository_export_task(task: ComplianceRepositoryExportTask) -> dict:
    """序列化导出任务状态，供前端轮询展示。"""
    return {
        "id": str(task.id),
        "scope": task.scope,
        "fingerprint": task.fingerprint,
        "status": task.status,
        "progress": int(task.progress or 0),
        "message": task.message or "",
        "error_message": task.error_message or "",
        "file_name": task.file_name or None,
        "file_size": int(task.file_size or 0),
        "started_at": task.started_at,
        "finished_at": task.finished_at,
        "sys_create_datetime": task.sys_create_datetime,
    }


def _is_repository_export_task_expired(task: ComplianceRepositoryExportTask) -> bool:
    """判断成功导出文件是否已超过保留时间。"""
    if task.finished_at is None:
        return False
    return timezone.now() > task.finished_at + timedelta(seconds=REPOSITORY_EXPORT_FILE_TTL_SECONDS)


def _is_repository_export_task_downloadable(task: ComplianceRepositoryExportTask) -> bool:
    """判断任务是否具备可下载文件。"""
    if task.status != ComplianceRepositoryExportTask.STATUS_SUCCESS:
        return False
    if _is_repository_export_task_expired(task):
        return False
    file_path = _clean_text(task.file_path)
    return bool(file_path and Path(file_path).is_file())


def _cleanup_repository_export_files(limit: int = 100):
    """清理过期导出文件并保留任务历史。"""
    expire_before = timezone.now() - timedelta(seconds=REPOSITORY_EXPORT_FILE_TTL_SECONDS)
    stale_tasks = (
        ComplianceRepositoryExportTask.objects.filter(
            status=ComplianceRepositoryExportTask.STATUS_SUCCESS,
            is_deleted=False,
            finished_at__lt=expire_before,
        )
        .order_by("finished_at")[: max(int(limit or 0), 1)]
    )
    for task in stale_tasks:
        file_path = _clean_text(task.file_path)
        if file_path:
            try:
                path = Path(file_path)
                if path.exists():
                    path.unlink()
            except Exception:
                logger.warning("Cleanup repository export file failed task_id=%s", task.id, exc_info=True)
        ComplianceRepositoryExportTask.objects.filter(id=task.id).update(
            file_path="",
            file_name="",
            file_size=0,
        )


def _get_active_repository_export_task(user, fingerprint: str) -> Optional[ComplianceRepositoryExportTask]:
    """查找同一用户同一筛选条件下仍在执行的导出任务。"""
    if not user or not getattr(user, "id", None):
        return None
    return (
        ComplianceRepositoryExportTask.objects.filter(
            user=user,
            fingerprint=fingerprint,
            status__in=REPOSITORY_EXPORT_ACTIVE_STATUSES,
            is_deleted=False,
        )
        .order_by("-sys_create_datetime")
        .first()
    )


def _get_reusable_repository_export_task(user, fingerprint: str) -> Optional[ComplianceRepositoryExportTask]:
    """复用同一用户近期已完成且文件仍存在的导出任务。"""
    if not user or not getattr(user, "id", None):
        return None
    task = (
        ComplianceRepositoryExportTask.objects.filter(
            user=user,
            fingerprint=fingerprint,
            status=ComplianceRepositoryExportTask.STATUS_SUCCESS,
            is_deleted=False,
        )
        .order_by("-finished_at", "-sys_create_datetime")
        .first()
    )
    if task and _is_repository_export_task_downloadable(task):
        return task
    return None


def _repository_export_queryset(payload: dict):
    """根据导出范围构建不分页的代码库查询。"""
    qs = _repository_queryset()
    if payload.get("scope") == REPOSITORY_EXPORT_SCOPE_FILTERED:
        qs = _apply_repository_filters(
            qs,
            organization_id=payload.get("organization_id") or None,
            keyword=payload.get("keyword") or None,
            mode=payload.get("mode") or None,
            domain=payload.get("domain") or None,
            repo_type=payload.get("repo_type") or None,
        )
    return qs.order_by("organization__sort", "organization__name", "sort", "project_name")


def _organization_path_map() -> dict[str, str]:
    """预加载组织路径，避免导出每一行递归查询父组织。"""
    organizations = {
        str(item.id): item
        for item in _active_organizations().select_related("parent").order_by("sort", "name")
    }
    cache: dict[str, str] = {}

    def build_path(org: ComplianceOrganization) -> str:
        """递归构建组织路径并缓存结果。"""
        org_id = str(org.id)
        if org_id in cache:
            return cache[org_id]
        if org.parent_id and str(org.parent_id) in organizations:
            value = f"{build_path(organizations[str(org.parent_id)])} / {org.name}"
        else:
            value = org.name
        cache[org_id] = value
        return value

    for organization in organizations.values():
        build_path(organization)
    return cache


def _format_export_datetime(value) -> str:
    """把日期时间转换成 Excel 中稳定可读的字符串。"""
    if not value:
        return ""
    return value.strftime("%Y-%m-%d %H:%M:%S") if hasattr(value, "strftime") else str(value)


def _repository_export_row(
    repo: ComplianceRepository,
    path_map: dict[str, str],
    repo_type_map: dict[str, str],
) -> list:
    """把单个代码库及其组织信息转换为 Excel 行。"""
    organization = repo.organization
    repo_type_label = repo_type_map.get(repo.repo_type, repo.repo_type or "")
    responsibility_groups = [group for group in repo.responsibility_groups.all() if group.status]
    return [
        organization.group_id,
        organization.name,
        organization.parent.group_id if organization.parent else "",
        organization.parent.name if organization.parent else "",
        path_map.get(str(organization.id), organization.name),
        MODE_LABELS.get(organization.mode, organization.mode),
        DOMAIN_LABELS.get(organization.domain, organization.domain),
        organization.remark or "",
        repo.project_id,
        repo.project_name,
        repo.project_url or "",
        MODE_LABELS.get(repo.mode, repo.mode),
        DOMAIN_LABELS.get(repo.domain, repo.domain),
        repo_type_label,
        "、".join(group.name for group in responsibility_groups),
        int(getattr(repo, "branch_count", 0) or 0),
        repo.remark or "",
        _format_export_datetime(repo.sys_create_datetime),
        _format_export_datetime(repo.sys_update_datetime),
    ]


def _build_repository_export_workbook(payload: dict, task_id: str) -> openpyxl.Workbook:
    """按任务 payload 生成组织+代码库 Excel。"""
    qs = _repository_export_queryset(payload)
    total = qs.count()
    workbook = openpyxl.Workbook(write_only=True)
    worksheet = workbook.create_sheet(title="组织代码库清单")
    worksheet.append(REPOSITORY_EXPORT_HEADERS)
    path_map = _organization_path_map()
    repo_type_map = _repo_type_label_map()
    for index, repo in enumerate(qs.iterator(chunk_size=500), start=1):
        worksheet.append(_repository_export_row(repo, path_map, repo_type_map))
        if index == 1 or index == total or index % 200 == 0:
            progress = 10 + int((index / max(total, 1)) * 80)
            _update_repository_export_task_progress(
                task_id,
                message=f"正在生成导出文件：{index}/{total}",
                progress=min(progress, 95),
            )
    if total == 0:
        _update_repository_export_task_progress(task_id, message="暂无匹配代码库，正在生成空文件", progress=90)
    return workbook


def _update_repository_export_task_progress(task_id: str, *, message: str, progress: int):
    """更新导出任务进度，限制成功前最大进度为 99。"""
    ComplianceRepositoryExportTask.objects.filter(id=task_id).update(
        message=message,
        progress=max(0, min(int(progress or 0), 99)),
    )


def _run_repository_export_task(task_id: str):
    """后台线程执行代码库导出任务。"""
    close_old_connections()
    generated_file_path: Optional[Path] = None
    try:
        task = ComplianceRepositoryExportTask.objects.filter(id=task_id, is_deleted=False).first()
        if task is None:
            return
        ComplianceRepositoryExportTask.objects.filter(id=task_id).update(
            status=ComplianceRepositoryExportTask.STATUS_RUNNING,
            progress=3,
            message="正在准备导出数据",
            error_message="",
            started_at=timezone.now(),
            finished_at=None,
            file_path="",
            file_name="",
            file_size=0,
        )
        workbook = _build_repository_export_workbook(task.payload or {}, task_id)
        timestamp = timezone.now().strftime("%Y%m%d-%H%M%S")
        scope_label = "all" if task.scope == REPOSITORY_EXPORT_SCOPE_ALL else "filtered"
        file_name = f"code_compliance_repositories_{scope_label}_{timestamp}_{str(task.id)[:8]}.xlsx"
        generated_file_path = _repository_export_temp_dir() / file_name
        workbook.save(str(generated_file_path))
        file_size = generated_file_path.stat().st_size if generated_file_path.exists() else 0
        ComplianceRepositoryExportTask.objects.filter(id=task_id).update(
            status=ComplianceRepositoryExportTask.STATUS_SUCCESS,
            progress=100,
            message="导出文件生成完成",
            error_message="",
            file_path=str(generated_file_path),
            file_name=file_name,
            file_size=file_size,
            finished_at=timezone.now(),
        )
    except Exception as exc:
        if generated_file_path and generated_file_path.exists():
            try:
                generated_file_path.unlink()
            except Exception:
                logger.warning("Remove repository export temp file failed task_id=%s", task_id, exc_info=True)
        logger.exception("Repository export task failed: task_id=%s", task_id)
        ComplianceRepositoryExportTask.objects.filter(id=task_id).update(
            status=ComplianceRepositoryExportTask.STATUS_FAILED,
            message="导出任务失败",
            error_message=str(exc),
            finished_at=timezone.now(),
        )
    finally:
        connection.close()


def _start_repository_export_task_thread(task_id: str):
    """启动进程内后台线程执行导出。"""
    thread = threading.Thread(
        target=_run_repository_export_task,
        args=(task_id,),
        daemon=True,
    )
    thread.start()


def prepare_repository_export_task(user, payload) -> dict:
    """创建或复用组织+代码库异步导出任务。"""
    if not user or not getattr(user, "id", None):
        raise HttpError(401, "用户未登录")
    _cleanup_repository_export_files(limit=200)
    normalized_payload = _normalize_repository_export_payload(payload)
    fingerprint = _repository_export_fingerprint(normalized_payload)
    active_task = _get_active_repository_export_task(user, fingerprint)
    if active_task is not None:
        return {"mode": "async", "task": _serialize_repository_export_task(active_task)}
    reusable_task = _get_reusable_repository_export_task(user, fingerprint)
    if reusable_task is not None:
        return {"mode": "ready", "task": _serialize_repository_export_task(reusable_task)}

    task = ComplianceRepositoryExportTask.objects.create(
        user=user,
        sys_creator=user,
        scope=normalized_payload["scope"],
        fingerprint=fingerprint,
        payload=normalized_payload,
        status=ComplianceRepositoryExportTask.STATUS_PENDING,
        progress=0,
        message="导出任务已提交，正在排队执行",
    )
    _start_repository_export_task_thread(str(task.id))
    return {"mode": "async", "task": _serialize_repository_export_task(task)}


def get_repository_export_task(user, task_id: str) -> dict:
    """查询当前用户的导出任务状态。"""
    if not user or not getattr(user, "id", None):
        raise HttpError(401, "用户未登录")
    task = ComplianceRepositoryExportTask.objects.filter(
        id=task_id,
        user=user,
        is_deleted=False,
    ).first()
    if task is None:
        raise HttpError(404, "导出任务不存在")
    return _serialize_repository_export_task(task)


def download_repository_export_task_file(user, task_id: str) -> FileResponse:
    """下载当前用户已完成的导出文件。"""
    if not user or not getattr(user, "id", None):
        raise HttpError(401, "用户未登录")
    _cleanup_repository_export_files(limit=200)
    task = ComplianceRepositoryExportTask.objects.filter(
        id=task_id,
        user=user,
        is_deleted=False,
    ).first()
    if task is None:
        raise HttpError(404, "导出任务不存在")
    if task.status != ComplianceRepositoryExportTask.STATUS_SUCCESS:
        raise HttpError(409, "导出任务尚未完成")
    if _is_repository_export_task_expired(task):
        raise HttpError(410, "导出文件已过期，请重新导出")
    file_path = _clean_text(task.file_path)
    if not file_path:
        raise HttpError(404, "导出文件不存在")
    path = Path(file_path)
    if not path.is_file():
        raise HttpError(404, "导出文件不存在")
    return FileResponse(
        path.open("rb"),
        as_attachment=True,
        filename=_clean_text(task.file_name) or path.name,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def get_repository(repo_id: str) -> dict:
    """读取单个代码库详情。"""
    item = get_object_or_404(_repository_queryset(), id=repo_id)
    return serialize_repository(item)


def get_repository_branches(repo_id: str) -> dict:
    """读取代码库绑定的分支列表，供分支演进 Dialog 使用。"""
    repository = get_object_or_404(_repository_queryset(), id=repo_id)
    links = (
        ComplianceRepositoryBranch.objects.filter(
            repository=repository,
            is_deleted=False,
            branch__is_deleted=False,
        )
        .select_related("branch")
        .order_by("branch__branch_name")
    )
    branches = [link.branch for link in links]
    # Python 排序用于兼容 MySQL 对 NULLS LAST 的表达差异。
    branches.sort(
        key=lambda item: (
            item.created_date is None,
            item.created_date or date.max,
            item.branch_name,
        )
    )
    return {
        "repository": serialize_repository(repository),
        "branches": [serialize_branch(item) for item in branches],
    }


@transaction.atomic
def create_repository(user, payload) -> dict:
    """创建代码库并写入责任 PL 组多对多关系。"""
    data = payload.dict()
    project_id = _clean_text(data["project_id"])
    project_name = _clean_text(data["project_name"])
    if not project_id:
        raise HttpError(400, "代码库ID不能为空")
    if not project_name:
        raise HttpError(400, "代码库名不能为空")

    organization = _get_active_organization(data["organization_id"])
    repo_type = _ensure_repo_type_valid(data.get("repo_type"))
    pl_groups = _get_pl_groups(data.get("responsibility_group_ids") or [])
    existing = ComplianceRepository.objects.filter(project_id=project_id).first()
    if existing and not existing.is_deleted:
        raise HttpError(400, f"代码库ID已存在: {project_id}")

    item = existing or ComplianceRepository(project_id=project_id)
    item.project_id = project_id
    item.project_name = project_name
    item.project_url = _clean_text(data.get("project_url"))
    item.organization = organization
    item.mode = _normalize_mode(data.get("mode"))
    item.repo_type = repo_type
    item.domain = _normalize_domain(data.get("domain"))
    item.remark = _optional_text(data.get("remark"))
    item.sort = int(data.get("sort") or 0)
    item.is_deleted = False
    _apply_audit_fields(item, user, is_create=not existing)
    item.save()
    item.responsibility_groups.set(pl_groups)
    return get_repository(str(item.id))


@transaction.atomic
def update_repository(user, repo_id: str, payload) -> dict:
    """更新代码库字段，责任 PL 组仅在请求显式传入时覆盖。"""
    item = _get_active_repository(repo_id)
    data = payload.dict(exclude_unset=True)

    if "project_id" in data:
        project_id = _clean_text(data["project_id"])
        if not project_id:
            raise HttpError(400, "代码库ID不能为空")
        exists = (
            ComplianceRepository.objects.filter(project_id=project_id, is_deleted=False)
            .exclude(id=repo_id)
            .exists()
        )
        if exists:
            raise HttpError(400, f"代码库ID已存在: {project_id}")
        item.project_id = project_id
    if "project_name" in data:
        project_name = _clean_text(data["project_name"])
        if not project_name:
            raise HttpError(400, "代码库名不能为空")
        item.project_name = project_name
    if "project_url" in data:
        item.project_url = _clean_text(data.get("project_url"))
    if "organization_id" in data:
        item.organization = _get_active_organization(data["organization_id"])
    if "mode" in data:
        item.mode = _normalize_mode(data.get("mode"))
    if "repo_type" in data:
        item.repo_type = _ensure_repo_type_valid(data.get("repo_type"))
    if "domain" in data:
        item.domain = _normalize_domain(data.get("domain"))
    if "remark" in data:
        item.remark = _optional_text(data.get("remark"))
    if "sort" in data:
        item.sort = int(data.get("sort") or 0)

    _apply_audit_fields(item, user)
    item.save()
    if "responsibility_group_ids" in data:
        item.responsibility_groups.set(_get_pl_groups(data.get("responsibility_group_ids") or []))
    return get_repository(str(item.id))


@transaction.atomic
def delete_repository(repo_id: str) -> dict:
    """软删除代码库，同时软删除已有绑定关系。"""
    item = _get_active_repository(repo_id)
    item.branch_links.filter(is_deleted=False).update(is_deleted=True)
    item.soft_delete()
    return {"id": str(item.id)}


def _branch_queryset():
    """返回分支列表基础查询，集中挂载关联代码库统计。"""
    return _active_branches().annotate(
        repository_count=Count(
            "repository_links",
            filter=Q(
                repository_links__is_deleted=False,
                repository_links__repository__is_deleted=False,
            ),
            distinct=True,
        )
    )


def list_branches(
    *,
    page: int = 1,
    page_size: int = 20,
    keyword: Optional[str] = None,
    branch_type: Optional[str] = None,
    domain: Optional[str] = None,
    is_active: Optional[bool] = None,
) -> dict:
    """按页面筛选条件分页查询分支主数据。"""
    qs = _branch_queryset()
    if keyword:
        word = keyword.strip()
        qs = qs.filter(
            Q(branch_name__icontains=word)
            | Q(alias__icontains=word)
            | Q(purpose__icontains=word)
        )
    if branch_type:
        qs = qs.filter(branch_type=_normalize_branch_type(branch_type))
    if domain:
        qs = qs.filter(domain=_normalize_domain(domain))
    if is_active is not None:
        qs = qs.filter(is_active=bool(is_active))

    total = qs.count()
    offset = max(page - 1, 0) * page_size
    items = list(qs.order_by("sort", "domain", "branch_name")[offset : offset + page_size])
    return {"items": [serialize_branch(item) for item in items], "total": total}


def get_branch(branch_id: str) -> dict:
    """读取单个分支详情。"""
    item = get_object_or_404(_branch_queryset(), id=branch_id)
    return serialize_branch(item)


def get_branch_repositories(branch_id: str) -> dict:
    """读取分支绑定的组织树和代码库列表，供关联仓库 Dialog 使用。"""
    branch = get_object_or_404(_branch_queryset(), id=branch_id)
    links = (
        ComplianceRepositoryBranch.objects.filter(
            branch=branch,
            is_deleted=False,
            repository__is_deleted=False,
            repository__organization__is_deleted=False,
        )
        .select_related("repository__organization")
        .prefetch_related("repository__responsibility_groups")
    )
    repositories = [link.repository for link in links]
    organizations = _build_repository_organization_tree(repositories)
    return {
        "branch": serialize_branch(branch),
        "organizations": organizations,
    }


def _build_repository_organization_tree(repositories: list[ComplianceRepository]) -> list[dict]:
    """按绑定代码库归属组织构建只包含相关路径的组织树。"""
    if not repositories:
        return []

    repo_map: dict[str, list[dict]] = {}
    organization_ids = {str(repo.organization_id) for repo in repositories}
    for repo in repositories:
        repo_map.setdefault(str(repo.organization_id), []).append(serialize_repository(repo))

    # 补齐所有命中组织的祖先节点，让左侧树能表达完整组织路径。
    all_orgs = {
        str(item.id): item
        for item in _active_organizations().select_related("parent").order_by("sort", "name")
    }
    needed_ids: set[str] = set()
    for org_id in organization_ids:
        current = all_orgs.get(org_id)
        while current:
            current_id = str(current.id)
            if current_id in needed_ids:
                break
            needed_ids.add(current_id)
            current = all_orgs.get(str(current.parent_id)) if current.parent_id else None

    nodes = {org_id: all_orgs[org_id] for org_id in needed_ids if org_id in all_orgs}
    children_map: dict[str | None, list[ComplianceOrganization]] = {}
    for node in nodes.values():
        parent_id = str(node.parent_id) if node.parent_id and str(node.parent_id) in nodes else None
        children_map.setdefault(parent_id, []).append(node)

    for siblings in children_map.values():
        siblings.sort(key=lambda item: (item.sort, item.name))

    def build_node(item: ComplianceOrganization) -> dict:
        """递归序列化相关组织节点，并挂载直接绑定代码库。"""
        item.repository_count = len(repo_map.get(str(item.id), []))
        payload = serialize_organization(item, include_children=False)
        payload["children"] = [build_node(child) for child in children_map.get(str(item.id), [])]
        payload["repositories"] = sorted(
            repo_map.get(str(item.id), []),
            key=lambda repo: (repo["sort"], repo["project_name"]),
        )
        return payload

    return [build_node(item) for item in children_map.get(None, [])]


@transaction.atomic
def create_branch(user, payload) -> dict:
    """创建分支主数据，同一领域内分支名称保持唯一。"""
    data = payload.dict()
    branch_name = _clean_text(data["branch_name"])
    if not branch_name:
        raise HttpError(400, "分支名称不能为空")
    domain = _normalize_domain(data.get("domain"))

    existing = ComplianceManagedBranch.objects.filter(
        branch_name=branch_name,
        domain=domain,
    ).first()
    if existing and not existing.is_deleted:
        raise HttpError(400, f"该领域下分支已存在: {branch_name}")

    item = existing or ComplianceManagedBranch(branch_name=branch_name, domain=domain)
    item.branch_name = branch_name
    item.domain = domain
    item.created_date = _parse_date(data.get("created_date"))
    item.branch_type = _normalize_branch_type(data.get("branch_type"))
    item.alias = _clean_text(data.get("alias"))
    item.purpose = _clean_text(data.get("purpose"))
    item.remark = _optional_text(data.get("remark"))
    item.is_active = _normalize_branch_active(data.get("is_active"), default=True)
    item.sort = int(data.get("sort") or 0)
    item.is_deleted = False
    _apply_audit_fields(item, user, is_create=not existing)
    item.save()
    return get_branch(str(item.id))


@transaction.atomic
def update_branch(user, branch_id: str, payload) -> dict:
    """更新分支主数据，并校验领域内分支名唯一。"""
    item = _get_active_branch(branch_id)
    data = payload.dict(exclude_unset=True)

    next_name = _clean_text(data.get("branch_name")) if "branch_name" in data else item.branch_name
    next_domain = _normalize_domain(data.get("domain")) if "domain" in data else item.domain
    if not next_name:
        raise HttpError(400, "分支名称不能为空")
    exists = (
        ComplianceManagedBranch.objects.filter(
            branch_name=next_name,
            domain=next_domain,
            is_deleted=False,
        )
        .exclude(id=branch_id)
        .exists()
    )
    if exists:
        raise HttpError(400, f"该领域下分支已存在: {next_name}")

    item.branch_name = next_name
    item.domain = next_domain
    if "created_date" in data:
        item.created_date = _parse_date(data.get("created_date"))
    if "branch_type" in data:
        item.branch_type = _normalize_branch_type(data.get("branch_type"))
    if "alias" in data:
        item.alias = _clean_text(data.get("alias"))
    if "purpose" in data:
        item.purpose = _clean_text(data.get("purpose"))
    if "remark" in data:
        item.remark = _optional_text(data.get("remark"))
    if "is_active" in data:
        item.is_active = _normalize_branch_active(data.get("is_active"), default=item.is_active)
    if "sort" in data:
        item.sort = int(data.get("sort") or 0)

    _apply_audit_fields(item, user)
    item.save()
    return get_branch(str(item.id))


@transaction.atomic
def delete_branch(branch_id: str) -> dict:
    """软删除分支，同时软删除已有绑定关系。"""
    item = _get_active_branch(branch_id)
    item.repository_links.filter(is_deleted=False).update(is_deleted=True)
    item.soft_delete()
    return {"id": str(item.id)}


def _normalize_bind_mode(mode: Optional[str]) -> str:
    """规范化批量绑定模式。"""
    value = _clean_text(mode).lower() or BIND_MODE_APPEND
    if value not in SUPPORTED_BIND_MODES:
        raise HttpError(400, "批量绑定模式仅支持 append 或 replace")
    return value


def _load_active_repositories(ids: Iterable[str]) -> list[ComplianceRepository]:
    """批量读取活跃代码库，并返回缺失 ID 的明确错误。"""
    id_list = [str(item) for item in ids if str(item or "").strip()]
    repos = list(_active_repositories().filter(id__in=id_list))
    found = {str(item.id) for item in repos}
    missing = [item for item in id_list if item not in found]
    if missing:
        raise HttpError(404, f"代码库不存在: {', '.join(missing)}")
    return repos


def _load_active_branches(ids: Iterable[str]) -> list[ComplianceManagedBranch]:
    """批量读取活跃分支，并返回缺失 ID 的明确错误。"""
    id_list = [str(item) for item in ids if str(item or "").strip()]
    branches = list(_active_branches().filter(id__in=id_list))
    found = {str(item.id) for item in branches}
    missing = [item for item in id_list if item not in found]
    if missing:
        raise HttpError(404, f"分支不存在: {', '.join(missing)}")
    return branches


def _ensure_link(repository, branch, counters: BindCounters):
    """创建或恢复单个代码库-分支绑定。"""
    # Soft-deleted link rows still occupy the unique key, so restore them instead of creating duplicates.
    link = ComplianceRepositoryBranch.objects.filter(
        repository=repository,
        branch=branch,
    ).first()
    if link:
        if link.is_deleted:
            link.is_deleted = False
            link.save(update_fields=["is_deleted", "sys_update_datetime"])
            counters.restored_count += 1
        else:
            counters.ignored_count += 1
        return
    ComplianceRepositoryBranch.objects.create(repository=repository, branch=branch)
    counters.created_count += 1


@transaction.atomic
def bind_branches_to_repositories(repository_ids, branch_ids, mode: str) -> dict:
    """从代码库侧批量绑定分支，replace 会移除未选中的旧绑定。"""
    normalized_mode = _normalize_bind_mode(mode)
    repositories = _load_active_repositories(repository_ids)
    branches = _load_active_branches(branch_ids)
    counters = BindCounters()

    branch_id_set = {item.id for item in branches}
    if normalized_mode == BIND_MODE_REPLACE:
        # replace 只替换选中代码库的绑定范围，不影响其他代码库。
        for repository in repositories:
            qs = repository.branch_links.filter(is_deleted=False)
            if branch_id_set:
                qs = qs.exclude(branch_id__in=branch_id_set)
            removed = qs.update(is_deleted=True)
            counters.removed_count += removed

    for repository in repositories:
        for branch in branches:
            _ensure_link(repository, branch, counters)
    return counters.as_dict()


@transaction.atomic
def bind_repositories_to_branches(branch_ids, repository_ids, mode: str) -> dict:
    """从分支侧批量绑定代码库，replace 会移除未选中的旧绑定。"""
    normalized_mode = _normalize_bind_mode(mode)
    branches = _load_active_branches(branch_ids)
    repositories = _load_active_repositories(repository_ids)
    counters = BindCounters()

    repository_id_set = {item.id for item in repositories}
    if normalized_mode == BIND_MODE_REPLACE:
        # replace 只替换选中分支的绑定范围，不影响其他分支。
        for branch in branches:
            qs = branch.repository_links.filter(is_deleted=False)
            if repository_id_set:
                qs = qs.exclude(repository_id__in=repository_id_set)
            removed = qs.update(is_deleted=True)
            counters.removed_count += removed

    for branch in branches:
        for repository in repositories:
            _ensure_link(repository, branch, counters)
    return counters.as_dict()


def _load_excel_rows(file_obj) -> list[tuple[int, dict[str, object]]]:
    """读取 .xlsx 首个 sheet，并保留原始行号用于错误提示。"""
    filename = (getattr(file_obj, "name", "") or "").lower()
    if not filename.endswith(".xlsx"):
        raise HttpError(400, "仅支持 .xlsx 文件")
    try:
        content = file_obj.read()
        workbook = openpyxl.load_workbook(
            filename=io.BytesIO(content),
            read_only=True,
            data_only=True,
        )
    except Exception as exc:
        raise HttpError(400, f"Excel 解析失败: {exc}")

    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise HttpError(400, "Excel 内容为空")
    header = [_clean_text(item) for item in rows[0]]
    parsed: list[tuple[int, dict[str, object]]] = []
    for row_no, row in enumerate(rows[1:], start=2):
        if not row or not any(_clean_text(item) for item in row):
            continue
        # 通过表头映射单元格，允许后续按中文或英文别名取值。
        row_map = {
            header[index]: row[index] if index < len(row) else None
            for index in range(len(header))
            if header[index]
        }
        parsed.append((row_no, row_map))
    return parsed


def _cell(row: dict[str, object], *names: str):
    """按多个候选表头取单元格值。"""
    for name in names:
        if name in row:
            return row[name]
    return None


def _same_values(instance, values: dict) -> bool:
    """判断导入行是否与现有记录一致，用于统计忽略行。"""
    for field, value in values.items():
        if getattr(instance, field) != value:
            return False
    return True


@transaction.atomic
def import_organizations(user, file_obj) -> ImportResultOut:
    """导入组织基础字段，按外部 group_id 新增或更新。"""
    rows = _load_excel_rows(file_obj)
    created = updated = ignored = 0
    errors: list[ImportErrorRow] = []

    for row_no, row in rows:
        try:
            group_id = _clean_text(_cell(row, "组织ID", "group_id"))
            name = _clean_text(_cell(row, "组织名", "name"))
            parent_group_id = _clean_text(_cell(row, "父组织ID", "parent_group_id"))
            if not group_id:
                raise HttpError(400, "组织ID不能为空")
            if not name:
                raise HttpError(400, "组织名不能为空")

            parent = None
            if parent_group_id:
                # Excel 使用公司代码库系统的父组织 ID，而不是本系统主键。
                parent = _active_organizations().filter(group_id=parent_group_id).first()
                if not parent:
                    raise HttpError(400, f"父组织不存在: {parent_group_id}")
            item = ComplianceOrganization.objects.filter(group_id=group_id).first()
            is_create = item is None
            if item is None:
                item = ComplianceOrganization(group_id=group_id)
            _ensure_org_parent_valid(org_id=str(item.id) if item.id else None, parent=parent)
            values = {
                "group_id": group_id,
                "name": name,
                "parent": parent,
                "mode": _normalize_mode(_cell(row, "模式", "mode")),
                "domain": _normalize_domain(_cell(row, "领域", "domain")),
                "remark": _optional_text(_cell(row, "备注", "remark")),
                "is_deleted": False,
            }
            if not is_create and not item.is_deleted and _same_values(item, values):
                ignored += 1
                continue
            # 已软删除的同 group_id 记录会被重新启用，避免唯一键冲突。
            for field, value in values.items():
                setattr(item, field, value)
            _apply_audit_fields(item, user, is_create=is_create)
            item.save()
            created += 1 if is_create else 0
            updated += 0 if is_create else 1
        except Exception as exc:
            errors.append(ImportErrorRow(row_no=row_no, message=_error_message(exc)))

    return ImportResultOut(
        created_count=created,
        updated_count=updated,
        ignored_count=ignored,
        errors=errors,
    )


@transaction.atomic
def import_repositories(user, file_obj) -> ImportResultOut:
    """导入代码库基础字段，按外部 project_id 新增或更新。"""
    rows = _load_excel_rows(file_obj)
    created = updated = ignored = 0
    errors: list[ImportErrorRow] = []

    for row_no, row in rows:
        try:
            project_id = _clean_text(_cell(row, "代码库ID", "project_id"))
            project_name = _clean_text(_cell(row, "代码库名", "project_name"))
            org_group_id = _clean_text(_cell(row, "组织ID", "organization_group_id"))
            if not project_id:
                raise HttpError(400, "代码库ID不能为空")
            if not project_name:
                raise HttpError(400, "代码库名不能为空")
            # Excel 中组织字段使用外部 group_id，便于从公司代码库系统导出的表直接导入。
            organization = _active_organizations().filter(group_id=org_group_id).first()
            if not organization:
                raise HttpError(400, f"所属组织不存在: {org_group_id}")
            repo_type = _ensure_repo_type_valid(_cell(row, "代码仓类型", "repo_type"))
            pl_groups = _resolve_pl_groups_from_text(_cell(row, "责任PL组", "责任领域", "responsibility_groups"))

            item = ComplianceRepository.objects.filter(project_id=project_id).first()
            is_create = item is None
            if item is None:
                item = ComplianceRepository(project_id=project_id)
            values = {
                "project_id": project_id,
                "project_name": project_name,
                "project_url": _clean_text(_cell(row, "代码库URL", "project_url")),
                "organization": organization,
                "mode": _normalize_mode(_cell(row, "模式", "mode")),
                "repo_type": repo_type,
                "domain": _normalize_domain(_cell(row, "领域", "domain")),
                "remark": _optional_text(_cell(row, "备注", "remark")),
                "is_deleted": False,
            }
            current_group_ids = set(item.responsibility_groups.values_list("id", flat=True)) if item.id else set()
            next_group_ids = {group.id for group in pl_groups}
            if (
                not is_create
                and not item.is_deleted
                and _same_values(item, values)
                and current_group_ids == next_group_ids
            ):
                ignored += 1
                continue
            # 责任 PL 组是多对多关系，需要在主记录保存后单独 set。
            for field, value in values.items():
                setattr(item, field, value)
            _apply_audit_fields(item, user, is_create=is_create)
            item.save()
            item.responsibility_groups.set(pl_groups)
            created += 1 if is_create else 0
            updated += 0 if is_create else 1
        except Exception as exc:
            errors.append(ImportErrorRow(row_no=row_no, message=_error_message(exc)))

    return ImportResultOut(
        created_count=created,
        updated_count=updated,
        ignored_count=ignored,
        errors=errors,
    )


@transaction.atomic
def import_branches(user, file_obj) -> ImportResultOut:
    """导入分支基础字段，按领域 + 分支名称新增或更新。"""
    rows = _load_excel_rows(file_obj)
    created = updated = ignored = 0
    errors: list[ImportErrorRow] = []

    for row_no, row in rows:
        try:
            branch_name = _clean_text(_cell(row, "分支名称", "branch_name"))
            if not branch_name:
                raise HttpError(400, "分支名称不能为空")
            domain = _normalize_domain(_cell(row, "领域", "domain"))
            item = ComplianceManagedBranch.objects.filter(
                branch_name=branch_name,
                domain=domain,
            ).first()
            is_create = item is None
            if item is None:
                item = ComplianceManagedBranch(branch_name=branch_name, domain=domain)
            values = {
                "branch_name": branch_name,
                "domain": domain,
                "created_date": _parse_date(_cell(row, "创建日期", "created_date")),
                "branch_type": _normalize_branch_type(_cell(row, "分支类型", "branch_type")),
                "alias": _clean_text(_cell(row, "分支别名", "alias")),
                "purpose": _clean_text(_cell(row, "分支用途", "purpose")),
                "remark": _optional_text(_cell(row, "备注", "remark")),
                "is_active": _normalize_branch_active(_cell(row, "是否活跃", "is_active"), default=True),
                "is_deleted": False,
            }
            if not is_create and not item.is_deleted and _same_values(item, values):
                ignored += 1
                continue
            # 分支导入只维护主数据，不在这里处理代码库绑定关系。
            for field, value in values.items():
                setattr(item, field, value)
            _apply_audit_fields(item, user, is_create=is_create)
            item.save()
            created += 1 if is_create else 0
            updated += 0 if is_create else 1
        except Exception as exc:
            errors.append(ImportErrorRow(row_no=row_no, message=_error_message(exc)))

    return ImportResultOut(
        created_count=created,
        updated_count=updated,
        ignored_count=ignored,
        errors=errors,
    )


def _error_message(exc: Exception) -> str:
    """把导入过程中捕获到的异常转换为前端可读文本。"""
    if isinstance(exc, HttpError):
        return str(getattr(exc, "message", "") or exc)
    return str(exc)


def _build_workbook_response(workbook, filename: str) -> HttpResponse:
    """把 openpyxl workbook 包装成浏览器下载响应。"""
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f"attachment; filename*=UTF-8''{filename}"
    workbook.save(response)
    return response


def build_organization_template_response() -> HttpResponse:
    """构建组织导入模板。"""
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "组织导入模板"
    sheet.append(["组织ID", "组织名", "父组织ID", "模式", "领域", "备注"])
    sheet.append(["10001", "示例组织", "", "CR", "cockpit", ""])
    sheet.append(["10002", "示例子组织", "10001", "MR", "vehicle", ""])
    return _build_workbook_response(workbook, "code_compliance_organization_template.xlsx")


def build_repository_template_response() -> HttpResponse:
    """构建代码库导入模板。"""
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "代码库导入模板"
    sheet.append(["代码库ID", "代码库名", "代码库URL", "组织ID", "模式", "领域", "代码仓类型", "责任PL组", "备注"])
    sheet.append(["20001", "demo-repo", "https://git.example.com/demo-repo", "10001", "CR", "cockpit", "", "PL组A,PL组B", ""])
    return _build_workbook_response(workbook, "code_compliance_repository_template.xlsx")


def build_branch_template_response() -> HttpResponse:
    """构建分支导入模板。"""
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "分支导入模板"
    sheet.append(["分支名称", "创建日期", "分支类型", "分支别名", "分支用途", "领域", "是否活跃", "备注"])
    sheet.append(["master", "2026-01-01", "主干", "主线", "主干开发", "cockpit", "是", ""])
    return _build_workbook_response(workbook, "code_compliance_branch_template.xlsx")
