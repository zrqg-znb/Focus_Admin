from __future__ import annotations

import hashlib
import json
import math
import tempfile
import threading
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any

import openpyxl
from django.conf import settings
from django.db import close_old_connections, transaction
from django.db.models import Count, Q, Sum
from django.http import FileResponse, HttpResponse
from django.utils import dateparse
from django.shortcuts import get_object_or_404
from django.utils import timezone
from ninja.errors import HttpError

from core.pl.pl_model import PlGroup
from core.user.user_model import User
from scheduler.module.executor import scheduler_task

from . import base_services
from .contribution_client import CodeComplianceCRClient, CodeComplianceMRClient, DEFAULT_PAGE_SIZE
from .models import (
    COMPLIANCE_MODE_CR,
    COMPLIANCE_MODE_MR,
    COMPLIANCE_BRANCH_TYPE_CHOICES,
    COMPLIANCE_DOMAIN_CHOICES,
    CONTRIBUTION_EXPORT_SCOPE_RECORDS,
    CONTRIBUTION_EXPORT_SCOPE_SUMMARY,
    CONTRIBUTION_BASELINE_SOURCE_IMPORT,
    CONTRIBUTION_BASELINE_SOURCE_MANUAL,
    CONTRIBUTION_TASK_STATUS_FAILED,
    CONTRIBUTION_TASK_STATUS_PENDING,
    CONTRIBUTION_TASK_STATUS_RUNNING,
    CONTRIBUTION_TASK_STATUS_SUCCESS,
    CONTRIBUTION_TASK_TRIGGER_BACKFILL,
    CONTRIBUTION_TASK_TRIGGER_CHOICES,
    CONTRIBUTION_TASK_TRIGGER_MANUAL,
    CONTRIBUTION_TASK_TRIGGER_SCHEDULED,
    ComplianceContributionCodeBaseline,
    ComplianceContributionCollectTask,
    ComplianceContributionDailyAggregate,
    ComplianceContributionExportTask,
    ComplianceContributionRecord,
    ComplianceManagedBranch,
    ComplianceRepository,
    ComplianceRepositoryBranch,
)


UNKNOWN_PL_GROUP_ID = "unknown"
UNKNOWN_PL_GROUP_NAME = "非底软领域"
CONTRIBUTION_SOURCE_MODES = {COMPLIANCE_MODE_CR, COMPLIANCE_MODE_MR}
MR_COLLECT_MAX_WORKERS = 5
TASK_STATUS_LABELS = {
    CONTRIBUTION_TASK_STATUS_PENDING: "待执行",
    CONTRIBUTION_TASK_STATUS_RUNNING: "执行中",
    CONTRIBUTION_TASK_STATUS_SUCCESS: "成功",
    CONTRIBUTION_TASK_STATUS_FAILED: "失败",
}
TASK_TRIGGER_LABELS = dict(CONTRIBUTION_TASK_TRIGGER_CHOICES)
BRANCH_TYPE_LABELS = dict(COMPLIANCE_BRANCH_TYPE_CHOICES)
DOMAIN_LABELS = dict(COMPLIANCE_DOMAIN_CHOICES)
EXPORT_ACTIVE_STATUSES = {CONTRIBUTION_TASK_STATUS_PENDING, CONTRIBUTION_TASK_STATUS_RUNNING}
EXPORT_TTL_SECONDS = 24 * 60 * 60
BASELINE_SOURCE_LABELS = {
    CONTRIBUTION_BASELINE_SOURCE_MANUAL: "手工维护",
    CONTRIBUTION_BASELINE_SOURCE_IMPORT: "Excel导入",
}


@dataclass(frozen=True)
class AuthorAssignment:
    user_id: str | None
    user_name: str
    pl_group_id: str | None
    pl_group_name: str


def _clean_text(value: Any) -> str:
    """把页面参数、上游字段统一转换成去空格字符串。"""
    if value is None:
        return ""
    return str(value).strip()


def _normalize_id_list(values: Any) -> list[str]:
    """兼容逗号字符串、数组和空值，输出去重后的 ID 列表。"""
    if not values:
        return []
    candidates = values.split(",") if isinstance(values, str) else values
    result: list[str] = []
    seen: set[str] = set()
    for item in candidates:
        value = _clean_text(item)
        if value and value not in seen:
            seen.add(value)
            result.append(value)
    return result


def _default_assignment() -> AuthorAssignment:
    """未匹配 Focus 用户或启用 PL 组时统一归到非底软领域。"""
    return AuthorAssignment(None, "", None, UNKNOWN_PL_GROUP_NAME)


def _load_author_assignments(usernames: list[str]) -> dict[str, AuthorAssignment]:
    """批量解析 CR 创建人的 Focus 用户与 PL 组归属，避免逐条查询。"""
    unique_usernames = _normalize_id_list(usernames)
    users: dict[str, dict[str, Any]] = {}
    # 历史数据若存在重复工号，优先使用最近创建的 Focus 用户，保证归属确定。
    for row in User.objects.filter(username__in=unique_usernames).order_by(
        "username", "-sys_create_datetime", "-id"
    ).values("id", "username", "name"):
        users.setdefault(row["username"], row)
    group_rows = (
        PlGroup.objects.filter(status=True, members__username__in=unique_usernames)
        .values("id", "name", "sort", "members__username")
        .order_by("-sort", "name", "id")
    )
    group_by_username: dict[str, dict[str, Any]] = {}
    for row in group_rows:
        username = _clean_text(row.get("members__username"))
        if username and username not in group_by_username:
            group_by_username[username] = row

    result: dict[str, AuthorAssignment] = {}
    for username in unique_usernames:
        user_row = users.get(username)
        group_row = group_by_username.get(username)
        if not user_row:
            result[username] = _default_assignment()
            continue
        result[username] = AuthorAssignment(
            user_id=str(user_row["id"]),
            user_name=_clean_text(user_row.get("name")) or username,
            pl_group_id=str(group_row["id"]) if group_row else None,
            pl_group_name=_clean_text(group_row.get("name")) if group_row else UNKNOWN_PL_GROUP_NAME,
        )
    return result


def _author_display_name(user_name: str, username: str) -> str:
    """按产品口径展示姓名（工号），未匹配姓名时保留原始工号。"""
    user_name = _clean_text(user_name)
    username = _clean_text(username)
    if user_name and username:
        return f"{user_name}（{username}）"
    return username or "-"


def _date_from_datetime(value: datetime | None) -> datetime.date:
    """贡献日期优先取 merged_at 的本地日期，空值兜底当前日期。"""
    if value:
        return timezone.localtime(value).date() if timezone.is_aware(value) else value.date()
    return timezone.localdate()


def _to_model_datetime(value: Any):
    """根据 USE_TZ 配置把数据湖或前端时间转换成 MySQL 可保存形态。"""
    if value is None:
        return None
    if isinstance(value, str):
        value = dateparse.parse_datetime(value) or dateparse.parse_date(value)
        if value is None:
            return None
        if not isinstance(value, datetime):
            value = datetime.combine(value, datetime.min.time())
    if settings.USE_TZ:
        if timezone.is_naive(value):
            return timezone.make_aware(value, timezone.get_current_timezone())
        return value
    if timezone.is_aware(value):
        return timezone.localtime(value, timezone.get_current_timezone()).replace(tzinfo=None)
    return value


def _line_metrics(row: dict[str, Any]) -> tuple[int, int, int, int]:
    """从 CR 明细计算新增、删除、净增和总变更行数。"""
    added = int(row.get("added_lines") or 0)
    removed = int(row.get("removed_lines") or 0)
    return added, removed, added - removed, added + removed


def _filter_payload_from_kwargs(**kwargs) -> dict:
    """把 API 查询参数压缩成长期存储和任务指纹使用的 payload。"""
    payload: dict[str, Any] = {}
    for key, value in kwargs.items():
        if value in (None, "", [], {}):
            continue
        payload[key] = value
    return payload


def _normalize_source_mode(value: Any, *, allow_empty: bool = True) -> str:
    """规范化贡献来源模式，空值表示 CR 与 MR 汇总查询。"""
    mode = _clean_text(value).upper()
    if not mode and allow_empty:
        return ""
    if mode in CONTRIBUTION_SOURCE_MODES:
        return mode
    raise HttpError(400, "source_mode 仅支持 CR 或 MR")


def _stock_payload(payload: dict) -> dict:
    """存量口径忽略时间、人员和 CR 关键词，只保留仓库分支范围。"""
    return {
        key: value
        for key, value in (payload or {}).items()
        if key in {"organization_ids", "repository_ids", "branch_ids", "branch_type", "repo_type", "domain"}
    }


def _apply_record_filters(queryset, payload: dict):
    """将看板筛选统一应用到贡献明细查询。"""
    organization_ids = _normalize_id_list(payload.get("organization_ids"))
    repository_ids = _normalize_id_list(payload.get("repository_ids"))
    branch_ids = _normalize_id_list(payload.get("branch_ids"))
    pl_group_ids = _normalize_id_list(payload.get("pl_group_ids"))
    source_mode = _normalize_source_mode(payload.get("source_mode"))
    if source_mode:
        queryset = queryset.filter(source_mode=source_mode)
    if organization_ids:
        queryset = queryset.filter(organization_id__in=organization_ids)
    if repository_ids:
        queryset = queryset.filter(repository_id__in=repository_ids)
    if branch_ids:
        queryset = queryset.filter(branch_id__in=branch_ids)
    if payload.get("branch_type"):
        queryset = queryset.filter(branch_type=payload["branch_type"])
    if payload.get("repo_type"):
        queryset = queryset.filter(repo_type=payload["repo_type"])
    if payload.get("domain"):
        queryset = queryset.filter(domain=payload["domain"])
    if pl_group_ids:
        real_ids = [item for item in pl_group_ids if item != UNKNOWN_PL_GROUP_ID]
        condition = Q(author_pl_group_id__in=real_ids) if real_ids else Q()
        if UNKNOWN_PL_GROUP_ID in pl_group_ids:
            condition |= Q(author_pl_group_id__isnull=True) | Q(author_pl_group_name=UNKNOWN_PL_GROUP_NAME)
        queryset = queryset.filter(condition)
    if payload.get("author_username"):
        keyword = _clean_text(payload["author_username"])
        queryset = queryset.filter(Q(author_username__icontains=keyword) | Q(author_user_name__icontains=keyword))
    if payload.get("keyword"):
        keyword = _clean_text(payload["keyword"])
        queryset = queryset.filter(
            Q(title__icontains=keyword)
            | Q(change_key__icontains=keyword)
            | Q(repository_name__icontains=keyword)
            | Q(branch_name__icontains=keyword)
        )
    if payload.get("merged_after"):
        queryset = queryset.filter(merged_at__gte=_to_model_datetime(payload["merged_after"]))
    if payload.get("merged_before"):
        queryset = queryset.filter(merged_at__lte=_to_model_datetime(payload["merged_before"]))
    return queryset


def _base_record_queryset(payload: dict | None = None):
    """返回贡献明细基础查询，统一过滤软删除数据。"""
    queryset = ComplianceContributionRecord.objects.filter(is_deleted=False)
    return _apply_record_filters(queryset, payload or {})


def _sum_lines(queryset) -> dict:
    """聚合常用代码变更指标，并把 None 归零。"""
    data = queryset.aggregate(
        cr_count=Count("id"),
        contributor_count=Count("author_username", distinct=True),
        active_repository_count=Count("repository_id", distinct=True),
        active_branch_count=Count("branch_name", distinct=True),
        added_lines=Sum("added_lines"),
        removed_lines=Sum("removed_lines"),
        net_lines=Sum("net_lines"),
        changed_lines=Sum("changed_lines"),
    )
    return {key: int(value or 0) for key, value in data.items()}


def _load_current_baselines(payload: dict) -> dict[tuple[str, str], ComplianceContributionCodeBaseline]:
    """按筛选范围加载当前生效基线，key 为 repository_id + branch_name。"""
    queryset = ComplianceContributionCodeBaseline.objects.filter(is_deleted=False, is_current=True).select_related(
        "repository", "branch", "operator"
    )
    payload = _stock_payload(payload)
    if _normalize_id_list(payload.get("organization_ids")):
        queryset = queryset.filter(organization_id__in=_normalize_id_list(payload.get("organization_ids")))
    if _normalize_id_list(payload.get("repository_ids")):
        queryset = queryset.filter(repository_id__in=_normalize_id_list(payload.get("repository_ids")))
    if _normalize_id_list(payload.get("branch_ids")):
        queryset = queryset.filter(branch_id__in=_normalize_id_list(payload.get("branch_ids")))
    if payload.get("branch_type"):
        queryset = queryset.filter(branch_type=payload["branch_type"])
    if payload.get("repo_type"):
        queryset = queryset.filter(repo_type=payload["repo_type"])
    if payload.get("domain"):
        queryset = queryset.filter(domain=payload["domain"])
    return {(str(item.repository_id), item.branch_name): item for item in queryset}


def _load_stock_bindings(payload: dict) -> list[ComplianceRepositoryBranch]:
    """加载存量看板的仓库 x 分支范围，作为基线覆盖率分母。"""
    queryset = (
        ComplianceRepositoryBranch.objects.filter(
            is_deleted=False,
            repository__is_deleted=False,
            branch__is_deleted=False,
            branch__is_active=True,
        )
        .select_related("repository", "repository__organization", "branch")
        .prefetch_related("repository__responsibility_groups")
    )
    payload = _stock_payload(payload)
    if _normalize_id_list(payload.get("organization_ids")):
        queryset = queryset.filter(repository__organization_id__in=_normalize_id_list(payload.get("organization_ids")))
    if _normalize_id_list(payload.get("repository_ids")):
        queryset = queryset.filter(repository_id__in=_normalize_id_list(payload.get("repository_ids")))
    if _normalize_id_list(payload.get("branch_ids")):
        queryset = queryset.filter(branch_id__in=_normalize_id_list(payload.get("branch_ids")))
    if payload.get("branch_type"):
        queryset = queryset.filter(branch__branch_type=payload["branch_type"])
    if payload.get("repo_type"):
        queryset = queryset.filter(repository__repo_type=payload["repo_type"])
    if payload.get("domain"):
        queryset = queryset.filter(repository__domain=payload["domain"])
    return list(queryset)


def _stock_increment_for_baseline(baseline: ComplianceContributionCodeBaseline) -> int:
    """计算单条基线时点之后的净增代码量。"""
    return int(
        ComplianceContributionRecord.objects.filter(
            is_deleted=False,
            repository_id=baseline.repository_id,
            branch_name=baseline.branch_name,
            merged_at__gt=baseline.baseline_at,
        ).aggregate(value=Sum("net_lines"))["value"]
        or 0
    )


def _stock_metrics(payload: dict) -> dict:
    """按当前基线和基线后净增计算存量指标。"""
    bindings = _load_stock_bindings(payload)
    baselines = _load_current_baselines(payload)
    binding_keys = {(str(link.repository_id), link.branch.branch_name) for link in bindings}
    stock_lines = 0
    covered_keys = set()
    for key, baseline in baselines.items():
        if key not in binding_keys:
            continue
        covered_keys.add(key)
        stock_lines += int(baseline.baseline_lines or 0) + _stock_increment_for_baseline(baseline)
    return {
        "baseline_repository_count": len({key[0] for key in covered_keys}),
        "baseline_branch_count": len(covered_keys),
        "missing_baseline_count": max(len(binding_keys) - len(covered_keys), 0),
        "stock_lines": stock_lines,
    }


def get_dashboard_summary(**filters) -> dict:
    """查询贡献看板核心指标，当前看板只统计筛选期内 CR 新增贡献。"""
    payload = _filter_payload_from_kwargs(**filters)
    return _sum_lines(_base_record_queryset(payload))


def get_dashboard_trend(**filters) -> list[dict]:
    """按贡献日期返回新增、删除、总变更趋势，新增行数是看板主口径。"""
    payload = _filter_payload_from_kwargs(**filters)
    rows = (
        _base_record_queryset(payload)
        .values("contribution_date")
        .annotate(
            cr_count=Count("id"),
            added_lines=Sum("added_lines"),
            removed_lines=Sum("removed_lines"),
            net_lines=Sum("net_lines"),
            changed_lines=Sum("changed_lines"),
        )
        .order_by("contribution_date")
    )
    return [
        {
            "date": row["contribution_date"],
            "cr_count": int(row.get("cr_count") or 0),
            "added_lines": int(row.get("added_lines") or 0),
            "removed_lines": int(row.get("removed_lines") or 0),
            "net_lines": int(row.get("net_lines") or 0),
            "changed_lines": int(row.get("changed_lines") or 0),
        }
        for row in rows
    ]


def get_pl_group_trend(**filters) -> list[dict]:
    """按贡献日期和作者 PL 组返回新增贡献趋势。"""
    payload = _filter_payload_from_kwargs(**filters)
    rows = (
        _base_record_queryset(payload)
        .values("contribution_date", "author_pl_group_name")
        .annotate(
            cr_count=Count("id"),
            added_lines=Sum("added_lines"),
            removed_lines=Sum("removed_lines"),
            changed_lines=Sum("changed_lines"),
        )
        .order_by("contribution_date", "author_pl_group_name")
    )
    return [
        {
            "date": row["contribution_date"],
            "pl_group_name": row.get("author_pl_group_name") or UNKNOWN_PL_GROUP_NAME,
            "cr_count": int(row.get("cr_count") or 0),
            "added_lines": int(row.get("added_lines") or 0),
            "removed_lines": int(row.get("removed_lines") or 0),
            "changed_lines": int(row.get("changed_lines") or 0),
        }
        for row in rows
    ]


def _repository_ranking_queryset(filters: dict):
    """构造仓库和分支新增贡献排行聚合查询。"""
    payload = _filter_payload_from_kwargs(**filters)
    return (
        _base_record_queryset(payload)
        .values("repository_id", "branch_id", "repository_name", "repository_project_id", "branch_name", "source_mode")
        .annotate(
            cr_count=Count("id"),
            contributor_count=Count("author_username", distinct=True),
            added_lines=Sum("added_lines"),
            removed_lines=Sum("removed_lines"),
            net_lines=Sum("net_lines"),
            changed_lines=Sum("changed_lines"),
        )
        .order_by("-added_lines", "-cr_count", "repository_name", "branch_name")
    )


def _serialize_repository_ranking_row(row: dict) -> dict:
    """序列化单条仓库和分支贡献排行。"""
    return {
        "id": f"{row['repository_id']}:{row['branch_name']}",
        "repository_id": str(row["repository_id"]),
        "branch_id": str(row["branch_id"]) if row.get("branch_id") else None,
        "name": row.get("repository_name") or "",
        "project_id": row.get("repository_project_id") or "",
        "branch_name": row.get("branch_name") or "",
        "repository_name": row.get("repository_name") or "",
        "source_mode": row.get("source_mode") or COMPLIANCE_MODE_CR,
        # 基线字段保留为兼容输出，当前看板不再按基线或存量排序。
        "baseline_id": None,
        "baseline_at": None,
        "baseline_lines": 0,
        "stock_lines": 0,
        "has_baseline": False,
        **{
            key_name: int(row.get(key_name) or 0)
            for key_name in ("cr_count", "contributor_count", "added_lines", "removed_lines", "net_lines", "changed_lines")
        },
    }


def get_repository_ranking(limit: int = 20, **filters) -> list[dict]:
    """按仓库和分支维度返回统计期内新增贡献 Top 排行。"""
    rows = _repository_ranking_queryset(filters)[: max(min(int(limit or 20), 100), 1)]
    return [_serialize_repository_ranking_row(row) for row in rows]


def list_repository_rankings(page: int = 1, page_size: int = 20, **filters) -> dict:
    """分页查询完整仓库和分支贡献排行。"""
    queryset = _repository_ranking_queryset(filters)
    safe_page = max(int(page or 1), 1)
    safe_size = max(min(int(page_size or 20), 100), 1)
    total = queryset.count()
    rows = queryset[(safe_page - 1) * safe_size : safe_page * safe_size]
    return {"items": [_serialize_repository_ranking_row(row) for row in rows], "total": total}


def _person_ranking_queryset(filters: dict):
    """构造创建人新增贡献排行聚合查询。"""
    payload = _filter_payload_from_kwargs(**filters)
    return (
        _base_record_queryset(payload)
        .values("author_user_id", "author_username", "author_user_name", "author_pl_group_id", "author_pl_group_name")
        .annotate(
            repository_count=Count("repository_id", distinct=True),
            branch_count=Count("branch_name", distinct=True),
            cr_count=Count("id"),
            added_lines=Sum("added_lines"),
            removed_lines=Sum("removed_lines"),
            net_lines=Sum("net_lines"),
            changed_lines=Sum("changed_lines"),
        )
        .order_by("-added_lines", "-cr_count", "author_username")
    )


def _serialize_person_ranking_row(row: dict) -> dict:
    """序列化单条创建人贡献排行。"""
    return {
        **row,
        "author_user_id": str(row["author_user_id"]) if row.get("author_user_id") else None,
        "author_pl_group_id": str(row["author_pl_group_id"]) if row.get("author_pl_group_id") else None,
        "author_pl_group_name": row.get("author_pl_group_name") or UNKNOWN_PL_GROUP_NAME,
        "author_display_name": _author_display_name(row.get("author_user_name"), row.get("author_username")),
        **{key: int(row.get(key) or 0) for key in ("repository_count", "branch_count", "cr_count", "added_lines", "removed_lines", "net_lines", "changed_lines")},
    }


def get_person_ranking(limit: int = 20, **filters) -> list[dict]:
    """按创建人返回新增行数贡献 Top 排行。"""
    rows = _person_ranking_queryset(filters)[: max(min(int(limit or 20), 100), 1)]
    return [_serialize_person_ranking_row(row) for row in rows]


def list_person_rankings(page: int = 1, page_size: int = 20, **filters) -> dict:
    """分页查询完整创建人贡献排行。"""
    queryset = _person_ranking_queryset(filters)
    safe_page = max(int(page or 1), 1)
    safe_size = max(min(int(page_size or 20), 100), 1)
    total = queryset.count()
    rows = queryset[(safe_page - 1) * safe_size : safe_page * safe_size]
    return {"items": [_serialize_person_ranking_row(row) for row in rows], "total": total}


def list_pl_group_rankings(page: int = 1, page_size: int = 20, **filters) -> dict:
    """分页查询 PL 组贡献排行，未知归属使用稳定键 unknown。"""
    payload = _filter_payload_from_kwargs(**filters)
    queryset = (
        _base_record_queryset(payload)
        .values("author_pl_group_id", "author_pl_group_name")
        .annotate(
            contributor_count=Count("author_username", distinct=True),
            repository_count=Count("repository_id", distinct=True),
            branch_count=Count("branch_name", distinct=True),
            cr_count=Count("id"),
            added_lines=Sum("added_lines"),
            removed_lines=Sum("removed_lines"),
            net_lines=Sum("net_lines"),
            changed_lines=Sum("changed_lines"),
        )
        .order_by("-added_lines", "-cr_count", "author_pl_group_name")
    )
    safe_page = max(int(page or 1), 1)
    safe_size = max(min(int(page_size or 20), 100), 1)
    total = queryset.count()
    rows = queryset[(safe_page - 1) * safe_size : safe_page * safe_size]
    items = [
        {
            "pl_group_id": str(row["author_pl_group_id"]) if row.get("author_pl_group_id") else UNKNOWN_PL_GROUP_ID,
            "pl_group_name": row.get("author_pl_group_name") or UNKNOWN_PL_GROUP_NAME,
            **{key: int(row.get(key) or 0) for key in ("contributor_count", "repository_count", "branch_count", "cr_count", "added_lines", "removed_lines", "net_lines", "changed_lines")},
        }
        for row in rows
    ]
    return {"items": items, "total": total}


def get_category_distribution(**filters) -> dict:
    """返回仓库类型、领域和 PL 组的新增贡献分布数据。"""
    payload = _filter_payload_from_kwargs(**filters)
    queryset = _base_record_queryset(payload)

    def build(field: str, label_func=None) -> list[dict]:
        rows = (
            queryset.values(field)
            .annotate(
                count=Count("id"),
                cr_count=Count("id"),
                added_lines=Sum("added_lines"),
                removed_lines=Sum("removed_lines"),
                net_lines=Sum("net_lines"),
                changed_lines=Sum("changed_lines"),
            )
            .order_by("-added_lines", field)
        )
        result = []
        for row in rows:
            value = row.get(field) or ""
            result.append(
                {
                    "category": str(value or "unknown"),
                    "category_label": label_func(value) if label_func else str(value or "未设置"),
                    **{key: int(row.get(key) or 0) for key in ("count", "cr_count", "added_lines", "removed_lines", "net_lines", "changed_lines")},
                }
            )
        return result

    repo_type_map = base_services._repo_type_label_map()
    return {
        "repo_types": build("repo_type", lambda value: repo_type_map.get(value, value or "未设置")),
        "domains": build("domain", lambda value: DOMAIN_LABELS.get(value, value or "未设置")),
        "pl_groups": build("author_pl_group_name", lambda value: value or UNKNOWN_PL_GROUP_NAME),
    }


def serialize_record(item: ComplianceContributionRecord) -> dict:
    """把贡献明细转换为 CR/MR 兼容输出。"""
    repo_type_label = base_services._repo_type_label_map().get(item.repo_type, item.repo_type or "")
    return {
        "id": str(item.id),
        "contribution_date": item.contribution_date,
        "organization_id": str(item.organization_id) if item.organization_id else None,
        "organization_group_id": item.organization_group_id,
        "organization_name": item.organization_name,
        "repository_id": str(item.repository_id) if item.repository_id else None,
        "repository_project_id": item.repository_project_id,
        "repository_name": item.repository_name,
        "branch_id": str(item.branch_id) if item.branch_id else None,
        "branch_name": item.branch_name,
        "branch_type": item.branch_type,
        "branch_type_label": BRANCH_TYPE_LABELS.get(item.branch_type, item.branch_type),
        "repo_type": item.repo_type,
        "repo_type_label": repo_type_label,
        "domain": item.domain,
        "domain_label": DOMAIN_LABELS.get(item.domain, item.domain),
        "source_mode": item.source_mode,
        "source_change_id": item.source_change_id,
        "change_request_iid": item.change_request_iid,
        "change_key": item.change_key,
        "title": item.title,
        "web_url": item.web_url,
        "merged_at": item.merged_at,
        "target_branch": item.target_branch,
        "author_username": item.author_username,
        "author_user_id": str(item.author_user_id) if item.author_user_id else None,
        "author_user_name": item.author_user_name,
        "author_display_name": _author_display_name(item.author_user_name, item.author_username),
        "author_pl_group_id": str(item.author_pl_group_id) if item.author_pl_group_id else None,
        "author_pl_group_name": item.author_pl_group_name or UNKNOWN_PL_GROUP_NAME,
        "added_lines": item.added_lines,
        "removed_lines": item.removed_lines,
        "net_lines": item.net_lines,
        "changed_lines": item.changed_lines,
    }


def list_records(page: int = 1, page_size: int = 20, **filters) -> dict:
    """分页查询 CR/MR 贡献明细，下钻表格复用该接口。"""
    safe_page = max(int(page or 1), 1)
    safe_size = max(min(int(page_size or 20), 100), 1)
    queryset = _base_record_queryset(_filter_payload_from_kwargs(**filters)).order_by("-merged_at", "-contribution_date")
    total = queryset.count()
    items = queryset[(safe_page - 1) * safe_size : safe_page * safe_size]
    return {"items": [serialize_record(item) for item in items], "total": total}


def _serialize_collect_task(task: ComplianceContributionCollectTask) -> dict:
    """序列化贡献采集任务。"""
    return {
        "id": str(task.id),
        "trigger_type": task.trigger_type,
        "trigger_type_label": TASK_TRIGGER_LABELS.get(task.trigger_type, task.trigger_type),
        "status": task.status,
        "status_label": TASK_STATUS_LABELS.get(task.status, task.status),
        "merged_after": task.merged_after,
        "merged_before": task.merged_before,
        "filter_payload": task.filter_payload or {},
        "collect_diagnostics": task.collect_diagnostics or {},
        "started_at": task.started_at,
        "finished_at": task.finished_at,
        "scanned_organization_count": task.scanned_organization_count,
        "scanned_repository_count": task.scanned_repository_count,
        "scanned_branch_count": task.scanned_branch_count,
        "fetched_count": task.fetched_count,
        "created_count": task.created_count,
        "updated_count": task.updated_count,
        "skipped_count": task.skipped_count,
        "aggregate_count": task.aggregate_count,
        "error_message": task.error_message,
        "sys_create_datetime": task.sys_create_datetime,
    }


def _serialize_export_task(task: ComplianceContributionExportTask) -> dict:
    """序列化贡献导出任务。"""
    return {
        "id": str(task.id),
        "scope": task.scope,
        "fingerprint": task.fingerprint,
        "status": task.status,
        "progress": task.progress,
        "message": task.message,
        "error_message": task.error_message,
        "file_name": task.file_name,
        "file_size": task.file_size,
        "started_at": task.started_at,
        "finished_at": task.finished_at,
        "sys_create_datetime": task.sys_create_datetime,
    }


def _load_collect_bindings(payload: dict) -> list[ComplianceRepositoryBranch]:
    """加载所有活跃绑定分支，作为贡献采集的仓库 x 分支范围。"""
    queryset = (
        ComplianceRepositoryBranch.objects.filter(
            is_deleted=False,
            repository__is_deleted=False,
            branch__is_deleted=False,
            branch__is_active=True,
        )
        .select_related("repository", "repository__organization", "branch")
        .prefetch_related("repository__responsibility_groups")
    )
    organization_ids = _normalize_id_list(payload.get("organization_ids"))
    repository_ids = _normalize_id_list(payload.get("repository_ids"))
    branch_ids = _normalize_id_list(payload.get("branch_ids"))
    source_mode = _normalize_source_mode(payload.get("source_mode"))
    if organization_ids:
        queryset = queryset.filter(repository__organization_id__in=organization_ids)
    if repository_ids:
        queryset = queryset.filter(repository_id__in=repository_ids)
    if branch_ids:
        queryset = queryset.filter(branch_id__in=branch_ids)
    if source_mode:
        queryset = queryset.filter(repository__mode=source_mode)
    return list(queryset)


def create_collect_task(user, payload, trigger_type: str = CONTRIBUTION_TASK_TRIGGER_MANUAL):
    """创建贡献采集任务；手动任务仅允许超级管理员提交。"""
    if trigger_type == CONTRIBUTION_TASK_TRIGGER_MANUAL and not getattr(user, "is_superuser", False):
        raise HttpError(403, "仅管理员可以手动同步代码贡献数据")
    data = payload.dict() if hasattr(payload, "dict") else dict(payload or {})
    filter_payload = {
        "organization_ids": _normalize_id_list(data.get("organization_ids")),
        "repository_ids": _normalize_id_list(data.get("repository_ids")),
        "branch_ids": _normalize_id_list(data.get("branch_ids")),
        "source_mode": _normalize_source_mode(data.get("source_mode")),
    }
    return ComplianceContributionCollectTask.objects.create(
        trigger_type=trigger_type,
        status=CONTRIBUTION_TASK_STATUS_PENDING,
        merged_after=_to_model_datetime(data["merged_after"]),
        merged_before=_to_model_datetime(data["merged_before"]),
        filter_payload=filter_payload,
    )


def run_collect_task(user, payload) -> dict:
    """提交管理员手动贡献采集任务，并由后台线程异步执行。"""
    active = ComplianceContributionCollectTask.objects.filter(
        status__in=[CONTRIBUTION_TASK_STATUS_PENDING, CONTRIBUTION_TASK_STATUS_RUNNING],
        is_deleted=False,
    ).first()
    if active:
        return {"accepted": False, "message": "已有代码贡献采集任务正在执行", "task": _serialize_collect_task(active)}
    task = create_collect_task(user, payload)
    _start_collect_thread(str(task.id))
    return {"accepted": True, "message": "代码贡献采集任务已提交", "task": _serialize_collect_task(task)}


def _start_collect_thread(task_id: str):
    """启动进程内后台线程执行贡献采集。"""
    thread = threading.Thread(target=execute_collect_task, args=(task_id,), daemon=True)
    thread.start()


def execute_collect_task(task_id: str) -> dict:
    """执行贡献采集任务，负责取数、明细 upsert 和日聚合重算。"""
    close_old_connections()
    task = ComplianceContributionCollectTask.objects.get(id=task_id)
    task.status = CONTRIBUTION_TASK_STATUS_RUNNING
    task.started_at = _to_model_datetime(timezone.now())
    task.error_message = ""
    task.save(update_fields=["status", "started_at", "error_message"])
    try:
        counters, diagnostics, affected_dates = _collect_contribution_records(task)
        aggregate_count = _rebuild_daily_aggregates(affected_dates, task.filter_payload)
        task.status = CONTRIBUTION_TASK_STATUS_SUCCESS
        task.finished_at = _to_model_datetime(timezone.now())
        task.collect_diagnostics = diagnostics
        for key, value in counters.items():
            setattr(task, key, value)
        task.aggregate_count = aggregate_count
        task.save()
    except Exception as exc:
        task.status = CONTRIBUTION_TASK_STATUS_FAILED
        task.finished_at = _to_model_datetime(timezone.now())
        task.error_message = str(exc)
        task.save(update_fields=["status", "finished_at", "error_message"])
    finally:
        close_old_connections()
    return _serialize_collect_task(ComplianceContributionCollectTask.objects.get(id=task_id))


def _collect_contribution_records(task: ComplianceContributionCollectTask):
    """按仓库模式拉取贡献明细；CR 按组织聚合，MR 按项目受控并发。"""
    bindings = _load_collect_bindings(task.filter_payload or {})
    diagnostics: dict[str, Any] = {"cr_groups": [], "cr_branches": [], "mr_projects": []}
    counters = {
        "scanned_organization_count": len({item.repository.organization_id for item in bindings}),
        "scanned_repository_count": len({item.repository_id for item in bindings}),
        "scanned_branch_count": len({item.branch_id for item in bindings}),
        "fetched_count": 0,
        "created_count": 0,
        "updated_count": 0,
        "skipped_count": 0,
    }
    if not bindings:
        diagnostics["reason"] = "未找到活跃代码库-分支绑定关系"
        return counters, diagnostics, set()

    cr_bindings = [item for item in bindings if item.repository.mode == COMPLIANCE_MODE_CR]
    mr_bindings = [item for item in bindings if item.repository.mode == COMPLIANCE_MODE_MR]
    by_group: dict[str, list[ComplianceRepositoryBranch]] = defaultdict(list)
    for link in cr_bindings:
        by_group[link.repository.organization.group_id].append(link)

    affected_dates: set[Any] = set()
    client = CodeComplianceCRClient()
    for group_id, group_links in by_group.items():
        project_ids = sorted({link.repository.project_id for link in group_links})
        branch_names = sorted({link.branch.branch_name for link in group_links})
        repo_by_project = {link.repository.project_id: link.repository for link in group_links}
        branch_by_name = {link.branch.branch_name: link.branch for link in group_links}
        diagnostics["cr_groups"].append({"group_id": group_id, "project_count": len(project_ids), "branch_count": len(branch_names)})
        for branch_name in branch_names:
            total = client.fetch_count(
                group_id=group_id,
                target_branch=branch_name,
                projects=project_ids,
                merged_after=task.merged_after,
                merged_before=task.merged_before,
            )
            rows = []
            if total > 0:
                total_pages = max(math.ceil(total / DEFAULT_PAGE_SIZE), 1)
                for page in range(1, total_pages + 1):
                    rows.extend(
                        client.fetch_page(
                            group_id=group_id,
                            page=page,
                            per_page=DEFAULT_PAGE_SIZE,
                            target_branch=branch_name,
                            projects=project_ids,
                            merged_after=task.merged_after,
                            merged_before=task.merged_before,
                        )
                    )
            branch_created = branch_updated = branch_skipped = 0
            assignments = _load_author_assignments([_clean_text(row.get("author_username")) for row in rows])
            for row in rows:
                repo = repo_by_project.get(_clean_text(row.get("project_id")))
                branch = branch_by_name.get(branch_name)
                change_key = _clean_text(row.get("change_key"))
                if not repo or not branch or not change_key:
                    branch_skipped += 1
                    continue
                created = _upsert_contribution_record(repo, branch, row, assignments, source_mode=COMPLIANCE_MODE_CR)
                contribution_date = _date_from_datetime(row.get("merged_at"))
                affected_dates.add(contribution_date)
                if created:
                    branch_created += 1
                else:
                    branch_updated += 1
            counters["fetched_count"] += len(rows)
            counters["created_count"] += branch_created
            counters["updated_count"] += branch_updated
            counters["skipped_count"] += branch_skipped
            diagnostics["cr_branches"].append(
                {
                    "group_id": group_id,
                    "target_branch": branch_name,
                    "project_count": len(project_ids),
                    "only_count": total,
                    "detail_count": len(rows),
                    "created_count": branch_created,
                    "updated_count": branch_updated,
                    "skipped_count": branch_skipped,
                }
            )

    if mr_bindings:
        # MR 数据湖不支持按组织批量查询；每个项目-分支请求由最多五个 worker 执行。
        mr_client = CodeComplianceMRClient()
        with ThreadPoolExecutor(max_workers=MR_COLLECT_MAX_WORKERS, thread_name_prefix="contribution-mr") as executor:
            futures = {
                executor.submit(
                    mr_client.fetch_all,
                    project_id=link.repository.project_id,
                    target_branch=link.branch.branch_name,
                    merged_after=task.merged_after,
                    merged_before=task.merged_before,
                    per_page=DEFAULT_PAGE_SIZE,
                ): link
                for link in mr_bindings
            }
            for future in as_completed(futures):
                link = futures[future]
                diagnostic = {
                    "project_id": link.repository.project_id,
                    "repository_id": str(link.repository_id),
                    "target_branch": link.branch.branch_name,
                    "only_count": 0,
                    "detail_count": 0,
                    "created_count": 0,
                    "updated_count": 0,
                    "skipped_count": 0,
                    "error": "",
                }
                try:
                    total, rows = future.result()
                    diagnostic["only_count"] = total
                    diagnostic["detail_count"] = len(rows)
                    assignments = _load_author_assignments([_clean_text(row.get("author_username")) for row in rows])
                    for row in rows:
                        if _clean_text(row.get("project_id")) not in {"", link.repository.project_id}:
                            diagnostic["skipped_count"] += 1
                            continue
                        if not _clean_text(row.get("source_change_id")):
                            diagnostic["skipped_count"] += 1
                            continue
                        created = _upsert_contribution_record(
                            link.repository,
                            link.branch,
                            row,
                            assignments,
                            source_mode=COMPLIANCE_MODE_MR,
                        )
                        affected_dates.add(_date_from_datetime(row.get("merged_at")))
                        diagnostic["created_count" if created else "updated_count"] += 1
                except Exception as exc:
                    # 单项目异常不影响其他 MR 项目，但任务诊断必须可用于生产排障。
                    diagnostic["error"] = str(exc)
                finally:
                    for key in ("detail_count", "created_count", "updated_count", "skipped_count"):
                        counters[{"detail_count": "fetched_count", "created_count": "created_count", "updated_count": "updated_count", "skipped_count": "skipped_count"}[key]] += diagnostic[key]
                    diagnostics["mr_projects"].append(diagnostic)
    return counters, diagnostics, affected_dates


def _upsert_contribution_record(
    repo: ComplianceRepository,
    branch: ComplianceManagedBranch,
    row: dict[str, Any],
    assignments: dict[str, AuthorAssignment],
    source_mode: str = COMPLIANCE_MODE_CR,
) -> bool:
    """按来源模式和上游稳定标识幂等写入 CR/MR 贡献明细。"""
    source_mode = _normalize_source_mode(source_mode, allow_empty=False)
    source_change_id = _clean_text(row.get("source_change_id") or row.get("change_key"))
    if not source_change_id:
        raise HttpError(400, "贡献明细缺少上游变更唯一标识")
    author_username = _clean_text(row.get("author_username"))
    assignment = assignments.get(author_username, _default_assignment())
    added, removed, net, changed = _line_metrics(row)
    merged_at = row.get("merged_at")
    contribution_date = _date_from_datetime(merged_at)
    responsibility_names = [item.name for item in repo.responsibility_groups.all()]
    defaults = {
        "contribution_date": contribution_date,
        "organization": repo.organization,
        "organization_group_id": repo.organization.group_id,
        "organization_name": repo.organization.name,
        "repository_project_id": repo.project_id,
        "repository_name": repo.project_name,
        "branch": branch,
        "branch_type": branch.branch_type,
        "repo_type": repo.repo_type,
        "domain": repo.domain,
        "responsibility_group_names": responsibility_names,
        "source_mode": source_mode,
        "source_change_id": source_change_id,
        "change_request_iid": _clean_text(row.get("change_request_iid")),
        # MR 数据湖没有 change_key；保留为空以避免把 MR 误用于 CR 漏合语义。
        "change_key": _clean_text(row.get("change_key")) if source_mode == COMPLIANCE_MODE_CR else "",
        "title": _clean_text(row.get("title")),
        "description": _clean_text(row.get("description")),
        "web_url": _clean_text(row.get("web_url")),
        "merged_at": _to_model_datetime(merged_at),
        "target_branch": _clean_text(row.get("target_branch")) or branch.branch_name,
        "author_username": author_username,
        "author_user_id": assignment.user_id,
        "author_user_name": assignment.user_name,
        "author_pl_group_id": assignment.pl_group_id,
        "author_pl_group_name": assignment.pl_group_name,
        "added_lines": added,
        "removed_lines": removed,
        "net_lines": net,
        "changed_lines": changed,
    }
    _, created = ComplianceContributionRecord.objects.update_or_create(
        repository=repo,
        branch_name=branch.branch_name,
        source_mode=source_mode,
        source_change_id=source_change_id,
        defaults=defaults,
    )
    return created


def _rebuild_daily_aggregates(dates: set[Any], payload: dict) -> int:
    """按受影响日期删除并重算日聚合，避免重复采集造成累加误差。"""
    if not dates:
        return 0
    record_queryset = _apply_record_filters(
        ComplianceContributionRecord.objects.filter(is_deleted=False, contribution_date__in=dates),
        payload or {},
    )
    aggregate_queryset = _apply_aggregate_scope(
        ComplianceContributionDailyAggregate.objects.filter(contribution_date__in=dates),
        payload or {},
    )
    aggregate_queryset.delete()
    rows = (
        record_queryset.values(
            "contribution_date",
            "source_mode",
            "organization_id",
            "organization_group_id",
            "organization_name",
            "repository_id",
            "repository_project_id",
            "repository_name",
            "branch_id",
            "branch_name",
            "branch_type",
            "repo_type",
            "domain",
            "author_user_id",
            "author_username",
            "author_user_name",
            "author_pl_group_id",
            "author_pl_group_name",
        )
        .annotate(
            cr_count=Count("id"),
            added_lines=Sum("added_lines"),
            removed_lines=Sum("removed_lines"),
            net_lines=Sum("net_lines"),
            changed_lines=Sum("changed_lines"),
        )
    )
    objects = []
    for row in rows:
        metrics = {
            key: row.pop(key)
            for key in ("cr_count", "added_lines", "removed_lines", "net_lines", "changed_lines")
        }
        objects.append(
            ComplianceContributionDailyAggregate(
                **row,
                contributor_count=1 if row.get("author_username") else 0,
                cr_count=int(metrics.get("cr_count") or 0),
                added_lines=int(metrics.get("added_lines") or 0),
                removed_lines=int(metrics.get("removed_lines") or 0),
                net_lines=int(metrics.get("net_lines") or 0),
                changed_lines=int(metrics.get("changed_lines") or 0),
            )
        )
    ComplianceContributionDailyAggregate.objects.bulk_create(objects, batch_size=500)
    return len(objects)


def _apply_aggregate_scope(queryset, payload: dict):
    """把采集范围应用到日聚合表删除查询。"""
    source_mode = _normalize_source_mode(payload.get("source_mode"))
    if source_mode:
        queryset = queryset.filter(source_mode=source_mode)
    if _normalize_id_list(payload.get("organization_ids")):
        queryset = queryset.filter(organization_id__in=_normalize_id_list(payload.get("organization_ids")))
    if _normalize_id_list(payload.get("repository_ids")):
        queryset = queryset.filter(repository_id__in=_normalize_id_list(payload.get("repository_ids")))
    if _normalize_id_list(payload.get("branch_ids")):
        queryset = queryset.filter(branch_id__in=_normalize_id_list(payload.get("branch_ids")))
    return queryset


def list_collect_tasks(page: int = 1, page_size: int = 20, status: str | None = None, trigger_type: str | None = None):
    """分页查询贡献采集任务历史。"""
    queryset = ComplianceContributionCollectTask.objects.filter(is_deleted=False)
    if status:
        queryset = queryset.filter(status=status)
    if trigger_type:
        queryset = queryset.filter(trigger_type=trigger_type)
    total = queryset.count()
    safe_page = max(int(page or 1), 1)
    safe_size = max(min(int(page_size or 20), 100), 1)
    items = queryset.order_by("-sys_create_datetime")[(safe_page - 1) * safe_size : safe_page * safe_size]
    return {"items": [_serialize_collect_task(item) for item in items], "total": total}


def get_collect_task(task_id: str) -> dict:
    """读取单个贡献采集任务详情。"""
    return _serialize_collect_task(get_object_or_404(ComplianceContributionCollectTask, id=task_id, is_deleted=False))


def _serialize_baseline(item: ComplianceContributionCodeBaseline) -> dict:
    """序列化代码量基线记录。"""
    return {
        "id": str(item.id),
        "organization_id": str(item.organization_id) if item.organization_id else None,
        "organization_group_id": item.organization_group_id,
        "organization_name": item.organization_name,
        "repository_id": str(item.repository_id),
        "repository_project_id": item.repository_project_id,
        "repository_name": item.repository_name,
        "branch_id": str(item.branch_id) if item.branch_id else None,
        "branch_name": item.branch_name,
        "branch_type": item.branch_type,
        "baseline_lines": int(item.baseline_lines or 0),
        "baseline_at": item.baseline_at,
        "source": item.source,
        "source_label": BASELINE_SOURCE_LABELS.get(item.source, item.source),
        "remark": item.remark,
        "is_current": item.is_current,
        "operator_name": getattr(item.operator, "name", "") or getattr(item.operator, "username", "") or "",
        "sys_create_datetime": item.sys_create_datetime,
    }


def list_code_baselines(page: int = 1, page_size: int = 20, current_only: bool = True, **filters) -> dict:
    """分页查询代码量基线，默认只看当前生效记录。"""
    payload = _stock_payload(_filter_payload_from_kwargs(**filters))
    queryset = ComplianceContributionCodeBaseline.objects.filter(is_deleted=False).select_related("operator")
    if current_only:
        queryset = queryset.filter(is_current=True)
    if _normalize_id_list(payload.get("organization_ids")):
        queryset = queryset.filter(organization_id__in=_normalize_id_list(payload.get("organization_ids")))
    if _normalize_id_list(payload.get("repository_ids")):
        queryset = queryset.filter(repository_id__in=_normalize_id_list(payload.get("repository_ids")))
    if _normalize_id_list(payload.get("branch_ids")):
        queryset = queryset.filter(branch_id__in=_normalize_id_list(payload.get("branch_ids")))
    if payload.get("branch_type"):
        queryset = queryset.filter(branch_type=payload["branch_type"])
    if payload.get("repo_type"):
        queryset = queryset.filter(repo_type=payload["repo_type"])
    if payload.get("domain"):
        queryset = queryset.filter(domain=payload["domain"])
    total = queryset.count()
    safe_page = max(int(page or 1), 1)
    safe_size = max(min(int(page_size or 20), 100), 1)
    items = queryset.order_by("-is_current", "-baseline_at")[(safe_page - 1) * safe_size : safe_page * safe_size]
    return {"items": [_serialize_baseline(item) for item in items], "total": total}


def save_code_baseline(user, payload, source: str = CONTRIBUTION_BASELINE_SOURCE_MANUAL) -> dict:
    """新增一次代码量基线校准，并把同仓库分支旧基线置为历史。"""
    data = payload.dict() if hasattr(payload, "dict") else dict(payload or {})
    repo = get_object_or_404(ComplianceRepository.objects.select_related("organization"), id=data.get("repository_id"), is_deleted=False)
    branch = get_object_or_404(ComplianceManagedBranch, id=data.get("branch_id"), is_deleted=False)
    baseline_lines = int(data.get("baseline_lines") or 0)
    if baseline_lines < 0:
        raise HttpError(400, "基线代码量不能小于 0")
    baseline_at = _to_model_datetime(data.get("baseline_at"))
    if not baseline_at:
        raise HttpError(400, "请选择基线统计时间")
    ComplianceContributionCodeBaseline.objects.filter(
        is_deleted=False,
        is_current=True,
        repository=repo,
        branch_name=branch.branch_name,
    ).update(is_current=False)
    item = ComplianceContributionCodeBaseline.objects.create(
        organization=repo.organization,
        repository=repo,
        branch=branch,
        organization_group_id=repo.organization.group_id,
        organization_name=repo.organization.name,
        repository_project_id=repo.project_id,
        repository_name=repo.project_name,
        branch_name=branch.branch_name,
        branch_type=branch.branch_type,
        repo_type=repo.repo_type,
        domain=repo.domain,
        baseline_lines=baseline_lines,
        baseline_at=baseline_at,
        source=source,
        remark=_clean_text(data.get("remark")),
        is_current=True,
        operator=user,
    )
    return _serialize_baseline(item)


def build_baseline_template_response() -> HttpResponse:
    """生成代码量基线导入模板。"""
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "代码量基线"
    sheet.append(["代码库ID", "分支名称", "基线代码量", "基线统计时间", "备注"])
    sheet.append(["project_id_1", "master", 120000, "2026-06-16 00:00:00", "初始化基线"])
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="code_contribution_baseline_template.xlsx"'
    workbook.save(response)
    return response


def import_code_baselines(user, file_obj) -> dict:
    """按 Excel 批量导入代码量基线。"""
    workbook = openpyxl.load_workbook(file_obj, data_only=True)
    sheet = workbook.active
    created_count = 0
    errors = []
    for row_no, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        project_id, branch_name, baseline_lines, baseline_at, remark = (list(row) + [None] * 5)[:5]
        if not any([project_id, branch_name, baseline_lines, baseline_at]):
            continue
        try:
            repo = ComplianceRepository.objects.get(project_id=_clean_text(project_id), is_deleted=False)
            branch = ComplianceManagedBranch.objects.get(branch_name=_clean_text(branch_name), domain=repo.domain, is_deleted=False)
            payload = {
                "repository_id": str(repo.id),
                "branch_id": str(branch.id),
                "baseline_lines": int(baseline_lines or 0),
                "baseline_at": baseline_at,
                "remark": _clean_text(remark),
            }
            save_code_baseline(user, payload, CONTRIBUTION_BASELINE_SOURCE_IMPORT)
            created_count += 1
        except Exception as exc:
            errors.append({"row_no": row_no, "message": str(exc)})
    return {"created_count": created_count, "updated_count": 0, "ignored_count": 0, "errors": errors}


@scheduler_task
def run_scheduled_contribution_collect():
    """每日凌晨采集前一天所有活跃绑定分支的贡献数据。"""
    end = _to_model_datetime(timezone.now()).replace(hour=0, minute=0, second=0, microsecond=0)
    start = end - timedelta(days=1)
    task = ComplianceContributionCollectTask.objects.create(
        trigger_type=CONTRIBUTION_TASK_TRIGGER_SCHEDULED,
        status=CONTRIBUTION_TASK_STATUS_PENDING,
        merged_after=start,
        merged_before=end,
        filter_payload={},
    )
    return execute_collect_task(str(task.id))


def create_backfill_tasks(user, months: int = 12) -> list[dict]:
    """按月创建 CR 历史回补任务；MR 按产品约束仅从上线后增量采集。"""
    if not getattr(user, "is_superuser", False):
        raise HttpError(403, "仅管理员可以创建历史回补任务")
    now = _to_model_datetime(timezone.now())
    tasks = []
    for index in range(max(min(int(months or 12), 24), 1)):
        end = (now.replace(day=1, hour=0, minute=0, second=0, microsecond=0) - timedelta(days=31 * index))
        start = (end - timedelta(days=31)).replace(day=1)
        task = ComplianceContributionCollectTask.objects.create(
            trigger_type=CONTRIBUTION_TASK_TRIGGER_BACKFILL,
            status=CONTRIBUTION_TASK_STATUS_PENDING,
            merged_after=start,
            merged_before=end,
            filter_payload={"source_mode": COMPLIANCE_MODE_CR},
        )
        tasks.append(_serialize_collect_task(task))
    return tasks


def _export_fingerprint(user, payload: dict) -> str:
    """根据用户、导出范围和筛选条件生成复用指纹。"""
    raw = json.dumps({"user": str(getattr(user, "id", "")), **payload}, sort_keys=True, ensure_ascii=False)
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()


def prepare_export_task(user, payload) -> dict:
    """创建或复用贡献看板异步导出任务。"""
    data = payload.dict() if hasattr(payload, "dict") else dict(payload or {})
    scope = data.get("scope") or CONTRIBUTION_EXPORT_SCOPE_SUMMARY
    if scope not in {CONTRIBUTION_EXPORT_SCOPE_SUMMARY, CONTRIBUTION_EXPORT_SCOPE_RECORDS}:
        raise HttpError(400, "导出范围仅支持 summary/records")
    normalized = {"scope": scope, "filters": data.get("filters") or {}}
    fingerprint = _export_fingerprint(user, normalized)
    active = ComplianceContributionExportTask.objects.filter(
        user=user,
        fingerprint=fingerprint,
        status__in=EXPORT_ACTIVE_STATUSES,
        is_deleted=False,
    ).first()
    if active:
        return {"mode": "async", "task": _serialize_export_task(active)}
    task = ComplianceContributionExportTask.objects.create(
        user=user,
        scope=scope,
        fingerprint=fingerprint,
        payload=normalized,
        status=CONTRIBUTION_TASK_STATUS_PENDING,
    )
    threading.Thread(target=_execute_export_task, args=(str(task.id),), daemon=True).start()
    return {"mode": "async", "task": _serialize_export_task(task)}


def _execute_export_task(task_id: str):
    """后台生成贡献看板 Excel 文件。"""
    close_old_connections()
    task = ComplianceContributionExportTask.objects.get(id=task_id)
    task.status = CONTRIBUTION_TASK_STATUS_RUNNING
    task.started_at = _to_model_datetime(timezone.now())
    task.progress = 10
    task.save(update_fields=["status", "started_at", "progress"])
    try:
        workbook = _build_export_workbook(task)
        export_dir = Path(tempfile.gettempdir()) / "code_compliance_contribution_exports"
        export_dir.mkdir(parents=True, exist_ok=True)
        file_name = f"code_contribution_{task.scope}_{task.id}.xlsx"
        file_path = export_dir / file_name
        workbook.save(file_path)
        task.status = CONTRIBUTION_TASK_STATUS_SUCCESS
        task.progress = 100
        task.message = "导出完成"
        task.file_name = file_name
        task.file_path = str(file_path)
        task.file_size = file_path.stat().st_size
        task.finished_at = _to_model_datetime(timezone.now())
        task.save()
    except Exception as exc:
        task.status = CONTRIBUTION_TASK_STATUS_FAILED
        task.error_message = str(exc)
        task.finished_at = _to_model_datetime(timezone.now())
        task.save(update_fields=["status", "error_message", "finished_at"])
    finally:
        close_old_connections()


def _build_export_workbook(task: ComplianceContributionExportTask):
    """根据导出范围生成 CR/MR 聚合排行或变更明细 Excel。"""
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    filters = (task.payload or {}).get("filters") or {}
    if task.scope == CONTRIBUTION_EXPORT_SCOPE_RECORDS:
        sheet.title = "变更明细"
        headers = ["来源", "上游变更ID", "日期", "代码库", "项目ID", "分支", "创建人", "PL组", "标题", "链接", "新增", "删除", "净增", "总变更", "合入时间"]
        sheet.append(headers)
        for item in _base_record_queryset(filters).order_by("-merged_at")[:100000]:
            sheet.append([
                item.source_mode,
                item.source_change_id,
                item.contribution_date,
                item.repository_name,
                item.repository_project_id,
                item.branch_name,
                _author_display_name(item.author_user_name, item.author_username),
                item.author_pl_group_name,
                item.title,
                item.web_url,
                item.added_lines,
                item.removed_lines,
                item.net_lines,
                item.changed_lines,
                item.merged_at,
            ])
    else:
        sheet.title = "代码贡献看板"
        headers = ["来源", "代码库", "项目ID", "分支", "变更数", "贡献人数", "新增行数", "删除行数", "总变更行数"]
        sheet.append(headers)
        for item in get_repository_ranking(limit=1000, **filters):
            sheet.append([
                item["source_mode"],
                item["repository_name"],
                item["project_id"],
                item["branch_name"],
                item["cr_count"],
                item["contributor_count"],
                item["added_lines"],
                item["removed_lines"],
                item["changed_lines"],
            ])
    return workbook


def get_export_task(user, task_id: str) -> dict:
    """查询当前用户的贡献导出任务。"""
    return _serialize_export_task(get_object_or_404(ComplianceContributionExportTask, id=task_id, user=user, is_deleted=False))


def download_export_task_file(user, task_id: str):
    """下载当前用户已完成的贡献导出文件。"""
    task = get_object_or_404(ComplianceContributionExportTask, id=task_id, user=user, is_deleted=False)
    if task.status != CONTRIBUTION_TASK_STATUS_SUCCESS:
        raise HttpError(409, "导出任务尚未完成")
    path = Path(task.file_path)
    if not path.exists():
        raise HttpError(404, "导出文件不存在或已过期")
    return FileResponse(open(path, "rb"), as_attachment=True, filename=task.file_name)
