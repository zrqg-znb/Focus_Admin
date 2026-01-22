from .models import PerformanceIndicator, PerformanceIndicatorData, PerformanceIndicatorImportTask, PerformanceRiskRecord
from django.db import transaction
from typing import List, Dict, Any
from datetime import date
from django.utils import timezone
from django.conf import settings

import csv
import os
import zipfile
from io import BytesIO, StringIO

import openpyxl

def upload_performance_data(payload: Dict[str, Any]) -> Dict[str, Any]:
    category = payload.get('category')
    project = payload.get('project')
    module = payload.get('module')
    chip_type = payload.get('chip_type')
    test_date = payload.get('date')
    data_items = payload.get('data', [])
    
    success_count = 0
    errors = []
    
    with transaction.atomic():
        for item in data_items:
            code = item.get('code')
            name = item.get('name')
            value = item.get('value')
            
            indicator = None
            
            # 1. Try match by code
            if code:
                indicator = PerformanceIndicator.objects.filter(code=code).first()
            
            # 2. Try match by composite keys
            if not indicator and name:
                qs = PerformanceIndicator.objects.filter(
                    project=project,
                    module=module,
                    chip_type=chip_type,
                    name=name
                )
                if category:
                    qs = qs.filter(category=category)
                indicator = qs.first()
                if not indicator and not category:
                    candidates = list(qs[:2])
                    if len(candidates) > 1:
                        errors.append(f"Indicator ambiguous without category: {name}")
                        continue
            
            if not indicator:
                identifier = code if code else name
                errors.append(f"Indicator not found: {identifier}")
                continue
            
            # Calculate fluctuation
            fluctuation_value = value - indicator.baseline_value
            
            data_obj, _ = PerformanceIndicatorData.objects.update_or_create(
                indicator=indicator,
                date=test_date,
                defaults={
                    'value': value,
                    'fluctuation_value': fluctuation_value
                }
            )
            _create_risk_if_violation(indicator, data_obj)
            success_count += 1
            
    return {"success_count": success_count, "errors": errors}

def _create_risk_if_violation(indicator: PerformanceIndicator, data_obj: PerformanceIndicatorData):
    dir = indicator.fluctuation_direction or 'none'
    rng = indicator.fluctuation_range or 0.0
    dev = data_obj.fluctuation_value or 0.0
    violated = False
    if dir == 'up':
        violated = dev < -rng
    elif dir == 'down':
        violated = dev > rng
    else:
        violated = abs(dev) > rng
    if not violated:
        return
    PerformanceRiskRecord.objects.update_or_create(
        indicator=indicator,
        data=data_obj,
        defaults={
            'occur_date': data_obj.date,
            'status': 'open',
            'owner': indicator.owner,
            'baseline_value': indicator.baseline_value,
            'measured_value': data_obj.value,
            'deviation_value': dev,
            'allowed_range': rng,
            'direction': dir,
            'message': ''
        }
    )

def import_indicators_service(data_list: List[Dict]) -> Dict[str, Any]:
    from core.user.user_model import User
    
    success_count = 0
    errors = []
    with transaction.atomic():
        for row in data_list:
            try:
                # Extract keys
                code = row.get('code')
                category = row.get('category') or 'vehicle'
                name = row.get('name')
                project = row.get('project')
                module = row.get('module')
                chip_type = row.get('chip_type')
                
                if category not in ('vehicle', 'cockpit'):
                    category = 'vehicle'

                if not name or not project or not module or not chip_type:
                    errors.append(f"Missing required fields for row: {row}")
                    continue

                defaults = row.copy()
                if 'code' in defaults: del defaults['code']
                if 'category' in defaults: del defaults['category']
                if 'name' in defaults: del defaults['name']
                if 'project' in defaults: del defaults['project']
                if 'module' in defaults: del defaults['module']
                if 'chip_type' in defaults: del defaults['chip_type']
                
                # Handle Owner field (username to user_id)
                owner_username = defaults.pop('owner', None)
                if owner_username:
                    owner_user = User.objects.filter(username=owner_username).first()
                    if owner_user:
                        defaults['owner'] = owner_user
                
                # Ensure value_type is valid
                if 'value_type' not in defaults:
                    defaults['value_type'] = 'avg'
                
                # Ensure fluctuation_direction is valid
                if 'fluctuation_direction' not in defaults:
                    defaults['fluctuation_direction'] = 'none'

                if code:
                    PerformanceIndicator.objects.update_or_create(
                        code=code,
                        defaults={
                            'category': category,
                            'name': name,
                            'project': project,
                            'module': module,
                            'chip_type': chip_type,
                            **defaults
                        }
                    )
                else:
                    PerformanceIndicator.objects.update_or_create(
                        category=category,
                        project=project,
                        module=module,
                        chip_type=chip_type,
                        name=name,
                        defaults=defaults
                    )
                success_count += 1
            except Exception as e:
                errors.append(f"Error processing row {row}: {str(e)}")
    return {"success_count": success_count, "errors": errors}


def _header_map():
    return {
        "业务唯一标识": "code",
        "分类": "category",
        "指标名称": "name",
        "所属模块": "module",
        "所属项目": "project",
        "芯片类型": "chip_type",
        "值类型": "value_type",
        "基线值": "baseline_value",
        "单位": "baseline_unit",
        "允许浮动范围": "fluctuation_range",
        "浮动方向": "fluctuation_direction",
        "责任人": "owner",
        "Code": "code",
        "Category": "category",
        "Name": "name",
        "Module": "module",
        "Project": "project",
        "Chip Type": "chip_type",
        "Value Type": "value_type",
        "Baseline Value": "baseline_value",
        "Baseline Unit": "baseline_unit",
        "Fluctuation Range": "fluctuation_range",
        "Fluctuation Direction": "fluctuation_direction",
        "Owner": "owner",
    }


def _coerce_float(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if not s:
        return None
    return float(s)


def _normalize_category(v):
    if not v:
        return "vehicle"
    s = str(v).strip().lower()
    if s in ("vehicle", "car", "car_control", "carcontrol", "车控", "车控类", "车控端"):
        return "vehicle"
    if s in ("cockpit", "cabin", "座舱", "座舱类", "座舱端"):
        return "cockpit"
    if s in ("vehicle", "cockpit"):
        return s
    return "vehicle"


def _resolve_owner(owner_value):
    if not owner_value:
        return None
    from core.user.user_model import User

    s = str(owner_value).strip()
    if len(s) == 36:
        u = User.objects.filter(id=s).first()
        if u:
            return u
    return User.objects.filter(username=s).first()


def _iter_rows_from_file(file_path: str):
    filename = os.path.basename(file_path).lower()
    mapping = _header_map()

    if filename.endswith(".xlsx"):
        wb = openpyxl.load_workbook(filename=file_path, read_only=True, data_only=True)
        ws = wb.active
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        total = max((ws.max_row or 1) - 1, 0)
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_dict = {}
            for i, value in enumerate(row):
                if i < len(headers):
                    header_name = headers[i]
                    if header_name and header_name in mapping:
                        row_dict[mapping[header_name]] = value
            if row_dict:
                yield row_dict, total
        return

    if filename.endswith(".csv"):
        with open(file_path, "rb") as f:
            content = f.read()
        content_str = content.decode("utf-8-sig")
        sio = StringIO(content_str)
        reader = csv.reader(sio)
        rows_iter = iter(reader)
        headers = next(rows_iter, [])
        data_rows = list(rows_iter)
        total = len(data_rows)
        for row in data_rows:
            row_dict = {}
            for i, value in enumerate(row):
                if i < len(headers):
                    header_name = headers[i]
                    if header_name and header_name in mapping:
                        row_dict[mapping[header_name]] = value
            if row_dict:
                yield row_dict, total
        return

    raise ValueError("不支持的文件格式，请上传 .xlsx 或 .csv 文件")


def run_indicator_import_task(task_id: str):
    from django.db import connection

    task = PerformanceIndicatorImportTask.objects.filter(id=task_id).first()
    if not task:
        return

    PerformanceIndicatorImportTask.objects.filter(id=task_id).update(
        status="running",
        progress=0,
        processed_rows=0,
        success_count=0,
        error_count=0,
        message="任务已开始",
        started_at=timezone.now(),
    )

    errors: list[str] = []
    processed = 0
    success_count = 0
    error_count = 0
    total_rows = None

    try:
        for row, total in _iter_rows_from_file(task.file_path):
            if total_rows is None:
                total_rows = total
                PerformanceIndicatorImportTask.objects.filter(id=task_id).update(
                    total_rows=total_rows,
                    message="已读取文件，开始导入",
                )

            processed += 1
            try:
                code = row.get("code") or None
                category = _normalize_category(row.get("category"))
                name = (row.get("name") or "").strip()
                project = (row.get("project") or "").strip()
                module = (row.get("module") or "").strip()
                chip_type = (row.get("chip_type") or "").strip()

                if not name or not project or not module or not chip_type:
                    raise ValueError(f"缺少必填字段: {row}")

                value_type = (row.get("value_type") or "avg").strip()
                if value_type not in ("avg", "min", "max"):
                    value_type = "avg"

                fluctuation_direction = (row.get("fluctuation_direction") or "none").strip()
                if fluctuation_direction not in ("up", "down", "none"):
                    fluctuation_direction = "none"

                baseline_value = _coerce_float(row.get("baseline_value"))
                fluctuation_range = _coerce_float(row.get("fluctuation_range"))
                if baseline_value is None:
                    raise ValueError(f"基线值为空或无效: {row}")
                if fluctuation_range is None:
                    fluctuation_range = 0.0

                defaults = {
                    "category": category,
                    "name": name,
                    "project": project,
                    "module": module,
                    "chip_type": chip_type,
                    "value_type": value_type,
                    "baseline_value": baseline_value,
                    "baseline_unit": (row.get("baseline_unit") or "").strip(),
                    "fluctuation_range": fluctuation_range,
                    "fluctuation_direction": fluctuation_direction,
                }

                owner_user = _resolve_owner(row.get("owner"))
                if owner_user:
                    defaults["owner"] = owner_user

                if code:
                    PerformanceIndicator.objects.update_or_create(
                        code=str(code).strip(),
                        defaults=defaults,
                    )
                else:
                    PerformanceIndicator.objects.update_or_create(
                        category=category,
                        project=project,
                        module=module,
                        chip_type=chip_type,
                        name=name,
                        defaults=defaults,
                    )
                success_count += 1
            except Exception as e:
                error_count += 1
                if len(errors) < 50:
                    errors.append(str(e))

            if total_rows and total_rows > 0:
                progress = int((processed / total_rows) * 100)
                progress = max(0, min(progress, 99))
            else:
                progress = 0

            if processed % 10 == 0 or processed == total_rows:
                PerformanceIndicatorImportTask.objects.filter(id=task_id).update(
                    processed_rows=processed,
                    success_count=success_count,
                    error_count=error_count,
                    progress=progress,
                    message="正在导入",
                    errors="\n".join(errors),
                )

        PerformanceIndicatorImportTask.objects.filter(id=task_id).update(
            status="success",
            progress=100,
            processed_rows=processed,
            success_count=success_count,
            error_count=error_count,
            message="导入完成",
            errors="\n".join(errors),
            finished_at=timezone.now(),
        )
    except zipfile.BadZipFile:
        PerformanceIndicatorImportTask.objects.filter(id=task_id).update(
            status="failed",
            progress=100,
            message="文件损坏或格式不正确，请确保是有效的 .xlsx 文件",
            errors="\n".join(errors),
            finished_at=timezone.now(),
        )
    except UnicodeDecodeError:
        PerformanceIndicatorImportTask.objects.filter(id=task_id).update(
            status="failed",
            progress=100,
            message="CSV文件编码错误，请使用UTF-8编码",
            errors="\n".join(errors),
            finished_at=timezone.now(),
        )
    except Exception as e:
        PerformanceIndicatorImportTask.objects.filter(id=task_id).update(
            status="failed",
            progress=100,
            message=f"导入失败: {str(e)}",
            errors="\n".join(errors),
            finished_at=timezone.now(),
        )
    finally:
        connection.close()
