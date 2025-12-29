from ninja import Router, UploadedFile, File
from ninja.pagination import paginate
from ninja.errors import HttpError
from typing import List
from .schemas import (
    PerformanceIndicatorSchema, 
    PerformanceIndicatorCreateSchema, 
    PerformanceIndicatorUpdateSchema,
    PerformanceDataUploadSchema,
    PerformanceDataUploadResponse,
    PerformanceDataTrendSchema
)
from .models import PerformanceIndicator, PerformanceIndicatorData
from .services import upload_performance_data, import_indicators_service
from django.shortcuts import get_object_or_404
from common.fu_pagination import MyPagination
from common.fu_crud import create, delete, update, retrieve
from django.db.models import Q
import openpyxl
import zipfile
import csv
from io import BytesIO, StringIO

router = Router()

# --- Indicator CRUD ---

@router.post("/indicators/import", url_name="import_indicators")
def import_indicators(request, file: UploadedFile = File(...)):
    # Check extension
    filename = file.name.lower()
    if not (filename.endswith('.xlsx') or filename.endswith('.csv')):
        raise HttpError(400, "不支持的文件格式，请上传 .xlsx 或 .csv 文件")

    try:
        # Ensure pointer is at the beginning
        if hasattr(file, 'seek'):
            file.seek(0)
            
        content = file.read()
        if not content:
            raise HttpError(400, "上传的文件为空")

        data_rows = []
        headers = []

        if filename.endswith('.xlsx'):
            # Read Excel file
            wb = openpyxl.load_workbook(filename=BytesIO(content))
            ws = wb.active
            
            # Get headers from first row
            headers = [cell.value for cell in ws[1]]
            
            # Get data rows
            for row in ws.iter_rows(min_row=2, values_only=True):
                data_rows.append(row)
                
        elif filename.endswith('.csv'):
            # Read CSV file
            # Decode bytes to string, handling BOM if present
            content_str = content.decode('utf-8-sig')
            f = StringIO(content_str)
            reader = csv.reader(f)
            rows = list(reader)
            
            if not rows:
                raise HttpError(400, "CSV文件为空")
                
            headers = rows[0]
            data_rows = rows[1:]

        # Map headers to fields
        # Expected headers: "业务唯一标识", "指标名称", "所属模块", "所属项目", "芯片类型", etc.
        header_map = {
            "业务唯一标识": "code",
            "指标名称": "name",
            "所属模块": "module",
            "所属项目": "project",
            "芯片类型": "chip_type",
            "值类型": "value_type",
            "基线值": "baseline_value",
            "单位": "baseline_unit",
            "允许浮动范围": "fluctuation_range",
            "浮动方向": "fluctuation_direction",
            "责任人": "owner",
            
            # English headers support
            "Code": "code",
            "Name": "name",
            "Module": "module",
            "Project": "project",
            "Chip Type": "chip_type",
            "Value Type": "value_type",
            "Baseline Value": "baseline_value",
            "Baseline Unit": "baseline_unit",
            "Fluctuation Range": "fluctuation_range",
            "Fluctuation Direction": "fluctuation_direction",
            "Owner": "owner"
        }
        
        data_list = []
        for row in data_rows:
            row_dict = {}
            for i, value in enumerate(row):
                if i < len(headers):
                    header_name = headers[i]
                    # Handle potential None in headers (e.g. empty column in Excel)
                    if header_name and header_name in header_map:
                        field_name = header_map[header_name]
                        # CSV values are always strings, might need type conversion if service expects otherwise
                        # But for now we pass as is, assuming service handles it or fields are strings
                        row_dict[field_name] = value
            if row_dict:
                data_list.append(row_dict)
                
        result = import_indicators_service(data_list)
        return result
        
    except zipfile.BadZipFile:
        raise HttpError(400, "文件损坏或格式不正确，请确保是有效的 .xlsx 文件")
    except UnicodeDecodeError:
        raise HttpError(400, "CSV文件编码错误，请使用UTF-8编码")
    except Exception as e:
        raise HttpError(400, f"导入失败: {str(e)}")

@router.post("/indicators", response=PerformanceIndicatorSchema)
def create_indicator(request, payload: PerformanceIndicatorCreateSchema):
    return create(request, payload, PerformanceIndicator)

@router.delete("/indicators/{id}")
def delete_indicator(request, id: str):
    instance = get_object_or_404(PerformanceIndicator, id=id)
    # 权限检查：只有责任人或超级管理员可以删除
    user_id = request.auth.id
    is_superuser = getattr(request.auth, 'is_superuser', False)
    
    if str(instance.owner_id) != str(user_id) and not is_superuser:
        raise HttpError(403, "只有责任人才能删除该指标")
        
    delete(id, PerformanceIndicator)
    return {"success": True}

@router.put("/indicators/{id}", response=PerformanceIndicatorSchema)
def update_indicator(request, id: str, payload: PerformanceIndicatorUpdateSchema):
    instance = get_object_or_404(PerformanceIndicator, id=id)
    # 权限检查：只有责任人或超级管理员可以编辑
    user_id = request.auth.id
    is_superuser = getattr(request.auth, 'is_superuser', False)
    
    if str(instance.owner_id) != str(user_id) and not is_superuser:
        raise HttpError(403, "只有责任人才能编辑该指标")

    return update(request, id, payload, PerformanceIndicator)

@router.get("/indicators", response=List[PerformanceIndicatorSchema])
@paginate(MyPagination)
def list_indicators(request, search: str = None, module: str = None, chip_type: str = None, project: str = None):
    qs = PerformanceIndicator.objects.select_related('owner').all()
    if search:
        qs = qs.filter(name__icontains=search)
    if module:
        qs = qs.filter(module__icontains=module)
    if chip_type:
        qs = qs.filter(chip_type__icontains=chip_type)
    if project:
        qs = qs.filter(project__icontains=project)
    return qs

# --- Data Upload & Trend ---

@router.post("/data/upload", response=PerformanceDataUploadResponse)
def upload_data(request, payload: PerformanceDataUploadSchema):
    result = upload_performance_data(payload.dict())
    return result

@router.get("/data/trend", response=List[PerformanceDataTrendSchema])
def get_trend(request, indicator_id: str, days: int = 7):
    from datetime import timedelta, date
    end_date = date.today()
    start_date = end_date - timedelta(days=days)
    
    qs = PerformanceIndicatorData.objects.filter(
        indicator_id=indicator_id,
        date__gte=start_date,
        date__lte=end_date
    ).order_by('date')
    
    return list(qs)

# --- Dashboard Data ---

@router.get("/dashboard", response=List[dict]) # Simplified schema for dashboard table
@paginate(MyPagination)
def dashboard_data(request, project: str = None, module: str = None, chip_type: str = None, date: str = None):
    """
    Dashboard API v2
    Focuses on actual data records (PerformanceIndicatorData) rather than indicators.
    Returns indicators only if they have data matching the filters.
    """
    
    # Start with Data query
    qs = PerformanceIndicatorData.objects.select_related('indicator', 'indicator__owner').all()
    
    # Filter by Date (default to today if not specified?)
    # Requirement: "In this page can search... by date"
    # If date is not provided, we should probably return latest data or all data?
    # Usually dashboard shows a snapshot. Let's assume if date is provided, filter by it.
    # If not, maybe show latest? Or just all records paginated?
    # The user said "see the data uploaded every day... search by date"
    # So we list Data records.
    
    from datetime import date as dt_date
    
    if date:
        try:
            target_date = dt_date.fromisoformat(str(date))
            qs = qs.filter(date=target_date)
        except ValueError:
            pass
    
    # Filter by Indicator properties (Project, Module, Chip Type)
    if project:
        qs = qs.filter(indicator__project__icontains=project)
    if module:
        qs = qs.filter(indicator__module__icontains=module)
    if chip_type:
        qs = qs.filter(indicator__chip_type__icontains=chip_type)
        
    # Order by date desc, then indicator code
    qs = qs.order_by('-date', 'indicator__code')
    
    # Transform to result format
    # Since we are paginating the QuerySet directly via Ninja, we just need to return the QuerySet
    # But we need to shape the output to match what frontend expects or update frontend.
    # The current frontend expects: name, baseline_value, current_value, fluctuation_value, etc.
    # Let's map the Data objects to a dict structure compatible with the frontend, 
    # but we can't use simple paginate decorator on a list generator if we want efficient DB pagination.
    # We should probably define a Schema or return a list of dicts.
    # Given the previous implementation returned a list of dicts manually, we can do the same here but
    # apply pagination on the QS first if we were not using the decorator on the view function.
    # However, ninja's @paginate handles QS efficiently.
    # We need to return a list of objects that look like the dashboard items.
    
    # We can use a Schema for response, but for now let's stick to dict to match previous style
    # but we need to iterate over the paginated page.
    # Wait, @paginate on the function means the function should return the full QS or List,
    # and Ninja slices it.
    
    # If we return QS, Ninja tries to serialize it using the response_schema. 
    # The response schema is List[dict].
    # So we need to map the QS objects to dicts.
    # BUT, if we return a QS, Ninja will serialize model instances.
    # We need to customize the serialization or return a list of dicts (which loads all into memory).
    # For better performance with @paginate, we usually return QS.
    # Let's manually construct the list of dicts for the *entire* filtered result? 
    # No, that's bad for large datasets.
    
    # Better approach:
    # Use `values()` or `annotate()` to shape the data?
    # Or just loop over the sliced result? 
    # Ninja's pagination happens *after* the function returns.
    # So if we return a list, we load everything.
    # If we return a QS, Ninja slices it.
    
    # To support custom field mapping with efficient pagination:
    # We can return the QS, and use a Schema with `resolve_xxx` methods or properties.
    # OR, we can implement manual pagination here and return a list of dicts for just the page.
    # Let's stick to returning a list of dicts for now, assuming the dataset isn't huge yet,
    # OR better: use a Schema that maps from PerformanceIndicatorData.
    
    # Let's define the logic to return list of dicts to be safe with existing frontend,
    # but strictly filtering for data existence.
    
    # Ideally:
    # 1. Filter PerformanceIndicatorData
    # 2. Return that QuerySet
    # 3. Use a Schema to format the output
    
    # But since I can't easily add a new Schema file right now without more tool calls,
    # I will construct the list. If data grows large, we should switch to Schema-based serialization.
    
    results = []
    for data in qs:
        indicator = data.indicator
        row = {
            "id": indicator.id,
            "name": indicator.name,
            "code": indicator.code,
            "project": indicator.project,
            "module": indicator.module,
            "chip_type": indicator.chip_type,
            "baseline_value": indicator.baseline_value,
            "baseline_unit": indicator.baseline_unit,
            "fluctuation_range": indicator.fluctuation_range,
            "fluctuation_direction": indicator.fluctuation_direction,
            
            # Data specific fields
            "current_value": data.value,
            "fluctuation_value": data.fluctuation_value,
            "data_date": data.date,
            "owner_name": indicator.owner.name if indicator.owner else None
        }
        results.append(row)
        
    return results
